import openpyxl, sys

wb = openpyxl.load_workbook(r"C:\Users\HUAWEI\Downloads\план работ.xlsx", data_only=True)
out = []
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    out.append(f"=== Лист: {sheet_name} ({ws.max_row} строк) ===")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if any(v is not None for v in row):
            name = str(row[0] or "").replace("_x000D_", "").strip()
            start = str(row[1] or "").strip()
            end   = str(row[2] or "").strip()
            out.append(f"{name} | {start} | {end}")

text = "\n".join(out)
with open(r"C:\Users\HUAWEI\Downloads\plan_decoded.txt", "w", encoding="utf-8") as f:
    f.write(text)
print("OK")
