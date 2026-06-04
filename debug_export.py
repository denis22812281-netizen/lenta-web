"""Тест генерации Excel без сервера."""
import sys, os
sys.path.insert(0, r"c:\Users\HUAWEI\моя\lenta-web")
os.chdir(r"c:\Users\HUAWEI\моя\lenta-web")

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"

    hfill = PatternFill(start_color="1A5C22", end_color="1A5C22", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True, size=10)
    msfill = PatternFill(start_color="3D3000", end_color="3D3000", fill_type="solid")
    msfont = Font(color="FFD200", bold=True, size=10)
    done_fill = PatternFill(start_color="0F2A12", end_color="0F2A12", fill_type="solid")
    thin = Side(style="thin", color="2D3748")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(vertical="center", wrap_text=True)
    row_font = Font(color="E2E8F0", size=10)

    # Test PatternFill() with no args
    empty_fill = PatternFill()
    print("PatternFill() OK")

    # Test writing cells
    c = ws.cell(1, 1, "Test")
    c.fill = hfill
    c.font = hfont
    c.border = brd

    # Test empty fill on cell
    c2 = ws.cell(2, 1, "Row")
    c2.fill = empty_fill
    c2.font = row_font
    c2.border = brd

    # Test done_fill
    c3 = ws.cell(3, 1, "Done")
    c3.fill = done_fill
    c3.font = row_font

    # Test save
    out = r"C:\Users\HUAWEI\Downloads\test_smr.xlsx"
    wb.save(out)
    print(f"Сохранено: {out}")

except Exception as e:
    import traceback
    print(f"ОШИБКА: {e}")
    traceback.print_exc()
