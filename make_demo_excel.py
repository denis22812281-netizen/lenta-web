"""Генерирует демо-файл Констракшн совместимый с import_construction_excel."""
import sys
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "2026"

# ── Позиции колонок (должны совпадать и в заголовке и в данных) ────────────────
COL = {
    "n":         1,
    "tk":        2,   # "номер тк"
    "addr":      3,   # "адрес"
    "city":      4,   # "регион"
    "fmt":       5,   # "тип формата"
    "area":      6,   # "площадь тз"
    "reception": 14,  # "приёмка"
    "cmp":       16,  # "выход на смр"
    "vpk":       19,  # "впк 1"
    "open_plan": 20,  # "первоначальная"
    "open_fact": 21,  # "фактическая"
    "status":    22,  # "статус открытия"
    "mgr":       38,  # "менеджер ос"  ← заголовок тоже будет в col 38
}

# ── Стили ──────────────────────────────────────────────────────────────────────
hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
hdr_font = Font(color="FFFFFF", bold=True, size=10)
row_fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
done_fill= PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
thin     = Side(style="thin", color="BFBFBF")
brd      = Border(left=thin, right=thin, top=thin, bottom=thin)
center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
left     = Alignment(vertical="center", wrap_text=True)

# ── Строка заголовков (строка 3) ───────────────────────────────────────────────
HEADERS = {
    COL["n"]:         "No.",
    COL["tk"]:        "Номер ТК",
    COL["addr"]:      "Адрес объекта",
    COL["city"]:      "Регион",
    COL["fmt"]:       "Тип формата",
    COL["area"]:      "Площадь ТЗ, кв.м",
    COL["reception"]: "Приёмка готовности",
    COL["cmp"]:       "Выход на СМР",
    COL["vpk"]:       "ВПК 1",
    COL["open_plan"]: "Первоначальная дата открытия",
    COL["open_fact"]: "Фактическая дата открытия",
    COL["status"]:    "Статус открытия",
    COL["mgr"]:       "Менеджер ОС",
}

max_col = max(HEADERS.keys())
for c in range(1, max_col + 1):
    cell = ws.cell(3, c, HEADERS.get(c, ""))
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.alignment = center
    cell.border = brd

ws.row_dimensions[3].height = 32
ws.freeze_panes = "A4"

# ── Данные ─────────────────────────────────────────────────────────────────────
managers = [
    "Месмер Денис",
    "Митько Роберт",
    "Ловчиков Александр",
    "Валеев Борис",
    "Косило Сергей",
    "Студеникин Сергей",
]

cities_addrs = [
    ("Москва",           "ул. Центральная, 1"),
    ("Москва",           "ул. Ленина, 45"),
    ("Санкт-Петербург",  "Московский пр-т, 120"),
    ("Казань",           "ул. Кирова, 78"),
    ("Екатеринбург",     "ул. Советская, 33"),
    ("Новосибирск",      "пр-т Мира, 55"),
    ("Краснодар",        "ул. Гагарина, 12"),
    ("Нижний Новгород",  "ул. Победы, 8"),
    ("Самара",           "Северный пр-т, 200"),
    ("Уфа",              "ул. Садовая, 90"),
    ("Ростов-на-Дону",   "ул. Новая, 17"),
    ("Воронеж",          "пр-т Строителей, 44"),
    ("Пермь",            "ул. Молодёжная, 3"),
    ("Волгоград",        "ул. Заречная, 67"),
    ("Тюмень",           "пр-т Энергетиков, 88"),
]

today = date.today()
base  = date(2026, 1, 15)

for i in range(15):
    tk      = 9001 + i
    city, street = cities_addrs[i]
    addr    = f"{city}, {street}"
    mgr     = managers[i % len(managers)]
    shift   = timedelta(days=i * 14)

    reception = base + shift
    cmp_d     = reception  + timedelta(days=7)
    vpk_d     = cmp_d      + timedelta(days=90)
    open_plan = vpk_d      + timedelta(days=14)
    delta     = [-7, 0, 7][i % 3]               # раньше / вовремя / позже
    open_fact = open_plan + timedelta(days=delta)
    is_done   = open_fact <= today
    st        = ("Открыт раньше срока" if delta < 0
                 else "Открыт вовремя" if delta == 0
                 else "Открыт с опозданием") if is_done else ""

    row = 4 + i
    fill = done_fill if is_done else row_fill

    def cell(col_key, val, fmt=None):
        c = ws.cell(row, COL[col_key], val)
        c.fill = fill
        c.border = brd
        if fmt:
            c.number_format = fmt
            c.alignment = center
        elif col_key in ("n", "tk", "area", "fmt"):
            c.alignment = center
        else:
            c.alignment = left

    cell("n",         i + 1)
    cell("tk",        tk)
    cell("addr",      addr)
    cell("city",      city)
    cell("fmt",       "SM")
    cell("area",      3500 + i * 200)
    cell("reception", reception, "DD.MM.YYYY")
    cell("cmp",       cmp_d,     "DD.MM.YYYY")
    cell("vpk",       vpk_d,     "DD.MM.YYYY")
    cell("open_plan", open_plan, "DD.MM.YYYY")
    cell("open_fact", open_fact if is_done else None, "DD.MM.YYYY")
    cell("status",    st)
    cell("mgr",       mgr)

    ws.row_dimensions[row].height = 18

# ── Ширина колонок ─────────────────────────────────────────────────────────────
ws.column_dimensions[get_column_letter(COL["n"])        ].width = 5
ws.column_dimensions[get_column_letter(COL["tk"])       ].width = 10
ws.column_dimensions[get_column_letter(COL["addr"])     ].width = 38
ws.column_dimensions[get_column_letter(COL["city"])     ].width = 18
ws.column_dimensions[get_column_letter(COL["fmt"])      ].width = 10
ws.column_dimensions[get_column_letter(COL["area"])     ].width = 14
ws.column_dimensions[get_column_letter(COL["reception"])].width = 16
ws.column_dimensions[get_column_letter(COL["cmp"])      ].width = 14
ws.column_dimensions[get_column_letter(COL["vpk"])      ].width = 12
ws.column_dimensions[get_column_letter(COL["open_plan"])].width = 20
ws.column_dimensions[get_column_letter(COL["open_fact"])].width = 20
ws.column_dimensions[get_column_letter(COL["status"])   ].width = 22
ws.column_dimensions[get_column_letter(COL["mgr"])      ].width = 18

# ── Сохранение ─────────────────────────────────────────────────────────────────
out = r"C:\Users\HUAWEI\Downloads\Констракшн_DEMO_2026.xlsx"
wb.save(out)

# Верификация
wb2 = Workbook()
import openpyxl as _xl
wb2 = _xl.load_workbook(out, data_only=True)
ws2 = wb2["2026"]
print("=== ВЕРИФИКАЦИЯ ===")
print(f"Лист: {ws2.title}")
print(f"Заголовок col {COL['tk']}: {ws2.cell(3, COL['tk']).value!r}")
print(f"Заголовок col {COL['mgr']}: {ws2.cell(3, COL['mgr']).value!r}")
print(f"Заголовок col {COL['fmt']}: {ws2.cell(3, COL['fmt']).value!r}")
print()
print("Первые 5 строк данных:")
for r in range(4, 9):
    tk  = ws2.cell(r, COL["tk"]).value
    fmt = ws2.cell(r, COL["fmt"]).value
    mgr = ws2.cell(r, COL["mgr"]).value
    pln = ws2.cell(r, COL["open_plan"]).value
    print(f"  row {r}: tk={tk!r:6}  fmt={fmt!r:5}  mgr={mgr!r:22}  open_plan={pln!r}")
print()
print(f"Файл: {out}")
print("OК — готов к загрузке")
