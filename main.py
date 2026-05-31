from fastapi import FastAPI, Request, Form, UploadFile, File, Depends, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
from sqlalchemy.orm import Session
from datetime import datetime, date, timedelta
import models
import database
from database import get_db
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import hashlib
import io
import os
import secrets
import asyncio
from pathlib import Path

app = FastAPI(title="Лента — Управление проектами")

SECRET_KEY = os.getenv("SECRET_KEY", secrets.token_hex(32))
app.add_middleware(SessionMiddleware, secret_key=SECRET_KEY, max_age=86400 * 7)

app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

MANAGERS_SEED = [
    ("Гаврин Игорь", True),       # Руководитель проектов
    ("Месмер Денис", False),
    ("Митько Роберт", False),
    ("Ловчиков Александр", False),
    ("Шевченко Наталья", False),
    ("Хачатурова Жанна", False),
    ("Валеев Борис", False),
    ("Студеникин Сергей", False),
    ("Косило Сергей", False),
]

PROJECT_TYPES = ["Реконструкция", "Констракшн", "КСО", "Новое строительство", "Капитальный ремонт", "Техническое обслуживание"]
STATUSES = ["Активный", "Завершён", "Приостановлен", "Планирование"]
PRIORITIES = ["Высокий", "Средний", "Низкий"]
TASK_STATUSES = ["Открытая", "В работе", "На проверке", "Завершена"]
STAGE_NAMES = [
    "Подготовительный этап",
    "Демонтажные работы",
    "Фундаментные работы",
    "Конструктивные элементы",
    "Инженерные системы",
    "Чистовая отделка",
    "Благоустройство",
    "Сдача объекта",
]


# Онлайн-пользователи: display_name -> datetime последнего пинга
ONLINE_USERS: dict = {}
ONLINE_TIMEOUT = 120  # секунд без пинга = оффлайн


def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()


def normalize_phone(phone: str) -> str:
    """Normalize to +7XXXXXXXXXX format."""
    digits = ''.join(c for c in phone if c.isdigit())
    if len(digits) == 11 and digits.startswith(('7', '8')):
        return '+7' + digits[1:]
    if len(digits) == 10:
        return '+7' + digits
    if len(digits) == 12 and digits.startswith('7'):
        return '+' + digits
    return phone.strip()


def get_current_user(request: Request):
    return request.session.get("user")


def require_login(request: Request):
    user = request.session.get("user")
    if not user:
        raise HTTPException(status_code=302, headers={"Location": "/login"})
    return user


def _match_manager(name_str: str, managers: list) -> int | None:
    """Match name like 'Месмер Д.', 'Месмер', 'МЕСМЕР Д' to manager by last name."""
    if not name_str or not isinstance(name_str, str):
        return None
    name_clean = name_str.strip()
    # Extract last name (first word, strip dots and spaces)
    last_name = name_clean.split()[0].strip('.')
    if not last_name or len(last_name) < 3:
        return None
    last_lower = last_name.lower()
    for m in managers:
        m_last = m.name.split()[0].lower()
        if m_last == last_lower or m_last.startswith(last_lower) or last_lower.startswith(m_last):
            return m.id
    return None


def _safe_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    return None


def _row_to_dict(row) -> dict:
    return {cell.column: cell.value for cell in row if cell.value is not None}


def import_reconstruct_excel(content: bytes, db: Session) -> dict:
    """
    Import Лента Реконструкции Excel.
    Reads 4 target sheets:
      9  - Реконструкции 2026   (deadline=col41, mgr=col51)
      10 - Лайт реконструкции  (deadline=col37, mgr=col41)
      11 - Рисковые АЛ         (deadline=col8 or last date, mgr=col40)
      12 - Рисковые МП         (deadline=col8 or last date, mgr=col40)
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    managers = db.query(models.Manager).all()
    today = date.today()
    created = updated = 0

    # Sheet configs: (wb_index, data_start_row, col_tk, col_region, col_address,
    #                 col_type, col_area, col_deadline, col_manager, col_manager2, label)
    # Manager columns corrected by reading actual Excel structure
    sheet_configs = [
        # idx, start, tk, region, address, type, area, deadline, mgr,  mgr2, label
        (9,  4, 1, 3, 5, 9, 12, 41, 52, 51, "Реконструкция"),   # col52=our mgr, col51=supplier
        (10, 4, 1, 3, 5, 9, 12, 37, 42, 41, "Реконструкция"),   # col42=our mgr
        (11, 3, 1, 2, 3, 6,  7,  8, 41, 40, "Реконструкция"),   # col41 or col40
        (12, 3, 1, 2, 3, 5,  7,  8, 37, 40, "Реконструкция"),   # col37=our mgr
    ]

    # Stage columns per sheet type (pairs: start, end; or single date)
    stage_map_main = {
        "СИД": (14, 15),          # Сбор исходных данных
        "Зонирование": (16, 17),
        "Закрытие": (18, 19),
        "ВПК1": (34, 35),
        "Открытие": (40, 41),
    }
    stage_map_risk = {
        "СИД": (9, 10),
        "Зонирование": (11, 12),
        "Закрытие": (13, 14),
        "ВПК1": (15, 16),
        "Открытие": (8, None),
    }

    for cfg in sheet_configs:
        idx, start_row, c_tk, c_region, c_addr, c_type, c_area, c_deadline, c_mgr, c_mgr2, p_type = cfg
        if idx >= len(wb.worksheets):
            continue
        ws = wb.worksheets[idx]
        is_risk = idx in (11, 12)
        stage_map = stage_map_risk if is_risk else stage_map_main

        for row_idx in range(start_row, ws.max_row + 1):
            row = ws[row_idx]
            vals = _row_to_dict(row)

            tk = vals.get(c_tk)
            if not tk or not isinstance(tk, (int, float)):
                continue
            tk_num = str(int(tk))

            region = vals.get(c_region, "")
            full_address = str(vals.get(c_addr, "") or "").strip()
            rec_type = vals.get(c_type, "")
            area = vals.get(c_area)
            area = float(area) if isinstance(area, (int, float)) else None

            # City: use region group; address: save full address from col
            city = ""
            if region and isinstance(region, str) and region not in ("#REF!", "#N/A"):
                city = region
            elif full_address:
                city = full_address.split(",")[0].strip()

            # Deadline
            dl_val = vals.get(c_deadline)
            deadline = _safe_date(dl_val)
            if deadline is None:
                # Fallback: last valid date in row
                dates = [_safe_date(v) for v in vals.values() if _safe_date(v)]
                deadline = max(dates) if dates else None

            # Manager: try primary column, then fallback column
            mgr_val = vals.get(c_mgr, "")
            manager_id = _match_manager(str(mgr_val) if mgr_val else "", managers)
            if not manager_id:
                mgr_val2 = vals.get(c_mgr2, "")
                manager_id = _match_manager(str(mgr_val2) if mgr_val2 else "", managers)

            # Stage-specific dates
            sid_s = _safe_date(vals.get(stage_map["СИД"][0]))
            sid_e = _safe_date(vals.get(stage_map["СИД"][1]))
            zon_s = _safe_date(vals.get(stage_map["Зонирование"][0]))
            zon_e = _safe_date(vals.get(stage_map["Зонирование"][1]))
            clos  = _safe_date(vals.get(stage_map["Закрытие"][0]))
            vpk   = _safe_date(vals.get(stage_map["ВПК1"][0]))
            open_start = stage_map["Открытие"][0]
            open_end   = stage_map["Открытие"][1]
            opening = _safe_date(vals.get(open_end if open_end else open_start))

            # Project start = first stage start
            proj_start = sid_s

            # Build project name
            proj_name = f"ТК {tk_num}"
            if city:
                proj_name += f" {city}"

            # Auto-status based on opening date
            if opening and opening <= today:
                status = "Завершён"
            else:
                status = "Активный"

            # Find or create
            existing = db.query(models.Project).filter(
                models.Project.tk_number == tk_num).first()

            if existing:
                existing.city = city or existing.city
                existing.address = full_address or existing.address
                existing.project_type = p_type
                existing.start_date = proj_start or existing.start_date
                existing.end_date = deadline or existing.end_date
                existing.area = area or existing.area
                existing.sid_start = sid_s or existing.sid_start
                existing.sid_end = sid_e or existing.sid_end
                existing.zoning_start = zon_s or existing.zoning_start
                existing.zoning_end = zon_e or existing.zoning_end
                existing.closure_date = clos or existing.closure_date
                existing.vpk_date = vpk or existing.vpk_date
                existing.opening_date = opening or existing.opening_date
                if existing.status != "Приостановлен":
                    existing.status = status
                if manager_id and not existing.manager_id:
                    existing.manager_id = manager_id
                updated += 1
            else:
                p = models.Project(
                    name=proj_name, tk_number=tk_num, city=city,
                    address=full_address,
                    project_type=p_type, manager_id=manager_id, status=status,
                    start_date=proj_start, end_date=deadline, area=area,
                    sid_start=sid_s, sid_end=sid_e,
                    zoning_start=zon_s, zoning_end=zon_e,
                    closure_date=clos, vpk_date=vpk, opening_date=opening,
                )
                db.add(p)
                created += 1

        db.flush()

    db.commit()
    return {"created": created, "updated": updated}


def parse_excel_file(content: bytes, project_type: str, manager_id_default: int,
                      db: Session) -> dict:
    """Parse Excel, update existing projects by TK number, create new ones."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    created = 0
    updated = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        project_name = sheet_name
        tk_num = ""
        city_name = ""

        for row in ws.iter_rows(max_row=10, values_only=True):
            for cell in row:
                if cell and isinstance(cell, str):
                    text = cell.strip()
                    if text.upper().startswith("ТК") or text.upper().startswith("TK"):
                        project_name = text
                        parts = text.split()
                        if len(parts) >= 2:
                            tk_num = parts[1]
                        if len(parts) >= 3:
                            city_name = " ".join(parts[2:])

        # Find dates in the sheet (look for start/end dates of whole project)
        all_dates = []
        for row in ws.iter_rows(max_row=100, values_only=True):
            for cell in row:
                if isinstance(cell, datetime):
                    all_dates.append(cell.date())
                elif isinstance(cell, date) and not isinstance(cell, datetime):
                    all_dates.append(cell)

        proj_start = min(all_dates) if all_dates else None
        proj_end = max(all_dates) if all_dates else None

        # Try to find existing project by TK number
        existing = None
        if tk_num:
            existing = db.query(models.Project).filter(
                models.Project.tk_number == tk_num).first()
        if not existing and project_name:
            existing = db.query(models.Project).filter(
                models.Project.name == project_name).first()

        if existing:
            # Update dates only (don't overwrite other fields)
            if proj_start:
                existing.start_date = proj_start
            if proj_end:
                existing.end_date = proj_end
            if city_name and not existing.city:
                existing.city = city_name
            if project_type and not existing.project_type:
                existing.project_type = project_type
            db.flush()
            _sync_stages(ws, existing.id, db)
            updated += 1
        else:
            p = models.Project(
                name=project_name, tk_number=tk_num, city=city_name,
                project_type=project_type,
                manager_id=manager_id_default if manager_id_default else None,
                status="Активный", start_date=proj_start, end_date=proj_end,
            )
            db.add(p)
            db.flush()
            _sync_stages(ws, p.id, db)
            created += 1

    db.commit()
    return {"created": created, "updated": updated}


def _sync_stages(ws, project_id: int, db: Session):
    existing_names = {s.name for s in db.query(models.ProjectStage).filter(
        models.ProjectStage.project_id == project_id).all()}
    order = db.query(models.ProjectStage).filter(
        models.ProjectStage.project_id == project_id).count()

    for row in ws.iter_rows(values_only=True):
        if not row[0] or not isinstance(row[0], str):
            continue
        cell_text = row[0].strip()
        for sn in STAGE_NAMES:
            if sn.split()[0].lower() in cell_text.lower():
                start = end = None
                for v in row[1:]:
                    if isinstance(v, datetime):
                        if not start:
                            start = v.date()
                        elif not end:
                            end = v.date()
                            break
                    elif isinstance(v, date) and not isinstance(v, datetime):
                        if not start:
                            start = v
                        elif not end:
                            end = v
                            break
                if cell_text not in existing_names:
                    db.add(models.ProjectStage(
                        project_id=project_id, name=cell_text,
                        start_date=start, end_date=end, order=order))
                    order += 1
                else:
                    # Update existing stage dates
                    s = db.query(models.ProjectStage).filter(
                        models.ProjectStage.project_id == project_id,
                        models.ProjectStage.name == cell_text).first()
                    if s:
                        if start:
                            s.start_date = start
                        if end:
                            s.end_date = end
                break


def import_construction_excel(content: bytes, db: Session) -> dict:
    """Import Лента Констракшн Excel — листы 2026 и 2027."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    managers = db.query(models.Manager).all()
    today = date.today()
    created = updated = 0

    # Собираем листы 2026 и 2027 (гибкий поиск)
    target_sheets = [ws for ws in wb.worksheets
                     if ws.title.strip() in ('2026', '2027')
                     or '2026' in ws.title or '2027' in ws.title]
    # Убираем дубли если есть
    seen = set()
    target_sheets = [ws for ws in target_sheets
                     if ws.title not in seen and not seen.add(ws.title)]
    if not target_sheets:
        target_sheets = [wb.worksheets[-1]]

    skipped_fmt = 0
    skipped_mgr = 0
    sample_formats = set()
    sample_managers = set()
    rows_with_tk = 0

    for ws in target_sheets:
        # Найти строку заголовков — правильный break из обоих циклов
        header_row = 4
        found_hdr = False
        for r in range(1, 11):
            if found_hdr:
                break
            for c in range(1, 15):
                v = str(ws.cell(r, c).value or '')
                if 'Номер' in v or 'Адрес' in v:
                    header_row = r
                    found_hdr = True
                    break

        # Определяем позиции колонок — только в строке заголовка (не в баннерах сверху)
        col = {}
        for r in range(max(1, header_row - 1), header_row + 2):
            for c in range(1, 60):
                v = str(ws.cell(r, c).value or '').strip()
                vl = v.lower()
                if not v:
                    continue
                if ('номер тк' in vl or vl == 'тк') and 'tk' not in col:
                    col['tk'] = c
                elif 'адрес' in vl and 'addr' not in col and len(v) < 30:
                    col['addr'] = c
                elif 'тип формата' in vl and 'fmt' not in col:
                    col['fmt'] = c
                elif ('формат' in vl and len(v) < 20) and 'fmt' not in col:
                    col['fmt'] = c
                elif ('регион' in vl or 'город' in vl) and 'city' not in col:
                    col['city'] = c
                elif 'площадь' in vl and 'тз' in vl and 'area' not in col:
                    col['area'] = c
                elif ('приёмка' in vl or 'приемка' in vl) and 'reception' not in col:
                    col['reception'] = c
                elif 'выход на смр' in vl and 'cmp' not in col:
                    col['cmp'] = c
                elif vl.startswith('впк') and 'vpk' not in col:
                    col['vpk'] = c
                elif 'первоначальная' in vl and 'open_plan' not in col:
                    col['open_plan'] = c
                elif 'фактическая' in vl and 'open_fact' not in col:
                    col['open_fact'] = c
                elif 'статус' in vl and ('откр' in vl or 'open' in vl) and 'open_status' not in col:
                    col['open_status'] = c
                elif 'менеджер' in vl and 'ос' in vl and 'mgr_os' not in col:
                    col['mgr_os'] = c

        fmt_from_header = 'fmt' in col  # флаг: найден ли формат через заголовок

        # Жёсткие fallback по реальной структуре файла
        col.setdefault('tk', 2)           # B: Номер ТК
        col.setdefault('addr', 3)         # C: Адрес
        col.setdefault('reception', 14)   # N: Приёмка помещения
        col.setdefault('cmp', 16)         # P: Выход на СМР
        col.setdefault('vpk', 19)         # S: ВПК 1
        col.setdefault('open_plan', 20)   # T: Первоначальная дата открытия
        col.setdefault('open_fact', 21)   # U: Фактическая дата открытия
        col.setdefault('open_status', 22) # V: Статус открытия
        col.setdefault('mgr_os', 38)      # AL: Менеджер проектов ОС

        data_start = header_row + 1

        for row_idx in range(data_start, ws.max_row + 1):
            tk_val = ws.cell(row_idx, col['tk']).value
            if not tk_val:
                continue
            tk_num = str(tk_val).strip()
            if not tk_num or tk_num in ('None', 'Номер ТК', 'ТК'):
                continue
            if len(tk_num) > 20:
                continue

            rows_with_tk += 1
            address = str(ws.cell(row_idx, col['addr']).value or '').strip()
            fmt_type = str(ws.cell(row_idx, col['fmt']).value or '').strip()
            if fmt_type and len(sample_formats) < 8:
                sample_formats.add(fmt_type)
            mgr_raw = str(ws.cell(row_idx, col['mgr_os']).value or '').strip()
            if mgr_raw and len(sample_managers) < 8:
                sample_managers.add(mgr_raw)
            city = str(ws.cell(row_idx, col.get('city', 0)).value or '').strip() if col.get('city') else ''
            area_val = ws.cell(row_idx, col.get('area', 0)).value if col.get('area') else None
            area = float(area_val) if isinstance(area_val, (int, float)) else None

            reception = _safe_date(ws.cell(row_idx, col['reception']).value)
            cmp_date  = _safe_date(ws.cell(row_idx, col['cmp']).value)
            vpk_date  = _safe_date(ws.cell(row_idx, col['vpk']).value)
            open_plan = _safe_date(ws.cell(row_idx, col['open_plan']).value)
            open_fact = _safe_date(ws.cell(row_idx, col['open_fact']).value)
            open_st   = str(ws.cell(row_idx, col['open_status']).value or '').strip()

            # Менеджер проектов ОС (колонка AL)
            mgr_val = mgr_raw
            manager_id = _match_manager(mgr_val, managers)

            # Фильтр по формату только если колонка найдена через заголовок
            fmt_upper = fmt_type.upper()
            if fmt_from_header and fmt_type and fmt_upper not in ('SM', 'UTKONOS'):
                skipped_fmt += 1
                continue

            # Только проекты наших менеджеров
            if not manager_id:
                skipped_mgr += 1
                continue

            opening = open_fact or open_plan
            status = "Завершён" if (opening and opening <= today) else "Активный"

            proj_name = f"ТК {tk_num}"
            if city:
                proj_name += f" {city}"

            existing = db.query(models.Project).filter(
                models.Project.tk_number == tk_num,
                models.Project.project_type == "Констракшн"
            ).first()
            if existing:
                existing.address = address or existing.address
                existing.city = city or existing.city
                existing.area = area or existing.area
                existing.format_type = fmt_type or existing.format_type
                existing.open_status = open_st or existing.open_status
                existing.start_date = reception or existing.start_date
                existing.closure_date = cmp_date or existing.closure_date
                existing.vpk_date = vpk_date or existing.vpk_date
                existing.end_date = open_plan or existing.end_date
                existing.opening_date = open_fact or existing.opening_date
                if existing.status != "Приостановлен":
                    existing.status = status
                existing.manager_id = manager_id
                updated += 1
            else:
                db.add(models.Project(
                    name=proj_name, tk_number=tk_num,
                    city=city, address=address,
                    project_type="Констракшн",
                    manager_id=manager_id, status=status,
                    format_type=fmt_type, open_status=open_st,
                    start_date=reception,
                    closure_date=cmp_date,
                    vpk_date=vpk_date,
                    end_date=open_plan,
                    opening_date=open_fact,
                    area=area,
                ))
                created += 1

    db.commit()
    sheets_info = ", ".join(ws.title for ws in target_sheets)
    return {
        "created": created,
        "updated": updated,
        "sheets": sheets_info,
        "rows_with_tk": rows_with_tk,
        "skipped_fmt": skipped_fmt,
        "skipped_mgr": skipped_mgr,
        "sample_formats": sorted(sample_formats),
        "sample_managers": sorted(sample_managers),
    }


async def auto_sync_loop():
    """Background task: check sync configs and run file sync."""
    while True:
        await asyncio.sleep(60)  # check every minute
        try:
            db = database.SessionLocal()
            try:
                configs = db.query(models.SyncConfig).filter(
                    models.SyncConfig.auto_sync == True,
                    models.SyncConfig.file_path != "").all()
                for cfg in configs:
                    if not cfg.last_synced:
                        should_sync = True
                    else:
                        elapsed = (datetime.utcnow() - cfg.last_synced).total_seconds() / 60
                        should_sync = elapsed >= cfg.sync_interval_minutes

                    if should_sync:
                        path = Path(cfg.file_path)
                        if path.exists() and path.suffix in ('.xlsx', '.xls', '.xlsm'):
                            try:
                                content = path.read_bytes()
                                if cfg.project_type == "Реконструкция":
                                    result = import_reconstruct_excel(content, db)
                                elif cfg.project_type == "Констракшн":
                                    result = import_construction_excel(content, db)
                                else:
                                    result = parse_excel_file(content, cfg.project_type, None, db)
                                cfg.last_synced = datetime.utcnow()
                                cfg.last_status = f"OK: создано {result['created']}, обновлено {result['updated']}"
                            except Exception as e:
                                cfg.last_status = f"Ошибка: {str(e)[:100]}"
                        else:
                            cfg.last_status = f"Файл не найден: {cfg.file_path}"
                        db.commit()
            finally:
                db.close()
        except Exception as e:
            print(f"[auto_sync_loop] критическая ошибка: {e}")


@app.on_event("startup")
async def startup():
    models.Base.metadata.create_all(bind=database.engine)

    # Расширяем VARCHAR колонки до TEXT в PostgreSQL
    if "postgresql" in str(database.DATABASE_URL):
        try:
            with database.engine.begin() as conn:
                for sql in [
                    "ALTER TABLE projects ALTER COLUMN city TYPE TEXT",
                    "ALTER TABLE projects ALTER COLUMN stage TYPE TEXT",
                    "ALTER TABLE project_stages ALTER COLUMN name TYPE TEXT",
                    "ALTER TABLE projects ADD COLUMN IF NOT EXISTS format_type VARCHAR(50) DEFAULT ''",
                    "ALTER TABLE projects ADD COLUMN IF NOT EXISTS open_status VARCHAR(100) DEFAULT ''",
                    "ALTER TABLE projects ADD COLUMN IF NOT EXISTS delay_reason TEXT DEFAULT ''",
                    "ALTER TABLE projects ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP DEFAULT NOW()",
                    "ALTER TABLE managers ADD COLUMN IF NOT EXISTS photo VARCHAR(200) DEFAULT ''",
                    "ALTER TABLE managers ADD COLUMN IF NOT EXISTS position VARCHAR(150) DEFAULT ''",
                    "ALTER TABLE tasks ADD COLUMN IF NOT EXISTS completion_comment TEXT DEFAULT ''",
                    """CREATE TABLE IF NOT EXISTS task_notifications (
                        id SERIAL PRIMARY KEY,
                        recipient_name VARCHAR(100) NOT NULL,
                        task_id INTEGER REFERENCES tasks(id) ON DELETE CASCADE,
                        message TEXT NOT NULL,
                        is_read BOOLEAN DEFAULT FALSE,
                        created_at TIMESTAMP DEFAULT NOW()
                    )""",
                    "ALTER TABLE vpk_report_items ADD COLUMN IF NOT EXISTS comment TEXT DEFAULT ''",
                    "ALTER TABLE vpk_report_items ADD COLUMN IF NOT EXISTS photo_path VARCHAR(300) DEFAULT ''",
                    "ALTER TABLE chat_messages ADD COLUMN IF NOT EXISTS photo_path VARCHAR(300) DEFAULT ''",
                    "CREATE TABLE IF NOT EXISTS ai_chat_messages (id SERIAL PRIMARY KEY, user_name VARCHAR(100) NOT NULL, role VARCHAR(20) NOT NULL, text TEXT NOT NULL, provider VARCHAR(30) DEFAULT 'groq', created_at TIMESTAMP DEFAULT NOW())",
                ]:
                    try:
                        conn.exec_driver_sql(sql)
                    except Exception:
                        pass
        except Exception:
            pass
    db = database.SessionLocal()
    try:
        if db.query(models.Manager).count() == 0:
            for name, is_leader in MANAGERS_SEED:
                db.add(models.Manager(name=name, is_leader=is_leader))
            db.commit()

        # Seed whitelist
        admin_phone = os.getenv("ADMIN_PHONE", "+79997303914")
        if db.query(models.PhoneWhitelist).count() == 0:
            db.add(models.PhoneWhitelist(phone=admin_phone, display_name="Месмер Денис", is_admin=True))
            db.commit()

        # Добавляем новых пользователей если их ещё нет
        new_users = [
            ("+79150511700", "Гаврин Игорь",        True),
            ("+79112744420", "Ловчиков Александр",   False),
            ("+79236255705", "Митько Роберт",         False),
            ("+79261476125", "Шевченко Наталья",      False),
            ("+79091152729", "Валеев Борис",          False),
            ("+79265096687", "Хачатурова Жанна",      False),
            ("+79231290722", "Студеникин Сергей",     False),
            ("+79227450486", "Косило Сергей",          False),
            ("+79032014579", "Комаров Алексей",       False),
        ]
        for phone, name, is_admin in new_users:
            exists = db.query(models.PhoneWhitelist).filter(
                models.PhoneWhitelist.phone == phone).first()
            if not exists:
                db.add(models.PhoneWhitelist(phone=phone, display_name=name, is_admin=is_admin))
        db.commit()

        # Гаврин Игорь — руководитель проектов
        gavrin = db.query(models.Manager).filter(models.Manager.name == "Гаврин Игорь").first()
        if gavrin:
            gavrin.photo = "img/managers/gavrin.png"
            gavrin.position = "Руководитель проектов"

        # Месмер Денис — своя аватарка
        mesmer = db.query(models.Manager).filter(models.Manager.name == "Месмер Денис").first()
        if mesmer:
            mesmer.photo = "img/raccoon_mesmer.jpg"
        db.commit()

        # Комаров Алексей — директор, видит все проекты
        komarov = db.query(models.Manager).filter(models.Manager.name == "Комаров Алексей").first()
        if not komarov:
            db.add(models.Manager(name="Комаров Алексей", is_leader=True,
                                  photo="img/managers/komarov.png",
                                  position="Директор по эксплуатации и реконструкции"))
        else:
            komarov.photo = "img/managers/komarov.png"
            komarov.position = "Директор по эксплуатации и реконструкции"
        db.commit()

        # Seed VPK criteria
        if db.query(models.VpkCriterion).count() == 0:
            vpk1 = [
                "Документация на произведённые работы предоставлена в полном объёме в печатном виде, согласно технического задания.",
                "Температурный режим на объекте обеспечивается, согласно условий технического задания с учётом времени года.",
                "Подъёмное оборудование для перегрузки товара смонтировано и введено в эксплуатацию.",
                "Холодильное оборудование запущено и выведено в режим, функционирует без аварий более 24 часов.",
                "Строительство подъездных путей окончено, препятствий для подъезда к зоне разгрузки нет.",
                "Системы пожарной безопасности смонтированы, находятся в исправном и автоматическом режиме, готовы к проведению комплексных испытаний.",
                "Электроснабжение объекта осуществляется по постоянной схеме подключения.",
                "Объект обеспечен всеми необходимыми коммунальными ресурсами, согласно условий договора аренды.",
                "Периметр объекта замкнут, все двери/ворота установлены, исправны.",
                "Лотки и кабельные трассы системы видеонаблюдения смонтированы, проводятся пуско-наладочные работы.",
                "Объект обеспечен доступом в интернет (один канал), зона приёмки товара оборудована всем необходимым, препятствий для приёма товара нет.",
                "Основные строительно-монтажные работы закончены.",
                "Лотки и кабельные трассы системы охранной сигнализации смонтированы, проводятся пуско-наладочные работы.",
                "Охрана объекта обеспечена сотрудником ЧОП.",
                "ДДА с распределением зон эксплуатационной ответственности предоставлен.",
                "Согласование контейнерной площадки со стороны местной администрации предоставлено от АРДД.",
                "Укомплектованность собственным персоналом не ниже 60%.",
            ]
            vpk2 = [
                "Технологическое оборудование полностью смонтировано и запущено.",
                "Системы пожарной безопасности находятся в полностью исправном состоянии, все неисправности устранены или признаны инженером ПБ не значительными и не влияющими на общую работоспособность системы.",
                "Рекламная вывеска смонтирована, в ночное время светится.",
                "Всё кассовое оборудование запущено и исправно, все IT коммуникации смонтированы, исправны.",
                "Входная группа готова для открытия объекта, наружные работы и благоустройство завершены.",
                "Охранная сигнализация смонтирована, исправна.",
                "Система видеонаблюдения смонтирована, исправна.",
                "Уличное освещение объекта смонтировано, исправно.",
                "Маркетинговое оформление СМ полностью закончено.",
                "Строительство парковки для клиентов (при наличии) окончено, препятствий для размещения автомобилей нет.",
                "Укомплектованность собственным персоналом не ниже 70%.",
                "Объект обеспечен доступом в интернет (два канала передачи данных).",
                "В зоне кассовой линейки должен быть обеспечен устойчивый сигнал сотовой связи (мобильное приложение ЛЕНТА запускается, существует возможность оплаты по СБП с личного мобильного телефона покупателя).",
            ]
            for i, name in enumerate(vpk1):
                db.add(models.VpkCriterion(vpk_type=1, name=name, order=i))
            for i, name in enumerate(vpk2):
                db.add(models.VpkCriterion(vpk_type=2, name=name, order=i))
            db.commit()

        # Users are created on first login — no auto password here
    finally:
        db.close()

    asyncio.create_task(auto_sync_loop())


# ─── AUTH ────────────────────────────────────────────────────────────────────

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    if request.session.get("user"):
        return RedirectResponse("/", status_code=302)
    return templates.TemplateResponse("login.html", {"request": request})


@app.post("/login/check-phone")
async def check_phone(request: Request, db: Session = Depends(get_db),
                      phone: str = Form(...)):
    """Step 1: check if phone is in whitelist."""
    normalized = normalize_phone(phone)
    wl = db.query(models.PhoneWhitelist).filter(
        models.PhoneWhitelist.phone == normalized).first()

    if not wl:
        return templates.TemplateResponse("login.html", {
            "request": request,
            "error": "Доступ закрыт. Этот номер не авторизован.",
        })

    user = db.query(models.User).filter(models.User.phone == normalized).first()

    if user and user.password_hash:
        # Existing user → ask for password
        return templates.TemplateResponse("login.html", {
            "request": request,
            "step": "password",
            "phone": normalized,
            "display_name": user.display_name or wl.display_name,
        })
    else:
        # New user or no password yet → create password
        return templates.TemplateResponse("login.html", {
            "request": request,
            "step": "create_password",
            "phone": normalized,
            "display_name": wl.display_name,
        })


@app.post("/login/enter")
async def login_enter(request: Request, db: Session = Depends(get_db),
                      phone: str = Form(...), password: str = Form(...),
                      remember: str = Form("")):
    """Step 2a: login with existing password."""
    normalized = normalize_phone(phone)
    user = db.query(models.User).filter(models.User.phone == normalized).first()
    if not user or user.password_hash != hash_password(password):
        return templates.TemplateResponse("login.html", {
            "request": request,
            "step": "password",
            "phone": normalized,
            "display_name": user.display_name if user else "",
            "error": "Неверный пароль",
        })
    _set_session(request, user)
    return RedirectResponse("/", status_code=302)


@app.post("/login/create-password")
async def create_password(request: Request, db: Session = Depends(get_db),
                          phone: str = Form(...), password: str = Form(...),
                          password2: str = Form(...)):
    """Step 2b: set password for first-time user."""
    normalized = normalize_phone(phone)

    # Re-verify whitelist
    wl = db.query(models.PhoneWhitelist).filter(
        models.PhoneWhitelist.phone == normalized).first()
    if not wl:
        return RedirectResponse("/login", status_code=302)

    if len(password) < 6:
        return templates.TemplateResponse("login.html", {
            "request": request,
            "step": "create_password",
            "phone": normalized,
            "display_name": wl.display_name,
            "error": "Пароль должен быть не менее 6 символов",
        })
    if password != password2:
        return templates.TemplateResponse("login.html", {
            "request": request,
            "step": "create_password",
            "phone": normalized,
            "display_name": wl.display_name,
            "error": "Пароли не совпадают",
        })

    user = db.query(models.User).filter(models.User.phone == normalized).first()
    if user:
        user.password_hash = hash_password(password)
    else:
        user = models.User(
            phone=normalized,
            username=normalized,
            password_hash=hash_password(password),
            display_name=wl.display_name,
            is_admin=wl.is_admin,
        )
        db.add(user)
    db.commit()
    db.refresh(user)
    _set_session(request, user)
    return RedirectResponse("/", status_code=302)


def _set_session(request: Request, user: models.User):
    request.session["user"] = {
        "id": user.id,
        "username": user.username,
        "display_name": user.display_name,
        "is_admin": user.is_admin,
        "phone": user.phone,
    }


@app.get("/logout")
async def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/login", status_code=302)


# ─── DASHBOARD ───────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    today = date.today()
    total_projects = db.query(models.Project).count()
    active_projects = db.query(models.Project).filter(models.Project.status == "Активный").count()
    overdue_tasks = db.query(models.Task).filter(
        models.Task.deadline < today, models.Task.status != "Завершена").count()
    tasks_due_soon = db.query(models.Task).filter(
        models.Task.deadline >= today,
        models.Task.deadline <= today + timedelta(days=7),
        models.Task.status != "Завершена").count()
    projects_deadline_soon = db.query(models.Project).filter(
        models.Project.end_date >= today,
        models.Project.end_date <= today + timedelta(days=14),
        models.Project.status == "Активный",
        (models.Project.opening_date == None) | (models.Project.opening_date > today)
    ).order_by(models.Project.end_date).limit(6).all()
    recent_tasks = db.query(models.Task).order_by(models.Task.created_at.desc()).limit(6).all()
    return templates.TemplateResponse("index.html", {
        "request": request, "user": user,
        "total_projects": total_projects, "active_projects": active_projects,
        "overdue_tasks": overdue_tasks, "tasks_due_soon": tasks_due_soon,
        "projects_deadline_soon": projects_deadline_soon,
        "recent_tasks": recent_tasks, "today": today,
    })


# ─── SECTION VIEWS (Реконструкции / Констракшн / КСО) ───────────────────────

@app.get("/reconstruct", response_class=HTMLResponse)
async def reconstruct_view(request: Request, db: Session = Depends(get_db),
                           manager_id: str = None, status: str = None, search: str = None):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    q = db.query(models.Project).filter(models.Project.project_type == "Реконструкция")
    if manager_id and str(manager_id).isdigit():
        q = q.filter(models.Project.manager_id == int(manager_id))
    if status:
        q = q.filter(models.Project.status == status)
    if search:
        q = q.filter(models.Project.tk_number.contains(search))
    projects = q.order_by(models.Project.end_date.nullslast()).all()
    managers = db.query(models.Manager).all()
    return templates.TemplateResponse("section_projects.html", {
        "request": request, "user": user,
        "section_title": "Реконструкции",
        "section_icon": "bi-building-fill",
        "section_color": "red",
        "section_type": "Реконструкция",
        "section_url": "/reconstruct",
        "projects": projects, "managers": managers,
        "statuses": STATUSES, "today": date.today(),
        "filter_manager_id": manager_id, "filter_status": status, "search": search or "",
    })


@app.get("/construction", response_class=HTMLResponse)
async def construction_view(request: Request, db: Session = Depends(get_db),
                            manager_id: str = None, status: str = None, search: str = None):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    q = db.query(models.Project).filter(models.Project.project_type == "Констракшн")
    if manager_id and str(manager_id).isdigit():
        q = q.filter(models.Project.manager_id == int(manager_id))
    if status:
        q = q.filter(models.Project.status == status)
    if search:
        q = q.filter(models.Project.tk_number.contains(search))
    projects = q.order_by(models.Project.end_date.nullslast()).all()
    managers = db.query(models.Manager).all()
    return templates.TemplateResponse("section_projects.html", {
        "request": request, "user": user,
        "section_title": "Констракшн",
        "section_icon": "bi-buildings-fill",
        "section_color": "blue",
        "section_type": "Констракшн",
        "section_url": "/construction",
        "projects": projects, "managers": managers,
        "statuses": STATUSES, "today": date.today(),
        "filter_manager_id": manager_id, "filter_status": status, "search": search or "",
    })


@app.get("/kso", response_class=HTMLResponse)
async def kso_view(request: Request, db: Session = Depends(get_db),
                   manager_id: str = None, search: str = None, tab: str = "objects"):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)

    q = db.query(models.KsoObject)
    if manager_id and str(manager_id).isdigit():
        q = q.filter(models.KsoObject.manager_id == int(manager_id))
    if search:
        q = q.filter(models.KsoObject.tk_number.contains(search))
    objects = q.order_by(models.KsoObject.manager_id, models.KsoObject.tk_number).all()

    schedules = db.query(models.KsoSchedule).order_by(
        models.KsoSchedule.uploaded_at.desc()).all()
    managers = db.query(models.Manager).all()

    done_count = sum(1 for o in objects if o.done)
    total = len(objects)

    return templates.TemplateResponse("kso.html", {
        "request": request, "user": user,
        "objects": objects, "schedules": schedules,
        "managers": managers, "tab": tab,
        "filter_manager_id": manager_id, "search": search or "",
        "done_count": done_count, "total": total,
        "msg": request.query_params.get("msg"),
        "error": request.query_params.get("error"),
    })


@app.post("/kso/import")
async def kso_import(request: Request, db: Session = Depends(get_db),
                     file: UploadFile = File(...)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.worksheets[0]
        managers = db.query(models.Manager).all()
        created = 0
        # Ищем строку заголовков
        header_row = 1
        col_tk = col_addr = col_mgr = None
        for r in range(1, 6):
            for c in range(1, 20):
                v = str(ws.cell(r, c).value or '').strip().lower()
                if 'тк' in v or 'номер' in v: col_tk = c; header_row = r
                elif 'адрес' in v and not col_addr: col_addr = c
                elif 'менеджер' in v and not col_mgr: col_mgr = c
        col_tk = col_tk or 1; col_addr = col_addr or 2; col_mgr = col_mgr or 3
        for row_idx in range(header_row + 1, ws.max_row + 1):
            tk = str(ws.cell(row_idx, col_tk).value or '').strip()
            if not tk or tk in ('None', '—'): continue
            addr = str(ws.cell(row_idx, col_addr).value or '').strip()
            mgr_val = str(ws.cell(row_idx, col_mgr).value or '').strip()
            mgr_id = _match_manager(mgr_val, managers)
            existing = db.query(models.KsoObject).filter(
                models.KsoObject.tk_number == tk).first()
            if not existing:
                db.add(models.KsoObject(tk_number=tk, address=addr, manager_id=mgr_id))
                created += 1
        db.commit()
        return RedirectResponse(f"/kso?msg=Загружено: {created} объектов&tab=objects", status_code=303)
    except Exception as e:
        return RedirectResponse(f"/kso?error={str(e)[:100]}&tab=objects", status_code=303)


@app.post("/kso/objects/{obj_id}/toggle")
async def kso_toggle(obj_id: int, request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return {"error": "Не авторизован"}
    obj = db.query(models.KsoObject).filter(models.KsoObject.id == obj_id).first()
    if not obj:
        raise HTTPException(status_code=404)
    obj.done = not obj.done
    db.commit()
    return {"done": obj.done}


@app.post("/api/kso/{obj_id}/comment")
async def kso_comment(obj_id: int, request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return {"error": "Не авторизован"}
    data = await request.json()
    obj = db.query(models.KsoObject).filter(models.KsoObject.id == obj_id).first()
    if not obj:
        raise HTTPException(status_code=404)
    obj.comment = data.get("comment", "")
    db.commit()
    return {"ok": True}


@app.post("/kso/objects/{obj_id}/delete")
async def kso_delete_object(obj_id: int, request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user or not user.get("is_admin"):
        return RedirectResponse("/kso", status_code=302)
    obj = db.query(models.KsoObject).filter(models.KsoObject.id == obj_id).first()
    if obj:
        db.delete(obj)
        db.commit()
    return RedirectResponse("/kso?tab=objects", status_code=303)


@app.post("/kso/schedules/upload")
async def kso_schedule_upload(request: Request, db: Session = Depends(get_db),
                               file: UploadFile = File(...),
                               description: str = Form("")):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    save_dir = Path("static/uploads/kso")
    save_dir.mkdir(parents=True, exist_ok=True)
    safe_name = f"{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
    (save_dir / safe_name).write_bytes(await file.read())
    db.add(models.KsoSchedule(
        original_name=file.filename, filename=safe_name,
        description=description, uploaded_by=user.get("display_name", ""),
    ))
    db.commit()
    return RedirectResponse("/kso?tab=schedules&msg=Файл загружен", status_code=303)


@app.get("/kso/schedules/{sch_id}/download")
async def kso_schedule_download(sch_id: int, request: Request,
                                 db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    sch = db.query(models.KsoSchedule).filter(models.KsoSchedule.id == sch_id).first()
    if not sch:
        raise HTTPException(status_code=404)
    path = Path(f"static/uploads/kso/{sch.filename}")
    if not path.exists():
        raise HTTPException(status_code=404, detail="Файл не найден на сервере")
    return StreamingResponse(open(path, "rb"),
        headers={"Content-Disposition": f"attachment; filename=\"{sch.original_name}\""})


@app.post("/kso/schedules/{sch_id}/delete")
async def kso_schedule_delete(sch_id: int, request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    sch = db.query(models.KsoSchedule).filter(models.KsoSchedule.id == sch_id).first()
    if sch:
        path = Path(f"static/uploads/kso/{sch.filename}")
        if path.exists(): path.unlink()
        db.delete(sch)
        db.commit()
    return RedirectResponse("/kso?tab=schedules", status_code=303)


@app.post("/section/create-project")
async def section_create_project(request: Request, db: Session = Depends(get_db),
                                  name: str = Form(...), tk_number: str = Form(""),
                                  city: str = Form(""), project_type: str = Form(...),
                                  manager_id: str = Form(""), status: str = Form("Активный"),
                                  stage: str = Form(""), start_date: str = Form(""),
                                  end_date: str = Form(""), description: str = Form(""),
                                  budget: str = Form(""), redirect_to: str = Form("/")):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    p = models.Project(
        name=name, tk_number=tk_number, city=city, project_type=project_type,
        manager_id=int(manager_id) if manager_id else None,
        status=status, stage=stage, description=description,
        start_date=datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None,
        end_date=datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None,
        budget=float(budget.replace(",", ".")) if budget else None,
    )
    db.add(p)
    db.commit()
    return RedirectResponse(redirect_to, status_code=303)


# ─── PROJECTS ────────────────────────────────────────────────────────────────

@app.get("/projects", response_class=HTMLResponse)
async def projects_list(request: Request, db: Session = Depends(get_db),
                        manager_id: str = None, status: str = None,
                        project_type: str = None, search: str = None):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    q = db.query(models.Project)
    if manager_id and str(manager_id).isdigit():
        q = q.filter(models.Project.manager_id == int(manager_id))
    if status:
        q = q.filter(models.Project.status == status)
    if project_type:
        q = q.filter(models.Project.project_type == project_type)
    if search:
        q = q.filter(models.Project.tk_number.contains(search))
    projects = q.order_by(models.Project.end_date.nullslast()).all()
    managers = db.query(models.Manager).all()
    return templates.TemplateResponse("projects.html", {
        "request": request, "user": user,
        "projects": projects, "managers": managers,
        "project_types": PROJECT_TYPES, "statuses": STATUSES, "today": date.today(),
        "filter_manager_id": manager_id, "filter_status": status,
        "filter_type": project_type, "search": search or "",
    })


@app.post("/projects/create")
async def create_project(request: Request, db: Session = Depends(get_db),
                         name: str = Form(...), tk_number: str = Form(""),
                         city: str = Form(""), project_type: str = Form(""),
                         manager_id: str = Form(""), status: str = Form("Активный"),
                         stage: str = Form(""), start_date: str = Form(""),
                         end_date: str = Form(""), description: str = Form(""),
                         budget: str = Form("")):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    p = models.Project(
        name=name, tk_number=tk_number, city=city, project_type=project_type,
        manager_id=int(manager_id) if manager_id else None,
        status=status, stage=stage, description=description,
        start_date=datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None,
        end_date=datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None,
        budget=float(budget.replace(",", ".")) if budget else None,
    )
    db.add(p)
    db.commit()
    return RedirectResponse("/projects", status_code=303)


@app.get("/projects/{project_id}", response_class=HTMLResponse)
async def project_detail(request: Request, project_id: int, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404)
    managers = db.query(models.Manager).all()
    return templates.TemplateResponse("project_detail.html", {
        "request": request, "user": user,
        "project": project, "managers": managers,
        "project_types": PROJECT_TYPES, "statuses": STATUSES,
        "stage_names": STAGE_NAMES, "today": date.today(),
    })


@app.post("/projects/{project_id}/update")
async def update_project(project_id: int, request: Request, db: Session = Depends(get_db),
                         name: str = Form(...), tk_number: str = Form(""),
                         city: str = Form(""), project_type: str = Form(""),
                         manager_id: str = Form(""), status: str = Form("Активный"),
                         stage: str = Form(""), start_date: str = Form(""),
                         end_date: str = Form(""), description: str = Form(""),
                         budget: str = Form("")):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    p = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not p:
        raise HTTPException(status_code=404)
    p.name = name; p.tk_number = tk_number; p.city = city
    p.project_type = project_type; p.status = status; p.stage = stage
    p.manager_id = int(manager_id) if manager_id else None
    p.start_date = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
    p.end_date = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None
    p.description = description
    p.budget = float(budget.replace(",", ".")) if budget else None
    db.commit()
    return RedirectResponse(f"/projects/{project_id}", status_code=303)


@app.post("/projects/{project_id}/delete")
async def delete_project(project_id: int, request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    p = db.query(models.Project).filter(models.Project.id == project_id).first()
    if p:
        db.delete(p)
        db.commit()
    return RedirectResponse("/projects", status_code=303)


@app.post("/projects/{project_id}/stages/add")
async def add_stage(project_id: int, request: Request, db: Session = Depends(get_db),
                    name: str = Form(...), start_date: str = Form(""),
                    end_date: str = Form(""), stage_status: str = Form("Запланировано")):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    existing = db.query(models.ProjectStage).filter(
        models.ProjectStage.project_id == project_id).count()
    stage = models.ProjectStage(
        project_id=project_id, name=name, status=stage_status,
        start_date=datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None,
        end_date=datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None,
        order=existing,
    )
    db.add(stage)
    db.commit()
    return RedirectResponse(f"/projects/{project_id}", status_code=303)


@app.post("/stages/{stage_id}/delete")
async def delete_stage(stage_id: int, request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    s = db.query(models.ProjectStage).filter(models.ProjectStage.id == stage_id).first()
    project_id = s.project_id if s else None
    if s:
        db.delete(s)
        db.commit()
    return RedirectResponse(f"/projects/{project_id}", status_code=303)


# ─── DEADLINES ───────────────────────────────────────────────────────────────

@app.get("/deadlines", response_class=HTMLResponse)
async def deadlines(request: Request, db: Session = Depends(get_db),
                    manager_id: str = None, view: str = "all"):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    today = date.today()
    pq = db.query(models.Project).filter(
        models.Project.status != "Завершён",
        models.Project.end_date != None,
        # Исключаем уже открытые объекты
        (models.Project.opening_date == None) | (models.Project.opening_date > today)
    )
    if manager_id and str(manager_id).isdigit():
        pq = pq.filter(models.Project.manager_id == int(manager_id))
    projects = pq.order_by(models.Project.end_date).all()

    tq = db.query(models.Task).filter(
        models.Task.status != "Завершена", models.Task.deadline != None)
    if manager_id and str(manager_id).isdigit():
        tq = tq.filter(models.Task.assignee_id == int(manager_id))
    tasks = tq.order_by(models.Task.deadline).all()

    managers = db.query(models.Manager).all()
    return templates.TemplateResponse("deadlines.html", {
        "request": request, "user": user,
        "projects": projects, "tasks": tasks,
        "managers": managers, "today": today,
        "filter_manager_id": manager_id, "view": view,
    })


# ─── MANAGERS ────────────────────────────────────────────────────────────────

@app.get("/managers", response_class=HTMLResponse)
async def managers_view(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    today = date.today()
    _mgr_order = [
        "Месмер Денис", "Митько Роберт", "Ловчиков Александр",
        "Валеев Борис", "Косило Сергей", "Студеникин Сергей",
        "Хачатурова Жанна", "Шевченко Наталья",
    ]
    managers = db.query(models.Manager).all()
    managers.sort(key=lambda m: (
        0 if m.is_leader else 1,
        _mgr_order.index(m.name) if m.name in _mgr_order else 99
    ))
    stats = []
    leader_stats = []
    for m in managers:
        recon = [p for p in m.projects if p.project_type == "Реконструкция"]
        constr = [p for p in m.projects if p.project_type == "Констракшн"]
        active = sum(1 for p in m.projects if p.status == "Активный")
        open_t = sum(1 for t in m.tasks if t.status != "Завершена")
        overdue = sum(1 for t in m.tasks
                      if t.status != "Завершена" and t.deadline and t.deadline < today)
        urgent_p = [p for p in m.projects
                    if p.status == "Активный" and p.end_date
                    and 0 <= (p.end_date - today).days <= 14]
        stat = {"manager": m, "active_projects": active,
                "open_tasks": open_t, "overdue_tasks": overdue,
                "urgent_projects": urgent_p,
                "recon_projects": recon, "constr_projects": constr}
        if m.is_leader:
            leader_stats.append(stat)
        else:
            stats.append(stat)
    leader_stats.sort(key=lambda s: 0 if "Комаров" in s["manager"].name else 1)
    now = datetime.utcnow()
    online_set = {
        name for name, ts in ONLINE_USERS.items()
        if (now - ts).total_seconds() < ONLINE_TIMEOUT
    }
    return templates.TemplateResponse("managers.html", {
        "request": request, "user": user,
        "manager_stats": stats, "leader_stats": leader_stats,
        "today": today, "online_set": online_set,
    })


@app.post("/managers/add")
async def add_manager(request: Request, db: Session = Depends(get_db),
                      name: str = Form(...), phone: str = Form(""),
                      is_leader: str = Form("")):
    user = get_current_user(request)
    if not user or not user.get("is_admin"):
        return RedirectResponse("/managers", status_code=302)
    mgr = models.Manager(name=name.strip(), is_leader=bool(is_leader))
    db.add(mgr)
    db.flush()
    if phone.strip():
        normalized = normalize_phone(phone.strip())
        exists = db.query(models.PhoneWhitelist).filter(
            models.PhoneWhitelist.phone == normalized).first()
        if not exists:
            db.add(models.PhoneWhitelist(
                phone=normalized, display_name=name.strip(), is_admin=False))
    db.commit()
    return RedirectResponse("/managers", status_code=303)


@app.post("/managers/{manager_id}/delete")
async def delete_manager(manager_id: int, request: Request,
                         db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user or not user.get("is_admin"):
        return RedirectResponse("/managers", status_code=302)
    mgr = db.query(models.Manager).filter(models.Manager.id == manager_id).first()
    if mgr:
        # Обнуляем manager_id у проектов и задач
        for p in mgr.projects:
            p.manager_id = None
        for t in mgr.tasks:
            t.assignee_id = None
        db.delete(mgr)
        db.commit()
    return RedirectResponse("/managers", status_code=303)


@app.post("/managers/{manager_id}/photo")
async def upload_manager_photo(manager_id: int, request: Request,
                                file: UploadFile = File(...),
                                db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user or user.get("display_name") != "Месмер Денис":
        return RedirectResponse("/managers", status_code=302)
    mgr = db.query(models.Manager).filter(models.Manager.id == manager_id).first()
    if not mgr:
        raise HTTPException(status_code=404)

    ext = Path(file.filename).suffix.lower() if file.filename else ".jpg"
    if ext not in ('.jpg', '.jpeg', '.png', '.webp'):
        ext = '.jpg'

    save_dir = Path("static/img/managers")
    save_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{manager_id}{ext}"
    filepath = save_dir / filename

    content = await file.read()
    filepath.write_bytes(content)

    mgr.photo = f"img/managers/{filename}"
    db.commit()
    return RedirectResponse("/managers", status_code=303)


# ─── TASKS ───────────────────────────────────────────────────────────────────

@app.get("/tasks", response_class=HTMLResponse)
async def tasks_view(request: Request, db: Session = Depends(get_db),
                     manager_id: str = None, status: str = None):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    my_name = user.get("display_name", "")
    is_admin = user.get("is_admin", False)

    # Найти менеджера текущего пользователя
    my_manager = db.query(models.Manager).filter(models.Manager.name == my_name).first()

    from sqlalchemy import or_
    q = db.query(models.Task)
    if not is_admin:
        # Обычный менеджер видит только свои задачи
        conditions = [models.Task.created_by == my_name]
        if my_manager:
            conditions.append(models.Task.assignee_id == my_manager.id)
        q = q.filter(or_(*conditions))
    else:
        # Админ/руководитель видит все, но может фильтровать
        if manager_id and str(manager_id).isdigit():
            q = q.filter(models.Task.assignee_id == int(manager_id))

    if status:
        q = q.filter(models.Task.status == status)
    tasks = q.order_by(models.Task.deadline.nullslast()).all()
    managers = db.query(models.Manager).all()
    projects = db.query(models.Project).filter(models.Project.status == "Активный").all()

    # Непрочитанные уведомления по задачам
    unread_notifs = db.query(models.TaskNotification).filter(
        models.TaskNotification.recipient_name == my_name,
        models.TaskNotification.is_read == False,
    ).count()

    return templates.TemplateResponse("create_task.html", {
        "request": request, "user": user,
        "tasks": tasks, "managers": managers, "projects": projects,
        "priorities": PRIORITIES, "task_statuses": TASK_STATUSES,
        "today": date.today(),
        "filter_manager_id": manager_id, "filter_status": status,
        "unread_notifs": unread_notifs,
    })


@app.post("/tasks/create")
async def create_task(request: Request, db: Session = Depends(get_db),
                      title: str = Form(...), description: str = Form(""),
                      project_id: str = Form(""), assignee_id: str = Form(""),
                      deadline: str = Form(""), priority: str = Form("Средний")):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    creator = user.get("display_name", "")
    task = models.Task(
        title=title, description=description,
        project_id=int(project_id) if project_id else None,
        assignee_id=int(assignee_id) if assignee_id else None,
        deadline=datetime.strptime(deadline, "%Y-%m-%d").date() if deadline else None,
        priority=priority,
        created_by=creator,
    )
    db.add(task)
    db.flush()

    # Уведомление исполнителю
    if assignee_id:
        assignee = db.query(models.Manager).filter(
            models.Manager.id == int(assignee_id)).first()
        if assignee and assignee.name != creator:
            dl = f", дедлайн {task.deadline.strftime('%d.%m.%Y')}" if task.deadline else ""
            db.add(models.TaskNotification(
                recipient_name=assignee.name,
                task_id=task.id,
                message=f"📋 {creator} поставил вам задачу: «{title}»{dl}",
            ))
    db.commit()
    return RedirectResponse("/tasks", status_code=303)


@app.post("/tasks/{task_id}/update-status")
async def update_task_status(task_id: int, request: Request, db: Session = Depends(get_db),
                             status: str = Form(...),
                             completion_comment: str = Form("")):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    t = db.query(models.Task).filter(models.Task.id == task_id).first()
    if t:
        old_status = t.status
        t.status = status
        if status == "Завершена" and completion_comment.strip():
            t.completion_comment = completion_comment.strip()

        # Уведомление постановщику при смене статуса
        if t.created_by and t.created_by != user.get("display_name", "") and old_status != status:
            status_icon = "✅" if status == "Завершена" else "🔄"
            msg = f"{status_icon} Задача «{t.title}»: статус изменён на «{status}»"
            if status == "Завершена" and completion_comment.strip():
                msg += f"\nКомментарий: {completion_comment.strip()}"
            db.add(models.TaskNotification(
                recipient_name=t.created_by,
                task_id=t.id,
                message=msg,
            ))
        db.commit()
    return RedirectResponse("/tasks", status_code=303)


@app.post("/tasks/{task_id}/delete")
async def delete_task(task_id: int, request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    t = db.query(models.Task).filter(models.Task.id == task_id).first()
    if t:
        db.delete(t)
        db.commit()
    return RedirectResponse("/tasks", status_code=303)


# ─── EXCEL ───────────────────────────────────────────────────────────────────

@app.post("/projects/import-excel")
async def import_excel(request: Request, db: Session = Depends(get_db),
                       file: UploadFile = File(...), manager_id: str = Form("")):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        return RedirectResponse("/projects?error=invalid_excel", status_code=303)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        project_name = sheet_name
        tk_num = ""
        city_name = ""

        for row in ws.iter_rows(max_row=10, values_only=True):
            for cell in row:
                if cell and isinstance(cell, str):
                    text = cell.strip()
                    if text.upper().startswith("ТК") or text.upper().startswith("TK"):
                        project_name = text
                        parts = text.split()
                        if len(parts) >= 2:
                            tk_num = parts[1]
                        if len(parts) >= 3:
                            city_name = " ".join(parts[2:])

        p = models.Project(
            name=project_name, tk_number=tk_num, city=city_name,
            manager_id=int(manager_id) if manager_id else None,
            status="Активный",
        )
        db.add(p)
        db.flush()

        order = 0
        for row in ws.iter_rows(values_only=True):
            if not row[0] or not isinstance(row[0], str):
                continue
            cell_text = row[0].strip()
            for sn in STAGE_NAMES:
                if sn.split()[0].lower() in cell_text.lower():
                    start = end = None
                    for v in row[1:]:
                        if isinstance(v, datetime):
                            if not start:
                                start = v.date()
                            elif not end:
                                end = v.date()
                                break
                        elif isinstance(v, date):
                            if not start:
                                start = v
                            elif not end:
                                end = v
                                break
                    db.add(models.ProjectStage(
                        project_id=p.id, name=cell_text,
                        start_date=start, end_date=end, order=order))
                    order += 1
                    break

    db.commit()
    return RedirectResponse("/projects", status_code=303)


@app.get("/api/export/projects-excel")
async def export_excel(db: Session = Depends(get_db), type: str = None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Проекты"

    hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")

    headers = ["№", "ТК №", "Название проекта", "Город", "Тип", "Менеджер",
               "Статус", "Этап", "Дата начала", "Дата окончания", "Бюджет, руб."]
    ws.row_dimensions[1].height = 20
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont; c.alignment = center

    q = db.query(models.Project)
    if type:
        q = q.filter(models.Project.project_type == type)
    projects = q.all()
    today = date.today()
    for i, p in enumerate(projects, 2):
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=p.tk_number)
        ws.cell(row=i, column=3, value=p.name)
        ws.cell(row=i, column=4, value=p.city)
        ws.cell(row=i, column=5, value=p.project_type)
        ws.cell(row=i, column=6, value=p.manager.name if p.manager else "")
        ws.cell(row=i, column=7, value=p.status)
        ws.cell(row=i, column=8, value=p.stage)
        ws.cell(row=i, column=9, value=p.start_date)
        ws.cell(row=i, column=10, value=p.end_date)
        ws.cell(row=i, column=11, value=p.budget)

        if p.end_date:
            days = (p.end_date - today).days
            color = "FFCCCC" if days < 0 else "FFEB9C" if days <= 7 else "D4EDDA"
            row_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for col in range(1, 12):
                ws.cell(row=i, column=col).fill = row_fill

    col_widths = [4, 8, 40, 15, 18, 18, 14, 20, 14, 14, 15]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=lenta_projects.xlsx"},
    )


# ─── API ─────────────────────────────────────────────────────────────────────

# ─── ВПК ─────────────────────────────────────────────────────────────────────

@app.get("/vpk", response_class=HTMLResponse)
async def vpk_page(request: Request, db: Session = Depends(get_db),
                   tab: str = "vpk1"):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    projects = db.query(models.Project).filter(
        models.Project.project_type == "Констракшн"
    ).order_by(models.Project.tk_number).all()
    criteria1 = db.query(models.VpkCriterion).filter(
        models.VpkCriterion.vpk_type == 1).order_by(models.VpkCriterion.order).all()
    criteria2 = db.query(models.VpkCriterion).filter(
        models.VpkCriterion.vpk_type == 2).order_by(models.VpkCriterion.order).all()
    reports = db.query(models.VpkReport).order_by(
        models.VpkReport.submitted_at.desc()).limit(50).all()

    display_name = user.get("display_name", "")
    unread = 0
    if "Гаврин" in display_name:
        unread = db.query(models.VpkReport).filter(models.VpkReport.read_gavrin == False).count()
    elif "Месмер" in display_name:
        unread = db.query(models.VpkReport).filter(models.VpkReport.read_mesmer == False).count()

    return templates.TemplateResponse("vpk.html", {
        "request": request, "user": user,
        "tab": tab, "projects": projects,
        "criteria1": criteria1, "criteria2": criteria2,
        "reports": reports, "unread": unread,
        "msg": request.query_params.get("msg"),
    })


@app.post("/vpk/submit")
async def vpk_submit(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    form = await request.form()
    project_id = form.get("project_id")
    vpk_type = int(form.get("vpk_type", 1))

    report = models.VpkReport(
        project_id=int(project_id) if project_id else None,
        vpk_type=vpk_type,
        submitted_by=user.get("display_name", ""),
        read_gavrin=False,
        read_mesmer=False,
    )
    db.add(report)
    db.flush()

    criteria = db.query(models.VpkCriterion).filter(
        models.VpkCriterion.vpk_type == vpk_type
    ).order_by(models.VpkCriterion.order).all()

    photo_dir = Path("static/uploads/vpk")
    photo_dir.mkdir(parents=True, exist_ok=True)

    for c in criteria:
        done = form.get(f"criterion_{c.id}") == "on"
        comment = str(form.get(f"comment_{c.id}", "") or "").strip()

        photo_path = ""
        photo_file = form.get(f"photo_{c.id}")
        if photo_file and hasattr(photo_file, "filename") and photo_file.filename:
            ext = Path(photo_file.filename).suffix.lower() or ".jpg"
            fname = f"{report.id}_{c.id}{ext}"
            (photo_dir / fname).write_bytes(await photo_file.read())
            photo_path = f"uploads/vpk/{fname}"

        db.add(models.VpkReportItem(
            report_id=report.id, criterion_id=c.id,
            criterion_name=c.name, done=done,
            comment=comment, photo_path=photo_path,
        ))
    db.commit()

    proj = db.query(models.Project).filter(models.Project.id == project_id).first()
    tk = proj.tk_number if proj else "—"
    return RedirectResponse(
        f"/vpk?tab=reports&msg=Отчёт ВПК{vpk_type} по ТК {tk} отправлен",
        status_code=303)


@app.post("/vpk/reports/{report_id}/read")
async def vpk_mark_read(report_id: int, request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return {"error": "Не авторизован"}
    report = db.query(models.VpkReport).filter(models.VpkReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404)
    display_name = user.get("display_name", "")
    if "Гаврин" in display_name:
        report.read_gavrin = True
    elif "Месмер" in display_name:
        report.read_mesmer = True
    db.commit()
    return {"ok": True}


@app.get("/api/vpk/unread")
async def vpk_unread(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return {"reports": []}
    display_name = user.get("display_name", "")
    if "Гаврин" in display_name:
        reports = db.query(models.VpkReport).filter(
            models.VpkReport.read_gavrin == False).all()
    elif "Месмер" in display_name:
        reports = db.query(models.VpkReport).filter(
            models.VpkReport.read_mesmer == False).all()
    else:
        return {"reports": []}

    result = []
    for r in reports:
        tk = r.project.tk_number if r.project else "—"
        done = sum(1 for i in r.items if i.done)
        total = len(r.items)
        result.append({
            "id": r.id,
            "title": f"Новый отчёт ВПК{r.vpk_type} по ТК {tk}",
            "body": f"Подал: {r.submitted_by} | Выполнено: {done}/{total} критериев",
        })
    return {"reports": result}


@app.get("/api/notifications/construction")
async def construction_notifications(request: Request, db: Session = Depends(get_db)):
    """Returns upcoming Construction deadline notifications for the current user's manager."""
    user = get_current_user(request)
    if not user:
        return {"notifications": []}

    today = date.today()
    display_name = user.get("display_name", "")

    # Найти менеджера по имени из сессии
    managers = db.query(models.Manager).all()
    manager = None
    if display_name:
        name_part = display_name.split()[0].lower()
        for m in managers:
            if m.name.lower().startswith(name_part) or name_part in m.name.lower():
                manager = m
                break

    # Руководитель видит все проекты
    is_leader = user.get("is_admin") or (manager and manager.is_leader)

    q = db.query(models.Project).filter(models.Project.project_type == "Констракшн")
    if not is_leader and manager:
        q = q.filter(models.Project.manager_id == manager.id)
    elif not is_leader:
        return {"notifications": []}

    projects = q.all()
    notifications = []

    for p in projects:
        tk = p.tk_number or str(p.id)
        mgr_name = p.manager.name if p.manager else ""

        # За 2 дня до выхода на СМР
        if p.closure_date:
            days = (p.closure_date - today).days
            if days in (0, 1, 2):
                day_label = "сегодня" if days == 0 else ("завтра" if days == 1 else f"через {days} дня")
                notifications.append({
                    "type": "smr",
                    "title": f"{mgr_name}: На ТК {tk} {day_label} выход на СМР",
                    "body": f"Дата выхода на СМР: {p.closure_date.strftime('%d.%m.%Y')}",
                    "urgency": "high",
                    "date": str(p.closure_date),
                })

        # За 3 дня до ВПК1
        if p.vpk_date:
            days = (p.vpk_date - today).days
            if days in (0, 1, 2, 3):
                notifications.append({
                    "type": "vpk",
                    "title": f"{mgr_name}: Через {days} {'день' if days == 1 else 'дня'} на ТК {tk} ВПК1",
                    "body": f"Дата ВПК1: {p.vpk_date.strftime('%d.%m.%Y')}",
                    "urgency": "high",
                    "date": str(p.vpk_date),
                })

        # В день открытия — поздравление
        if p.opening_date and p.opening_date == today:
            notifications.append({
                "type": "opening",
                "title": f"Поздравляю с открытием! 🎉",
                "body": f"{mgr_name}, поздравляю с открытием ТК {tk}!!!",
                "urgency": "celebration",
                "date": str(p.opening_date),
            })

    return {"notifications": notifications, "manager": manager.name if manager else "Все"}


@app.get("/api/notifications/reconstruct")
async def reconstruct_notifications(request: Request, db: Session = Depends(get_db)):
    """Returns upcoming Reconstruction deadline notifications for the current user's manager."""
    user = get_current_user(request)
    if not user:
        return {"notifications": []}

    today = date.today()
    display_name = user.get("display_name", "")

    managers = db.query(models.Manager).all()
    manager = None
    if display_name:
        name_part = display_name.split()[0].lower()
        for m in managers:
            if m.name.lower().startswith(name_part) or name_part in m.name.lower():
                manager = m
                break

    is_leader = user.get("is_admin") or (manager and manager.is_leader)

    q = db.query(models.Project).filter(models.Project.project_type == "Реконструкция")
    if not is_leader and manager:
        q = q.filter(models.Project.manager_id == manager.id)
    elif not is_leader:
        return {"notifications": []}

    projects = q.all()
    notifications = []

    for p in projects:
        tk = p.tk_number or str(p.id)
        mgr_name = p.manager.name if p.manager else ""

        # За 3 дня до начала СИД
        if p.sid_start:
            days = (p.sid_start - today).days
            if days in (1, 2, 3):
                notifications.append({
                    "type": "sid",
                    "title": f"{mgr_name}: ТК {tk} приближается дата сбора данных",
                    "body": f"Начало СИД: {p.sid_start.strftime('%d.%m.%Y')} (через {days} {'день' if days == 1 else 'дня'})",
                    "urgency": "high",
                    "date": str(p.sid_start),
                })

        # За 1 день до начала Зонирования
        if p.zoning_start:
            days = (p.zoning_start - today).days
            if days == 1:
                notifications.append({
                    "type": "zoning",
                    "title": f"{mgr_name}: ТК {tk}, провести Зонирование",
                    "body": f"Начало зонирования завтра: {p.zoning_start.strftime('%d.%m.%Y')}",
                    "urgency": "high",
                    "date": str(p.zoning_start),
                })

    return {"notifications": notifications, "manager": manager.name if manager else "Все"}


@app.get("/api/data-version")
async def data_version(db: Session = Depends(get_db)):
    """Returns timestamp of latest project change — used by clients for polling."""
    from sqlalchemy import func
    result = db.query(func.max(models.Project.updated_at)).scalar()
    ts = result.isoformat() if result else "0"
    return {"version": ts}


@app.get("/api/deadlines/check")
async def check_deadlines(request: Request, db: Session = Depends(get_db)):
    if not get_current_user(request):
        return {"urgent_tasks": [], "overdue_tasks": [], "urgent_projects": []}
    today = date.today()
    urgent_tasks = db.query(models.Task).filter(
        models.Task.status != "Завершена", models.Task.deadline != None,
        models.Task.deadline >= today,
        models.Task.deadline <= today + timedelta(days=3)).all()
    overdue_tasks = db.query(models.Task).filter(
        models.Task.status != "Завершена", models.Task.deadline != None,
        models.Task.deadline < today).all()
    urgent_projects = db.query(models.Project).filter(
        models.Project.status != "Завершён",
        models.Project.end_date != None,
        models.Project.end_date >= today,
        models.Project.end_date <= today + timedelta(days=7),
        (models.Project.opening_date == None) | (models.Project.opening_date > today)
    ).all()
    return {
        "urgent_tasks": [{"id": t.id, "title": t.title, "deadline": str(t.deadline),
                          "assignee": t.assignee.name if t.assignee else "",
                          "days_left": (t.deadline - today).days} for t in urgent_tasks],
        "overdue_tasks": [{"id": t.id, "title": t.title, "deadline": str(t.deadline),
                           "assignee": t.assignee.name if t.assignee else "",
                           "days_overdue": (today - t.deadline).days} for t in overdue_tasks],
        "urgent_projects": [{"id": p.id, "name": p.name, "deadline": str(p.end_date),
                              "manager": p.manager.name if p.manager else "",
                              "days_left": (p.end_date - today).days} for p in urgent_projects],
    }


# ─── STATISTICS ──────────────────────────────────────────────────────────────

@app.get("/stats", response_class=HTMLResponse)
async def stats_page(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    today = date.today()
    managers = db.query(models.Manager).all()

    # Только Констракшн
    projects = db.query(models.Project).filter(
        models.Project.project_type == "Констракшн"
    ).all()

    def opening_color(p):
        """Зелёный=раньше срока, белый=вовремя/нет плана, красный=опоздание."""
        # В работе = ещё не открыт сегодня
        if not p.opening_date or p.opening_date > today:
            return "active"
        if not p.end_date:
            return "on-time"
        if p.opening_date < p.end_date:
            return "early"
        if p.opening_date == p.end_date:
            return "on-time"
        return "late"

    manager_stats = []
    for m in managers:
        m_proj = [p for p in projects if p.manager_id == m.id]
        if not m_proj:
            continue
        early   = sum(1 for p in m_proj if opening_color(p) == "early")
        on_time = sum(1 for p in m_proj if opening_color(p) == "on-time")
        late    = sum(1 for p in m_proj if opening_color(p) == "late")
        # В работе = ещё не открыты сегодня
        active  = sum(1 for p in m_proj if not p.opening_date or p.opening_date > today)
        manager_stats.append({
            "name": m.name, "early": early, "on_time": on_time,
            "late": late, "active": active, "total": len(m_proj),
        })

    projects_with_stats = []
    for p in sorted(projects, key=lambda x: (x.manager_id or 0, x.end_date or date.max)):
        color = opening_color(p)
        delta = None
        if p.opening_date and p.end_date:
            delta = (p.opening_date - p.end_date).days
        days_left = (p.end_date - today).days if p.end_date and not p.opening_date else None

        projects_with_stats.append({
            "id": p.id,
            "tk_number": p.tk_number,
            "address": p.address or p.city or "—",
            "format_type": p.format_type or "",
            "manager_name": p.manager.name if p.manager else "—",
            "end_date": p.end_date,
            "opening_date": p.opening_date,
            "color": color,
            "delta_days": delta,
            "days_left": days_left,
            "delay_reason": p.delay_reason or "",
        })

    total_early   = sum(s["early"]   for s in manager_stats)
    total_on_time = sum(s["on_time"] for s in manager_stats)
    total_late    = sum(s["late"]    for s in manager_stats)
    total_active  = sum(s["active"]  for s in manager_stats)

    return templates.TemplateResponse("stats.html", {
        "request": request, "user": user,
        "manager_stats": manager_stats,
        "projects_with_stats": projects_with_stats,
        "total_early": total_early, "total_on_time": total_on_time,
        "total_late": total_late, "total_active": total_active,
        "today": today,
    })


@app.post("/api/projects/{project_id}/delay-reason")
async def save_delay_reason(project_id: int, request: Request,
                             db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return {"error": "Не авторизован"}
    data = await request.json()
    p = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not p:
        raise HTTPException(status_code=404)
    p.delay_reason = data.get("reason", "")
    p.updated_at = datetime.utcnow()
    db.commit()
    return {"ok": True}


@app.post("/stats/upload")
async def stats_upload(request: Request, db: Session = Depends(get_db),
                       file: UploadFile = File(...)):
    """Upload Excel with actual opening dates to update projects."""
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        updated = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                tk_num = None
                open_date = None
                for val in row:
                    if isinstance(val, (int, float)) and 100 < val < 9999:
                        tk_num = str(int(val))
                    if isinstance(val, datetime):
                        open_date = val.date()
                    elif isinstance(val, date) and not isinstance(val, datetime):
                        open_date = val
                if tk_num and open_date:
                    p = db.query(models.Project).filter(
                        models.Project.tk_number == tk_num).first()
                    if p:
                        p.opening_date = open_date
                        today = date.today()
                        if open_date <= today:
                            p.status = "Завершён"
                        updated += 1
        db.commit()
    except Exception as e:
        return RedirectResponse(f"/stats?error={str(e)[:80]}", status_code=303)
    return RedirectResponse(f"/stats?updated={updated}", status_code=303)


@app.get("/stats/export")
async def stats_export(request: Request, db: Session = Depends(get_db),
                       date_from: str = "", date_to: str = ""):
    """Export opened Construction projects to Excel with color coding."""
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)

    today_date = date.today()
    q = db.query(models.Project).filter(
        models.Project.project_type == "Констракшн",
        models.Project.opening_date != None,
        models.Project.opening_date <= today_date,
    )
    if date_from:
        try:
            q = q.filter(models.Project.opening_date >= datetime.strptime(date_from, "%Y-%m-%d").date())
        except ValueError:
            pass
    if date_to:
        try:
            q = q.filter(models.Project.opening_date <= datetime.strptime(date_to, "%Y-%m-%d").date())
        except ValueError:
            pass
    projects = q.order_by(models.Project.manager_id, models.Project.opening_date).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Статистика открытий"

    # Стили заголовка
    hfill = PatternFill(start_color="1A5C22", end_color="1A5C22", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Цвета строк
    fill_early   = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    fill_on_time = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_late    = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    headers = ["№", "ТК №", "Адрес", "Формат", "Менеджер",
               "План открытия", "Факт открытия", "Отклонение (дн)", "Результат", "Комментарий"]
    col_widths = [5, 12, 40, 10, 20, 16, 16, 16, 18, 35]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill
        c.font = hfont
        c.alignment = center
        ws.column_dimensions[c.column_letter].width = w

    ws.row_dimensions[1].height = 30

    for i, p in enumerate(projects, 1):
        # Определяем результат
        if p.end_date:
            if p.opening_date < p.end_date:
                result = "Раньше срока"
                row_fill = fill_early
                delta = (p.opening_date - p.end_date).days
            elif p.opening_date == p.end_date:
                result = "Вовремя"
                row_fill = fill_on_time
                delta = 0
            else:
                result = "С опозданием"
                row_fill = fill_late
                delta = (p.opening_date - p.end_date).days
        else:
            result = "Вовремя"
            row_fill = fill_on_time
            delta = None

        row_data = [
            i,
            p.tk_number or "",
            p.address or p.city or "",
            p.format_type or "",
            p.manager.name if p.manager else "—",
            p.end_date.strftime("%d.%m.%Y") if p.end_date else "—",
            p.opening_date.strftime("%d.%m.%Y"),
            delta if delta is not None else "—",
            result,
            p.delay_reason or "",
        ]

        row_num = i + 1
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row_num, column=col, value=val)
            c.fill = row_fill
            if col in (6, 7):
                c.alignment = Alignment(horizontal="center")
            if col == 8 and isinstance(val, int):
                c.alignment = Alignment(horizontal="center")
                if val < 0:
                    c.font = Font(color="16A34A", bold=True)
                elif val > 0:
                    c.font = Font(color="DC2626", bold=True)

    # Итоговая строка
    total_row = len(projects) + 2
    early_count  = sum(1 for p in projects if p.end_date and p.opening_date < p.end_date)
    ontime_count = sum(1 for p in projects if not p.end_date or p.opening_date == p.end_date)
    late_count   = sum(1 for p in projects if p.end_date and p.opening_date > p.end_date)

    ws.cell(row=total_row, column=1, value="ИТОГО").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=len(projects)).font = Font(bold=True)
    ws.cell(row=total_row, column=9,
            value=f"Раньше: {early_count} | Вовремя: {ontime_count} | Опозд.: {late_count}"
            ).font = Font(bold=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    period = f"{date_from or 'начало'}_{date_to or date.today().strftime('%Y-%m-%d')}"
    filename = f"construction_stats_{period}.xlsx"
    return StreamingResponse(output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"})


# ─── ADMIN: WHITELIST & USERS ────────────────────────────────────────────────

@app.get("/admin/users", response_class=HTMLResponse)
async def admin_users(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user or not user.get("is_admin"):
        return RedirectResponse("/", status_code=302)
    whitelist = db.query(models.PhoneWhitelist).order_by(
        models.PhoneWhitelist.created_at).all()
    users = db.query(models.User).all()
    # Map phone → user
    user_by_phone = {u.phone: u for u in users}
    return templates.TemplateResponse("admin_users.html", {
        "request": request, "user": user,
        "whitelist": whitelist, "user_by_phone": user_by_phone,
    })


@app.post("/admin/whitelist/add")
async def add_to_whitelist(request: Request, db: Session = Depends(get_db),
                           phone: str = Form(...), display_name: str = Form(""),
                           is_admin: str = Form("")):
    user = get_current_user(request)
    if not user or not user.get("is_admin"):
        return RedirectResponse("/", status_code=302)
    normalized = normalize_phone(phone)
    if not db.query(models.PhoneWhitelist).filter(
            models.PhoneWhitelist.phone == normalized).first():
        db.add(models.PhoneWhitelist(
            phone=normalized, display_name=display_name.strip(),
            is_admin=bool(is_admin),
        ))
        db.commit()
    return RedirectResponse("/admin/users", status_code=303)


@app.post("/admin/whitelist/{wl_id}/delete")
async def remove_from_whitelist(wl_id: int, request: Request,
                                db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user or not user.get("is_admin"):
        return RedirectResponse("/", status_code=302)
    wl = db.query(models.PhoneWhitelist).filter(
        models.PhoneWhitelist.id == wl_id).first()
    if wl:
        # Also remove linked user account
        linked_user = db.query(models.User).filter(
            models.User.phone == wl.phone).first()
        if linked_user and linked_user.id != user.get("id"):
            db.delete(linked_user)
        db.delete(wl)
        db.commit()
    return RedirectResponse("/admin/users", status_code=303)


@app.post("/admin/users/{user_id}/reset-password")
async def reset_password(user_id: int, request: Request,
                         db: Session = Depends(get_db)):
    """Reset user's password so they have to create a new one on next login."""
    current = get_current_user(request)
    if not current or not current.get("is_admin"):
        return RedirectResponse("/", status_code=302)
    u = db.query(models.User).filter(models.User.id == user_id).first()
    if u:
        u.password_hash = None
        db.commit()
    return RedirectResponse("/admin/users", status_code=303)


# ─── СПЕЦИАЛИЗИРОВАННЫЙ ИМПОРТ РЕКОНСТРУКЦИЙ ─────────────────────────────────

@app.get("/import-reconstruct", response_class=HTMLResponse)
async def import_reconstruct_page(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    return templates.TemplateResponse("import_reconstruct.html", {
        "request": request, "user": user,
        "section_title": "Реконструкции",
        "form_action": "/import-reconstruct",
        "file_accept": ".xlsx,.xls",
        "msg": request.query_params.get("msg"),
        "error": request.query_params.get("error"),
    })


@app.post("/import-reconstruct")
async def do_import_reconstruct(request: Request, db: Session = Depends(get_db),
                                 file: UploadFile = File(...)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    content = await file.read()
    try:
        result = import_reconstruct_excel(content, db)
        return RedirectResponse(
            f"/import-reconstruct?msg=Импорт завершён: создано {result['created']}, обновлено {result['updated']} проектов",
            status_code=303)
    except Exception as e:
        return RedirectResponse(f"/import-reconstruct?error={str(e)[:120]}", status_code=303)


# ─── ИМПОРТ КОНСТРАКШН ───────────────────────────────────────────────────────

@app.get("/import-construction", response_class=HTMLResponse)
async def import_construction_page(request: Request):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    return templates.TemplateResponse("import_reconstruct.html", {
        "request": request, "user": user,
        "section_title": "Констракшн",
        "form_action": "/import-construction",
        "file_accept": ".xlsx,.xls,.xlsm",
        "msg": request.query_params.get("msg"),
        "error": request.query_params.get("error"),
    })


@app.post("/projects/clear-all")
async def clear_all_projects(request: Request, db: Session = Depends(get_db)):
    """Удаляет ВСЕ проекты Реконструкций и Констракшн."""
    user = get_current_user(request)
    if not user or not user.get("is_admin"):
        return RedirectResponse("/", status_code=302)
    projects = db.query(models.Project).filter(
        models.Project.project_type.in_(["Реконструкция", "Констракшн"])
    ).all()
    for p in projects:
        db.delete(p)
    db.commit()
    return RedirectResponse("/?msg=Все проекты удалены", status_code=303)


@app.post("/reconstruct/delete-tk-prefix")
async def delete_tk_prefix(request: Request, db: Session = Depends(get_db)):
    """Удаляет из Реконструкций проекты где номер ТК начинается с 'ТК' или 'L'."""
    user = get_current_user(request)
    if not user or not user.get("is_admin"):
        return RedirectResponse("/login", status_code=302)
    projects = db.query(models.Project).filter(
        models.Project.project_type == "Реконструкция",
        models.Project.tk_number.op("~")("^(ТК|TK|L)")
    ).all()
    count = len(projects)
    for p in projects:
        db.delete(p)
    db.commit()
    return RedirectResponse(f"/reconstruct?msg=Удалено: {count} проектов", status_code=303)


@app.post("/reconstruct/fix-types")
async def fix_reconstruct_types(request: Request, db: Session = Depends(get_db)):
    """Перемещает L-номера из Реконструкций в Констракшн."""
    user = get_current_user(request)
    if not user or not user.get("is_admin"):
        return RedirectResponse("/login", status_code=302)
    moved = 0
    wrong = db.query(models.Project).filter(
        models.Project.project_type == "Реконструкция",
        models.Project.tk_number.like("L%"),
    ).all()
    for p in wrong:
        p.project_type = "Констракшн"
        moved += 1
    db.commit()
    return RedirectResponse(f"/reconstruct?msg=Перемещено в Констракшн: {moved} проектов", status_code=303)


@app.post("/construction/clear-all")
async def clear_all_construction(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user or not user.get("is_admin"):
        return RedirectResponse("/login", status_code=302)
    projects = db.query(models.Project).filter(
        models.Project.project_type == "Констракшн"
    ).all()
    for p in projects:
        db.delete(p)
    db.commit()
    return RedirectResponse("/construction?msg=Все проекты Констракшн удалены. Загрузите файл заново.", status_code=303)


@app.post("/construction/delete-non-2026")
async def delete_construction_non_2026(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user or not user.get("is_admin"):
        return RedirectResponse("/login", status_code=302)
    deleted = db.query(models.Project).filter(
        models.Project.project_type == "Констракшн",
        models.Project.end_date != None,
        models.Project.end_date < date(2026, 1, 1),
    ).delete()
    # Также удаляем где дата открытия в 2025
    deleted2 = db.query(models.Project).filter(
        models.Project.project_type == "Констракшн",
        models.Project.opening_date != None,
        models.Project.opening_date < date(2026, 1, 1),
    ).delete()
    db.commit()
    return RedirectResponse(f"/construction?msg=Удалено объектов 2025 года: {deleted + deleted2}", status_code=303)


@app.post("/import-construction")
async def do_import_construction(request: Request, db: Session = Depends(get_db),
                                  file: UploadFile = File(...)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    content = await file.read()
    try:
        result = import_construction_excel(content, db)
        msg = (f"Создано:{result['created']} Обновлено:{result['updated']} "
               f"Строк:{result.get('rows_with_tk',0)} "
               f"Пропущено_формат:{result.get('skipped_fmt',0)} "
               f"Пропущено_менеджер:{result.get('skipped_mgr',0)} | "
               f"Форматы:[{','.join(result.get('sample_formats',[]))}] "
               f"Менеджеры:[{','.join(result.get('sample_managers',[]))}]")
        return RedirectResponse(f"/import-construction?msg={msg}", status_code=303)
    except Exception as e:
        return RedirectResponse(f"/import-construction?error={str(e)[:120]}", status_code=303)


# ─── SMART EXCEL IMPORT (section) ────────────────────────────────────────────

@app.post("/import-excel-section")
async def import_excel_section(request: Request, db: Session = Depends(get_db),
                                file: UploadFile = File(...),
                                project_type: str = Form(""),
                                manager_id: str = Form(""),
                                redirect_to: str = Form("/")):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    content = await file.read()
    try:
        result = parse_excel_file(content, project_type, int(manager_id) if manager_id else None, db)
        return RedirectResponse(
            f"{redirect_to}?msg=created:{result['created']},updated:{result['updated']}",
            status_code=303)
    except Exception as e:
        return RedirectResponse(f"{redirect_to}?error={str(e)[:80]}", status_code=303)


# ─── SYNC SETTINGS ───────────────────────────────────────────────────────────

@app.get("/sync-settings", response_class=HTMLResponse)
async def sync_settings(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    configs = {cfg.project_type: cfg for cfg in db.query(models.SyncConfig).all()}
    section_types = ["Реконструкция", "Констракшн", "КСО"]
    return templates.TemplateResponse("sync_settings.html", {
        "request": request, "user": user,
        "configs": configs, "section_types": section_types,
    })


@app.post("/sync-settings/save")
async def save_sync_settings(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    form = await request.form()
    section_types = ["Реконструкция", "Констракшн", "КСО"]
    for ptype in section_types:
        key = ptype.lower().replace(" ", "_")
        file_path = str(form.get(f"path_{key}", "")).strip()
        auto_sync = bool(form.get(f"auto_{key}"))
        interval = int(form.get(f"interval_{key}", 60) or 60)

        cfg = db.query(models.SyncConfig).filter(
            models.SyncConfig.project_type == ptype).first()
        if cfg:
            cfg.file_path = file_path
            cfg.auto_sync = auto_sync
            cfg.sync_interval_minutes = interval
        else:
            db.add(models.SyncConfig(
                project_type=ptype, file_path=file_path,
                auto_sync=auto_sync, sync_interval_minutes=interval))
    db.commit()
    return RedirectResponse("/sync-settings?saved=1", status_code=303)


@app.post("/sync-settings/run-now")
async def run_sync_now(request: Request, db: Session = Depends(get_db),
                       project_type: str = Form(...)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    cfg = db.query(models.SyncConfig).filter(
        models.SyncConfig.project_type == project_type).first()
    if cfg and cfg.file_path:
        path = Path(cfg.file_path)
        if path.exists():
            content = path.read_bytes()
            result = parse_excel_file(content, project_type, None, db)
            cfg.last_synced = datetime.utcnow()
            cfg.last_status = f"OK: создано {result['created']}, обновлено {result['updated']}"
            db.commit()
        else:
            cfg.last_status = f"Файл не найден: {cfg.file_path}"
            db.commit()
    return RedirectResponse("/sync-settings?ran=1", status_code=303)


# ─── AI ASSISTANT ─────────────────────────────────────────────────────────────

@app.get("/ai", response_class=HTMLResponse)
async def ai_page(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    return templates.TemplateResponse("ai_chat.html", {
        "request": request, "user": user,
    })


@app.post("/api/ai/chat")
async def ai_chat(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return {"error": "Не авторизован"}

    body = await request.json()
    user_message = body.get("message", "").strip()
    if not user_message:
        return {"error": "Пустое сообщение"}

    provider = body.get("provider", "groq")
    history = body.get("history", [])  # [{role, content}, ...]

    # ── Богатый контекст из БД ────────────────────────────────────────────────
    today = date.today()
    all_projects = db.query(models.Project).all()
    active = [p for p in all_projects if p.status == "Активный"]
    done   = [p for p in all_projects if p.status == "Завершён"]

    urgent = [p for p in active if p.end_date and 0 <= (p.end_date - today).days <= 7
              and not (p.opening_date and p.opening_date <= today)]
    overdue_p = [p for p in active if p.end_date and p.end_date < today
                 and not (p.opening_date and p.opening_date <= today)]

    overdue_tasks = db.query(models.Task).filter(
        models.Task.status != "Завершена",
        models.Task.deadline != None,
        models.Task.deadline < today).limit(15).all()
    open_tasks = db.query(models.Task).filter(
        models.Task.status != "Завершена").count()
    managers = db.query(models.Manager).all()

    ctx = f"""Сегодня: {today.strftime('%d.%m.%Y')}

ПРОЕКТЫ:
- Всего: {len(all_projects)} | Активных: {len(active)} | Завершённых: {len(done)}
- Реконструкции активных: {sum(1 for p in active if p.project_type=='Реконструкция')}
- Констракшн активных: {sum(1 for p in active if p.project_type=='Констракшн')}

СРОЧНЫЕ (дедлайн ≤7 дней): {len(urgent)} объектов
{chr(10).join(f'  • {p.name} | {p.manager.name if p.manager else "—"} | {p.end_date.strftime("%d.%m.%Y")}' for p in urgent[:8])}

ПРОСРОЧЕННЫЕ проекты: {len(overdue_p)}
{chr(10).join(f'  • {p.name} | {p.manager.name if p.manager else "—"} | просрочено на {(today-p.end_date).days} дн' for p in overdue_p[:5])}

ЗАДАЧИ: открытых {open_tasks}, просроченных {len(overdue_tasks)}
{chr(10).join(f'  • {t.title} | {t.assignee.name if t.assignee else "—"} | -{(today-t.deadline).days} дн' for t in overdue_tasks[:5])}

МЕНЕДЖЕРЫ ({len(managers)} чел):
{chr(10).join(f'  • {m.name}{"  [руководитель]" if m.is_leader else ""}: активных проектов {sum(1 for p in m.projects if p.status=="Активный")}' for m in managers)}
"""

    system_prompt = (
        "Ты — ИИ-ассистент системы управления проектами компании ЛЕНТА (сеть гипермаркетов России). "
        "Помогаешь команде из 9 человек: руководитель Гаврин Игорь + 8 менеджеров. "
        "Работаешь с двумя типами проектов: Реконструкции магазинов и Констракшн (новое строительство). "
        "Анализируешь статусы, дедлайны, риски. Даёшь конкретные рекомендации. "
        "Отвечай чётко, структурированно, на русском языке. Используй markdown: **жирный**, списки, заголовки. "
        f"\n\n=== ТЕКУЩИЕ ДАННЫЕ СИСТЕМЫ ===\n{ctx}"
    )

    # ── Собираем историю сообщений ────────────────────────────────────────────
    msg_history = []
    for h in history[:-1]:
        role = h.get("role", "user")
        if role in ("user", "assistant"):
            msg_history.append({"role": role, "content": h.get("content", "")})
    msg_history.append({"role": "user", "content": user_message})

    reply = ""
    try:
        if provider == "claude":
            import anthropic as ant
            api_key = os.getenv("ANTHROPIC_API_KEY", "")
            if not api_key:
                return {"reply": "⚠️ ANTHROPIC_API_KEY не задан в Railway Variables."}
            client_ant = ant.Anthropic(api_key=api_key)
            resp = client_ant.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=2048,
                system=system_prompt,
                messages=msg_history,
            )
            reply = resp.content[0].text

        elif provider == "groq":
            from openai import OpenAI
            api_key = os.getenv("GROQ_API_KEY", "")
            if not api_key:
                return {"reply": "⚠️ GROQ_API_KEY не задан. Получить бесплатно: console.groq.com"}
            client = OpenAI(api_key=api_key, base_url="https://api.groq.com/openai/v1")
            resp = client.chat.completions.create(
                model="llama-3.3-70b-versatile",
                max_tokens=2048,
                messages=[{"role": "system", "content": system_prompt}] + msg_history,
            )
            reply = resp.choices[0].message.content

        elif provider == "deepseek":
            from openai import OpenAI
            api_key = os.getenv("DEEPSEEK_API_KEY", "")
            if not api_key:
                return {"reply": "⚠️ DEEPSEEK_API_KEY не задан в Railway Variables."}
            client = OpenAI(api_key=api_key, base_url="https://api.deepseek.com")
            resp = client.chat.completions.create(
                model="deepseek-chat",
                max_tokens=2048,
                messages=[{"role": "system", "content": system_prompt}] + msg_history,
            )
            reply = resp.choices[0].message.content

    except Exception as e:
        reply = f"⚠️ Ошибка ({provider}): {str(e)[:300]}"

    # Сохраняем в БД
    if reply and not reply.startswith("⚠️"):
        user_name = user.get("display_name", "")
        db.add(models.AiChatMessage(user_name=user_name, role="user",
                                    text=user_message, provider=provider))
        db.add(models.AiChatMessage(user_name=user_name, role="assistant",
                                    text=reply, provider=provider))
        db.commit()

    return {"reply": reply}


@app.get("/api/ai/history")
async def ai_history(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return {"messages": []}
    user_name = user.get("display_name", "")
    msgs = db.query(models.AiChatMessage).filter(
        models.AiChatMessage.user_name == user_name
    ).order_by(models.AiChatMessage.id.desc()).limit(100).all()
    msgs.reverse()
    return {"messages": [
        {"role": m.role, "text": m.text,
         "time": m.created_at.strftime("%H:%M")}
        for m in msgs
    ]}


@app.post("/api/ai/clear-history")
async def ai_clear_history(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return {"error": "Не авторизован"}
    user_name = user.get("display_name", "")
    db.query(models.AiChatMessage).filter(
        models.AiChatMessage.user_name == user_name).delete()
    db.commit()
    return {"ok": True}


@app.post("/api/ai/check-excel")
async def ai_check_excel(request: Request, db: Session = Depends(get_db),
                          file: UploadFile = File(...)):
    user = get_current_user(request)
    if not user:
        return {"error": "Не авторизован"}

    api_key = os.getenv("ANTHROPIC_API_KEY", "")
    if not api_key:
        return {"report": "ANTHROPIC_API_KEY не задан."}

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
        issues = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        issues.append(f"Лист '{sheet_name}', ячейка {cell.coordinate}: {cell.value[:60]}")
        formula_text = "\n".join(issues[:50]) if issues else "Формул не обнаружено."
    except Exception as e:
        return {"report": f"Ошибка чтения файла: {e}"}

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=2048,
            system=(
                "Ты — эксперт по Excel и управлению строительными проектами. "
                "Проверяй формулы в таблицах на корректность. "
                "Отвечай на русском языке."
            ),
            messages=[{"role": "user", "content":
                f"Проверь формулы в Excel-файле и выяви проблемы:\n{formula_text}\n"
                "Составь краткий отчёт: какие формулы потенциально проблемные и почему."}],
        )
        report = response.content[0].text
    except Exception as e:
        report = f"Ошибка ИИ: {str(e)[:200]}"

    return {"report": report, "formulas_found": len(issues)}


# ─── ОТЧЁТЫ ВПК ──────────────────────────────────────────────────────────────

@app.get("/reports", response_class=HTMLResponse)
async def reports_page(request: Request, db: Session = Depends(get_db),
                       vpk_type: str = "", manager_name: str = ""):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)

    q = db.query(models.VpkReport).order_by(models.VpkReport.submitted_at.desc())
    if vpk_type in ("1", "2"):
        q = q.filter(models.VpkReport.vpk_type == int(vpk_type))
    if manager_name:
        q = q.filter(models.VpkReport.submitted_by == manager_name)
    reports = q.limit(100).all()

    # Список авторов для фильтра
    authors = db.query(models.VpkReport.submitted_by).distinct().all()
    authors = sorted({a[0] for a in authors if a[0]})

    return templates.TemplateResponse("reports.html", {
        "request": request, "user": user,
        "reports": reports, "authors": authors,
        "filter_vpk_type": vpk_type, "filter_manager": manager_name,
    })


@app.get("/reports/export")
async def reports_export(request: Request, db: Session = Depends(get_db),
                         date_from: str = "", date_to: str = "",
                         vpk_type: str = ""):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)

    q = db.query(models.VpkReport).order_by(models.VpkReport.submitted_at.desc())
    if vpk_type in ("1", "2"):
        q = q.filter(models.VpkReport.vpk_type == int(vpk_type))
    if date_from:
        try:
            q = q.filter(models.VpkReport.submitted_at >=
                         datetime.strptime(date_from, "%Y-%m-%d"))
        except ValueError:
            pass
    if date_to:
        try:
            q = q.filter(models.VpkReport.submitted_at <
                         datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1))
        except ValueError:
            pass
    reports = q.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёты ВПК"

    hfill  = PatternFill(start_color="1A5C22", end_color="1A5C22", fill_type="solid")
    hfont  = Font(color="FFFFFF", bold=True, size=12)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap   = Alignment(vertical="center", wrap_text=True)

    headers    = ["Менеджер", "Номер ТК", "Критерий ВПК", "Комментарий"]
    col_widths = [25, 12, 60, 50]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont; c.alignment = center
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"

    row_num = 2
    for r in reports:
        tk  = r.project.tk_number if r.project else "—"
        mgr = r.project.manager.name if (r.project and r.project.manager) else "—"

        for item in r.items:
            ws.cell(row=row_num, column=1, value=mgr).alignment = wrap
            ws.cell(row=row_num, column=2, value=tk).alignment  = center
            ws.cell(row=row_num, column=3, value=item.criterion_name).alignment = wrap
            ws.cell(row=row_num, column=4, value=item.comment or "").alignment  = wrap
            ws.row_dimensions[row_num].height = 18
            row_num += 1

        row_num += 1  # пустая строка между отчётами

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    period = f"{date_from or 'all'}_{date_to or 'now'}"
    fname = f"vpk_reports_{period}.xlsx"
    return StreamingResponse(output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


@app.post("/reports/clear-all")
async def clear_all_reports(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user or not user.get("is_admin"):
        return RedirectResponse("/reports", status_code=302)
    db.query(models.VpkReportItem).delete()
    db.query(models.VpkReport).delete()
    db.commit()
    return RedirectResponse("/reports", status_code=303)


# ─── ЧАТ ─────────────────────────────────────────────────────────────────────

@app.get("/chat", response_class=HTMLResponse)
async def chat_page(request: Request, db: Session = Depends(get_db),
                    partner: str = ""):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    _chat_order = [
        "Месмер Денис", "Митько Роберт", "Ловчиков Александр",
        "Валеев Борис", "Косило Сергей", "Студеникин Сергей",
        "Хачатурова Жанна", "Шевченко Наталья",
    ]
    managers = db.query(models.Manager).all()
    managers.sort(key=lambda m: (
        0 if m.is_leader else 1,
        _chat_order.index(m.name) if m.name in _chat_order else 99
    ))
    my_name = user.get("display_name", "")

    # Непрочитанные по каждому собеседнику
    unread_by = {}
    unread_msgs = db.query(models.ChatMessage).filter(
        models.ChatMessage.receiver_name == my_name,
        models.ChatMessage.is_read == False,
    ).all()
    for m in unread_msgs:
        unread_by[m.sender_name] = unread_by.get(m.sender_name, 0) + 1

    now = datetime.utcnow()
    online_set = {
        name for name, ts in ONLINE_USERS.items()
        if (now - ts).total_seconds() < ONLINE_TIMEOUT
    }

    return templates.TemplateResponse("chat.html", {
        "request": request, "user": user,
        "managers": managers, "partner": partner,
        "my_name": my_name, "unread_by": unread_by,
        "online_set": online_set,
    })


@app.get("/api/task-notifications")
async def get_task_notifications(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return {"notifications": [], "unread": 0}
    my_name = user.get("display_name", "")
    notifs = db.query(models.TaskNotification).filter(
        models.TaskNotification.recipient_name == my_name,
        models.TaskNotification.is_read == False,
    ).order_by(models.TaskNotification.created_at.desc()).limit(10).all()
    # Помечаем как прочитанные
    for n in notifs:
        n.is_read = True
    db.commit()
    return {
        "notifications": [{"id": n.id, "message": n.message, "task_id": n.task_id} for n in notifs],
        "unread": len(notifs),
    }


@app.post("/api/ping")
async def ping(request: Request):
    user = get_current_user(request)
    if user:
        ONLINE_USERS[user.get("display_name", "")] = datetime.utcnow()
    return {"ok": True}


@app.get("/api/online")
async def get_online(request: Request):
    user = get_current_user(request)
    if not user:
        return {"online": []}
    now = datetime.utcnow()
    online = [
        name for name, ts in ONLINE_USERS.items()
        if (now - ts).total_seconds() < ONLINE_TIMEOUT
    ]
    return {"online": online}


@app.get("/api/chat/messages")
async def chat_messages(request: Request, db: Session = Depends(get_db),
                        partner: str = "", since_id: int = 0):
    user = get_current_user(request)
    if not user:
        return {"messages": []}
    my_name = user.get("display_name", "")

    if partner == "":
        # Общий чат
        q = db.query(models.ChatMessage).filter(
            models.ChatMessage.receiver_name == "",
            models.ChatMessage.id > since_id,
        )
    else:
        # Личные — между мной и партнёром
        from sqlalchemy import or_, and_
        q = db.query(models.ChatMessage).filter(
            models.ChatMessage.id > since_id,
            or_(
                and_(models.ChatMessage.sender_name == my_name,
                     models.ChatMessage.receiver_name == partner),
                and_(models.ChatMessage.sender_name == partner,
                     models.ChatMessage.receiver_name == my_name),
            )
        )
        # Помечаем прочитанными входящие
        db.query(models.ChatMessage).filter(
            models.ChatMessage.sender_name == partner,
            models.ChatMessage.receiver_name == my_name,
            models.ChatMessage.is_read == False,
        ).update({"is_read": True})
        db.commit()

    msgs = q.order_by(models.ChatMessage.id).limit(200).all()
    return {"messages": [
        {"id": m.id, "sender": m.sender_name,
         "text": m.text,
         "photo": m.photo_path or "",
         "time": m.created_at.strftime("%H:%M"),
         "mine": m.sender_name == my_name}
        for m in msgs
    ]}


@app.post("/api/chat/send")
async def chat_send(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return {"error": "Не авторизован"}
    data = await request.json()
    text = data.get("text", "").strip()
    if not text:
        return {"error": "Пустое сообщение"}
    partner = data.get("partner", "")
    msg = models.ChatMessage(
        sender_name=user.get("display_name", ""),
        receiver_name=partner,
        text=text,
    )
    db.add(msg)
    db.commit()
    db.refresh(msg)
    return {"id": msg.id, "time": msg.created_at.strftime("%H:%M")}


@app.post("/api/chat/send-photo")
async def chat_send_photo(request: Request, db: Session = Depends(get_db),
                          file: UploadFile = File(...),
                          partner: str = Form(""),
                          text: str = Form("")):
    user = get_current_user(request)
    if not user:
        return {"error": "Не авторизован"}
    photo_dir = Path("static/uploads/chat")
    photo_dir.mkdir(parents=True, exist_ok=True)
    ext = Path(file.filename).suffix.lower() if file.filename else ".jpg"
    if ext not in ('.jpg', '.jpeg', '.png', '.webp', '.gif'):
        ext = '.jpg'
    fname = f"{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{user.get('display_name','u').replace(' ','_')}{ext}"
    (photo_dir / fname).write_bytes(await file.read())
    msg = models.ChatMessage(
        sender_name=user.get("display_name", ""),
        receiver_name=partner,
        text=text or "",
        photo_path=f"uploads/chat/{fname}",
    )
    db.add(msg)
    db.commit()
    db.refresh(msg)
    return {"id": msg.id, "time": msg.created_at.strftime("%H:%M"),
            "photo": msg.photo_path}


@app.get("/api/chat/unread")
async def chat_unread(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return {"total": 0}
    my_name = user.get("display_name", "")
    total = db.query(models.ChatMessage).filter(
        models.ChatMessage.receiver_name == my_name,
        models.ChatMessage.is_read == False,
    ).count()
    return {"total": total}
