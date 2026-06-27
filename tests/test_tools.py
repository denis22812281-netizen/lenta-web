"""Tests for tools service: Excel builders correctness and templates DB."""
import io

import pytest
from openpyxl import load_workbook

from services.tools_service import (
    BUILDERS,
    build_comparison,
    build_gantt,
    build_table,
)


def _load(wb_or_bytes):
    if isinstance(wb_or_bytes, bytes):
        return load_workbook(io.BytesIO(wb_or_bytes))
    buf = io.BytesIO()
    wb_or_bytes.save(buf)
    buf.seek(0)
    return load_workbook(buf)


# ── build_table ────────────────────────────────────────────────────────────────

def test_build_table_basic():
    wb = build_table({"title": "Тест", "headers": ["A", "B"], "rows": [["1", "x"], ["2", "y"]]})
    ws = wb.active
    assert ws.cell(1, 1).value == "Тест"
    assert ws.cell(2, 1).value == "A"
    assert ws.cell(3, 1).value == "1"
    assert ws.cell(4, 2).value == "y"


def test_build_table_empty_rows():
    wb = build_table({"title": "Пусто", "headers": ["X"], "rows": []})
    ws = wb.active
    assert ws.cell(2, 1).value == "X"


def test_build_table_fixes_row_numbers():
    wb = build_table({"title": "T", "headers": ["№", "Name"], "rows": [["?", "Иван"], ["?", "Пётр"]]})
    ws = wb.active
    assert ws.cell(3, 1).value == "1"
    assert ws.cell(4, 1).value == "2"


# ── build_gantt ────────────────────────────────────────────────────────────────

def test_build_gantt_creates_calendar_columns():
    data = {
        "title": "График", "year": 2026,
        "tasks": [
            {"id": 1, "name": "Задача 1", "object": "ТК 001", "start": "01.07", "end": "10.07"},
            {"id": 2, "name": "• Веха", "object": "", "start": "10.07", "end": "10.07"},
        ]
    }
    wb = build_gantt(data)
    ws = wb.active
    assert ws.title == "Ганта"
    assert ws.cell(1, 1).value == "График"
    # Должно быть >= 5 фиксированных + 10 дней (01-10 июля)
    assert ws.max_column >= 15


def test_build_gantt_falls_back_to_table_on_bad_dates():
    data = {"title": "Без дат", "year": 2026, "tasks": [
        {"id": 1, "name": "Задача", "start": "BAD", "end": "DATES"}
    ]}
    wb = build_gantt(data)
    ws = wb.active
    assert ws.title == "Данные"  # fallback to table


# ── build_comparison ──────────────────────────────────────────────────────────

def test_build_comparison_creates_three_sheets():
    d1 = {"title": "T1", "headers": ["ID", "Val"], "rows": [["1", "a"], ["2", "b"]]}
    d2 = {"title": "T2", "headers": ["ID", "Val"], "rows": [["1", "a"], ["3", "c"]]}
    wb = build_comparison(d1, d2)
    assert "Источник 1" in wb.sheetnames
    assert "Источник 2" in wb.sheetnames
    assert "Разница" in wb.sheetnames


def test_build_comparison_diff_marks_changes():
    d1 = {"headers": ["ID", "V"], "rows": [["1", "old"]]}
    d2 = {"headers": ["ID", "V"], "rows": [["1", "new"]]}
    wb = build_comparison(d1, d2)
    diff_ws = wb["Разница"]
    # Row 2 (key "1") has changed value in col 2 — should have yellow fill
    changed_cell = diff_ws.cell(2, 2)
    assert changed_cell.fill.fgColor.rgb in ("FFFACC", "00FFFACC")


# ── BUILDERS dict ──────────────────────────────────────────────────────────────

def test_builders_dict_has_all_types():
    assert set(BUILDERS.keys()) == {"table", "chart", "diagram", "gantt"}


# ── Templates API via auth_client ──────────────────────────────────────────────

def test_templates_list_empty_initially(auth_client):
    resp = auth_client.get("/api/tools/templates")
    assert resp.status_code == 200
    assert isinstance(resp.json()["templates"], list)


def test_templates_save_and_list(auth_client):
    resp = auth_client.post("/api/tools/templates", json={
        "name": "Мой шаблон", "output_type": "table", "mode": "photo"
    })
    assert resp.status_code == 200
    assert resp.json()["ok"] is True
    uid = resp.json()["id"]

    resp = auth_client.get("/api/tools/templates")
    names = [t["name"] for t in resp.json()["templates"]]
    assert "Мой шаблон" in names
    return uid


def test_templates_upsert_duplicate_name(auth_client):
    auth_client.post("/api/tools/templates", json={"name": "Дубль", "output_type": "table", "mode": "photo"})
    resp = auth_client.post("/api/tools/templates", json={"name": "Дубль", "output_type": "gantt", "mode": "text"})
    assert resp.status_code == 200

    resp = auth_client.get("/api/tools/templates")
    dups = [t for t in resp.json()["templates"] if t["name"] == "Дубль"]
    assert len(dups) == 1
    assert dups[0]["output_type"] == "gantt"


def test_templates_delete(auth_client):
    r = auth_client.post("/api/tools/templates", json={"name": "Удаляемый", "output_type": "chart", "mode": "photo"})
    uid = r.json()["id"]
    resp = auth_client.delete(f"/api/tools/templates/{uid}")
    assert resp.status_code == 200
    assert resp.json()["ok"] is True

    resp = auth_client.get("/api/tools/templates")
    names = [t["name"] for t in resp.json()["templates"]]
    assert "Удаляемый" not in names


def test_templates_requires_admin(client):
    # require_admin → require_login → 302 to /login for unauthenticated browser requests
    resp = client.get("/api/tools/templates", follow_redirects=False)
    assert resp.status_code in (302, 401, 403)
