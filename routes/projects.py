"""All project-related routes: list, detail, create, update, delete, stages, excel import/export."""
import io
from datetime import datetime, date

import os
from pathlib import Path

from fastapi import APIRouter, BackgroundTasks, Request, Form, Depends, UploadFile, File, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

import models
from database import get_db
from deps import templates, require_login, limiter
from config import PROJECT_TYPES, STATUSES, STAGE_NAMES
from services.excel_import import parse_excel_file, import_reconstruct_excel, import_construction_excel
from services.cloud_storage import upload_photo
from services.email_service import notify_opening_photos
from utils.files import read_limited

_OPENING_EMAIL = os.getenv("NOTIFY_PRECHECK_EMAIL", "").strip()

router = APIRouter()


# ─── Список всех проектов ────────────────────────────────────────────────────

_PER_PAGE = 25


@router.get("/projects", response_class=HTMLResponse)
async def projects_list(request: Request, db: Session = Depends(get_db),
                        user: dict = Depends(require_login),
                        manager_id: str = None, status: str = None,
                        project_type: str = None, search: str = None,
                        page: int = 1):
    q = db.query(models.Project)
    if manager_id and str(manager_id).isdigit():
        q = q.filter(models.Project.manager_id == int(manager_id))
    if status:
        q = q.filter(models.Project.status == status)
    if project_type:
        q = q.filter(models.Project.project_type == project_type)
    if search:
        like = f"%{search}%"
        q = q.filter(
            models.Project.tk_number.ilike(like) |
            models.Project.name.ilike(like) |
            models.Project.city.ilike(like)
        )
    total = q.count()
    total_pages = max(1, (total + _PER_PAGE - 1) // _PER_PAGE)
    page = max(1, min(page, total_pages))
    projects = (q.order_by(models.Project.end_date.nullslast())
                 .offset((page - 1) * _PER_PAGE).limit(_PER_PAGE).all())
    managers = db.query(models.Manager).all()
    return templates.TemplateResponse("projects.html", {
        "request": request, "user": user,
        "projects": projects, "managers": managers,
        "project_types": PROJECT_TYPES, "statuses": STATUSES, "today": date.today(),
        "filter_manager_id": manager_id, "filter_status": status,
        "filter_type": project_type, "search": search or "",
        "page": page, "total": total, "total_pages": total_pages,
    })


@router.post("/projects/create")
async def create_project(request: Request, db: Session = Depends(get_db),
                         user: dict = Depends(require_login),
                         name: str = Form(...), tk_number: str = Form(""),
                         city: str = Form(""), project_type: str = Form(""),
                         manager_id: str = Form(""), status: str = Form("Активный"),
                         stage: str = Form(""), start_date: str = Form(""),
                         end_date: str = Form(""), description: str = Form(""),
                         budget: str = Form("")):
    db.add(models.Project(
        name=name, tk_number=tk_number, city=city, project_type=project_type,
        manager_id=int(manager_id) if manager_id else None,
        status=status, stage=stage, description=description,
        start_date=datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None,
        end_date=datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None,
        budget=float(budget.replace(",", ".")) if budget else None,
    ))
    db.commit()
    return RedirectResponse("/projects", status_code=303)


@router.get("/projects/{project_id}", response_class=HTMLResponse)
async def project_detail(request: Request, project_id: int, db: Session = Depends(get_db),
                         user: dict = Depends(require_login)):
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404)
    managers = db.query(models.Manager).all()
    return templates.TemplateResponse("project_detail.html", {
        "request": request, "user": user,
        "project": project, "managers": managers,
        "project_types": PROJECT_TYPES, "statuses": STATUSES,
        "stage_names": STAGE_NAMES, "today": date.today(),
    })


def _mgr_name(db, mid):
    if not mid:
        return ""
    m = db.query(models.Manager).filter(models.Manager.id == mid).first()
    return m.name if m else str(mid)


def _track_changes(db, project, user, new_vals: dict):
    """Записывает изменения полей в project_history."""
    LABELS = {
        "name":         "Название",
        "tk_number":    "Номер ТК",
        "city":         "Город",
        "project_type": "Тип",
        "status":       "Статус",
        "stage":        "Этап",
        "manager_id":   "Менеджер",
        "start_date":   "Дата начала",
        "end_date":     "Дата окончания",
        "description":  "Описание",
        "budget":       "Бюджет",
    }
    author = user.get("display_name", "")
    for field, label in LABELS.items():
        old = getattr(project, field, None)
        new = new_vals.get(field)
        old_s = str(old) if old is not None else ""
        new_s = str(new) if new is not None else ""
        if field == "manager_id":
            old_s = _mgr_name(db, old)
            new_s = _mgr_name(db, new)
        if old_s != new_s:
            db.add(models.ProjectHistory(
                project_id=project.id,
                changed_by=author,
                field_label=label,
                old_value=old_s,
                new_value=new_s,
            ))


@router.post("/projects/{project_id}/update")
async def update_project(project_id: int, request: Request, db: Session = Depends(get_db),
                         user: dict = Depends(require_login),
                         name: str = Form(...), tk_number: str = Form(""),
                         city: str = Form(""), project_type: str = Form(""),
                         manager_id: str = Form(""), status: str = Form("Активный"),
                         stage: str = Form(""), start_date: str = Form(""),
                         end_date: str = Form(""), description: str = Form(""),
                         budget: str = Form("")):
    p = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not p:
        raise HTTPException(status_code=404)
    new_manager_id = int(manager_id) if manager_id else None
    new_start = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
    new_end = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None
    new_budget = float(budget.replace(",", ".")) if budget else None
    _track_changes(db, p, user, {
        "name": name, "tk_number": tk_number, "city": city,
        "project_type": project_type, "status": status, "stage": stage,
        "manager_id": new_manager_id, "start_date": new_start,
        "end_date": new_end, "description": description, "budget": new_budget,
    })
    p.name = name; p.tk_number = tk_number; p.city = city
    p.project_type = project_type; p.status = status; p.stage = stage
    p.manager_id = new_manager_id
    p.start_date = new_start
    p.end_date = new_end
    p.description = description
    p.budget = new_budget
    db.commit()
    return RedirectResponse(f"/projects/{project_id}", status_code=303)


@router.post("/projects/{project_id}/delete")
async def delete_project(project_id: int, request: Request, db: Session = Depends(get_db),
                         user: dict = Depends(require_login)):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403)
    p = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not p:
        raise HTTPException(status_code=404)
    db.delete(p)
    db.commit()
    return RedirectResponse("/projects", status_code=303)


@router.post("/api/projects/{project_id}/field")
async def update_project_field(
    project_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: dict = Depends(require_login),
):
    """Inline-редактирование одного поля проекта (JSON API)."""
    from fastapi.responses import JSONResponse
    _ALLOWED = {"name", "city", "status", "stage", "end_date", "start_date", "description", "budget", "tk_number"}
    data = await request.json()
    field = str(data.get("field", "")).strip()
    value = data.get("value", "")

    if field not in _ALLOWED:
        return JSONResponse({"ok": False, "error": "field not allowed"}, status_code=400)

    p = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not p:
        return JSONResponse({"ok": False}, status_code=404)

    if field in ("end_date", "start_date"):
        try:
            parsed = datetime.strptime(str(value), "%Y-%m-%d").date() if value else None
        except ValueError:
            return JSONResponse({"ok": False, "error": "invalid date"}, status_code=400)
        _track_changes(db, p, user, {field: parsed})
        setattr(p, field, parsed)
    elif field == "budget":
        try:
            parsed_b = float(str(value).replace(",", ".")) if value else None
        except ValueError:
            return JSONResponse({"ok": False, "error": "invalid budget"}, status_code=400)
        _track_changes(db, p, user, {field: parsed_b})
        setattr(p, field, parsed_b)
    else:
        _track_changes(db, p, user, {field: value or None})
        setattr(p, field, value or None)

    db.commit()
    return JSONResponse({"ok": True})


# ─── Комментарии к проекту ───────────────────────────────────────────────────

@router.post("/projects/{project_id}/comments/add")
async def add_comment(project_id: int, request: Request, db: Session = Depends(get_db),
                      user: dict = Depends(require_login),
                      text: str = Form(...)):
    text = text.strip()
    if text:
        comment = models.ProjectComment(
            project_id=project_id,
            author_name=user.get("display_name", ""),
            text=text,
        )
        db.add(comment)
        db.commit()
    return RedirectResponse(f"/projects/{project_id}#comments", status_code=303)


# ─── Фото открытия ───────────────────────────────────────────────────────────

@router.post("/projects/{project_id}/opening-photos")
async def upload_opening_photos(project_id: int, request: Request,
                                background_tasks: BackgroundTasks,
                                db: Session = Depends(get_db),
                                user: dict = Depends(require_login)):
    proj = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not proj:
        raise HTTPException(status_code=404)

    form = await request.form()
    files = form.getlist("photos")

    photo_urls = []
    for i, photo_file in enumerate(files):
        if not (hasattr(photo_file, "filename") and photo_file.filename):
            continue
        ext = Path(photo_file.filename).suffix.lower() or ".jpg"
        fname = f"open_{project_id}_{len(proj.opening_photos) + i + 1}{ext}"
        folder = f"opening/tk-{proj.tk_number or project_id}"
        url = upload_photo(await photo_file.read(), folder, fname)
        photo_urls.append(url)
        db.add(models.OpeningPhoto(
            project_id=project_id,
            photo_path=url,
            uploaded_by=user.get("display_name", ""),
        ))

    db.commit()

    if photo_urls:
        submitter = user.get("display_name", "")
        to_email = _OPENING_EMAIL
        if not to_email:
            mgr = db.query(models.Manager).filter(
                models.Manager.name == submitter,
                models.Manager.email.isnot(None),
                models.Manager.email != "",
            ).first()
            if mgr:
                to_email = mgr.email
        if to_email:
            background_tasks.add_task(
                notify_opening_photos, to_email,
                proj.tk_number or str(project_id),
                proj.city or proj.address or "",
                submitter, photo_urls,
            )

    count = len(photo_urls)
    return RedirectResponse(
        f"/projects/{project_id}?msg=Загружено+{count}+фото+открытия#opening-photos",
        status_code=303,
    )


@router.post("/projects/{project_id}/comments/{comment_id}/delete")
async def delete_comment(project_id: int, comment_id: int, request: Request,
                         db: Session = Depends(get_db),
                         user: dict = Depends(require_login)):
    c = db.query(models.ProjectComment).filter(models.ProjectComment.id == comment_id).first()
    if c and (c.author_name == user.get("display_name") or user.get("is_admin")):
        db.delete(c)
        db.commit()
    return RedirectResponse(f"/projects/{project_id}#comments", status_code=303)


@router.post("/projects/{project_id}/delete")
async def delete_project(project_id: int, request: Request, db: Session = Depends(get_db),
                         user: dict = Depends(require_login)):
    p = db.query(models.Project).filter(models.Project.id == project_id).first()
    if p:
        db.delete(p)
        db.commit()
    return RedirectResponse("/projects", status_code=303)


# ─── Этапы проекта ────────────────────────────────────────────────────────────

@router.post("/projects/{project_id}/stages/add")
async def add_stage(project_id: int, request: Request, db: Session = Depends(get_db),
                    user: dict = Depends(require_login),
                    name: str = Form(...), start_date: str = Form(""),
                    end_date: str = Form(""), stage_status: str = Form("Запланировано")):
    order = db.query(models.ProjectStage).filter(
        models.ProjectStage.project_id == project_id).count()
    db.add(models.ProjectStage(
        project_id=project_id, name=name, status=stage_status,
        start_date=datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None,
        end_date=datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None,
        order=order,
    ))
    db.commit()
    return RedirectResponse(f"/projects/{project_id}", status_code=303)


@router.post("/stages/{stage_id}/delete")
async def delete_stage(stage_id: int, request: Request, db: Session = Depends(get_db),
                       user: dict = Depends(require_login)):
    s = db.query(models.ProjectStage).filter(models.ProjectStage.id == stage_id).first()
    project_id = s.project_id if s else None
    if s:
        db.delete(s)
        db.commit()
    return RedirectResponse(f"/projects/{project_id}", status_code=303)


@router.post("/api/stages/{stage_id}/dates")
async def update_stage_dates(stage_id: int, request: Request, db: Session = Depends(get_db),
                             user: dict = Depends(require_login)):
    """Обновление дат этапа из интерактивного Gantt (AJAX)."""
    from fastapi.responses import JSONResponse
    body = await request.json()
    s = db.query(models.ProjectStage).filter(models.ProjectStage.id == stage_id).first()
    if not s:
        return JSONResponse({"ok": False, "error": "not found"}, status_code=404)
    try:
        if body.get("start"):
            s.start_date = datetime.strptime(body["start"][:10], "%Y-%m-%d").date()
        if body.get("end"):
            s.end_date = datetime.strptime(body["end"][:10], "%Y-%m-%d").date()
        db.commit()
        return JSONResponse({"ok": True})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=400)


# ─── Вложения к проекту (файлы и фото) ───────────────────────────────────────

_ALLOWED_MIME = {
    "image/jpeg", "image/png", "image/webp", "image/gif",
    "application/pdf",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/msword",
    "text/plain",
}
_MAX_ATTACH_SIZE = 20 * 1024 * 1024  # 20 МБ


def _file_type(content_type: str, filename: str) -> str:
    if content_type.startswith("image/"):
        return "image"
    ext = Path(filename).suffix.lower()
    if ext in (".pdf",):
        return "pdf"
    if ext in (".xlsx", ".xls", ".csv"):
        return "xls"
    if ext in (".docx", ".doc"):
        return "doc"
    return "file"


@router.post("/projects/{project_id}/attachments/upload")
async def upload_attachment(project_id: int, request: Request, db: Session = Depends(get_db),
                            user: dict = Depends(require_login),
                            file: UploadFile = File(...)):
    proj = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not proj:
        raise HTTPException(status_code=404)
    content = await file.read()
    if len(content) > _MAX_ATTACH_SIZE:
        return RedirectResponse(f"/projects/{project_id}?err=Файл+слишком+большой+(макс+20МБ)#attachments",
                                status_code=303)
    ct = file.content_type or ""
    ftype = _file_type(ct, file.filename or "")
    ext = Path(file.filename or "file").suffix.lower() or ".bin"
    import uuid as _uuid
    fname = f"{_uuid.uuid4().hex[:12]}{ext}"
    folder = f"attachments/proj-{project_id}"
    if ftype == "image":
        url = upload_photo(content, folder, fname)
    else:
        url = upload_file(content, folder, fname, original_name=file.filename or "")
    db.add(models.ProjectAttachment(
        project_id=project_id,
        original_name=file.filename or fname,
        file_url=url,
        file_type=ftype,
        file_size=len(content),
        uploaded_by=user.get("display_name", ""),
    ))
    db.commit()
    return RedirectResponse(f"/projects/{project_id}#attachments", status_code=303)


@router.post("/projects/{project_id}/attachments/{att_id}/delete")
async def delete_attachment(project_id: int, att_id: int, request: Request,
                            db: Session = Depends(get_db), user: dict = Depends(require_login)):
    att = db.query(models.ProjectAttachment).filter(
        models.ProjectAttachment.id == att_id,
        models.ProjectAttachment.project_id == project_id,
    ).first()
    if att and (att.uploaded_by == user.get("display_name") or user.get("is_admin")):
        from services.cloud_storage import delete_photo
        delete_photo(att.file_url)
        db.delete(att)
        db.commit()
    return RedirectResponse(f"/projects/{project_id}#attachments", status_code=303)


# ─── Создание проекта из секции (Реконструкции / Констракшн) ──────────────────

@router.post("/section/create-project")
async def section_create_project(request: Request, db: Session = Depends(get_db),
                                  user: dict = Depends(require_login),
                                  name: str = Form(...), tk_number: str = Form(""),
                                  city: str = Form(""), project_type: str = Form(...),
                                  manager_id: str = Form(""), status: str = Form("Активный"),
                                  stage: str = Form(""), start_date: str = Form(""),
                                  end_date: str = Form(""), description: str = Form(""),
                                  budget: str = Form(""), redirect_to: str = Form("/")):
    db.add(models.Project(
        name=name, tk_number=tk_number, city=city, project_type=project_type,
        manager_id=int(manager_id) if manager_id else None,
        status=status, stage=stage, description=description,
        start_date=datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None,
        end_date=datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None,
        budget=float(budget.replace(",", ".")) if budget else None,
    ))
    db.commit()
    return RedirectResponse(redirect_to, status_code=303)


# ─── Excel импорт/экспорт ─────────────────────────────────────────────────────

_MAX_EXCEL_BYTES = 10 * 1024 * 1024  # 10 MB


@router.post("/projects/import-excel")
@limiter.limit("5/minute")
async def import_excel(request: Request, db: Session = Depends(get_db),
                       user: dict = Depends(require_login),
                       file: UploadFile = File(...), manager_id: str = Form("")):
    try:
        content = await read_limited(file, _MAX_EXCEL_BYTES)
    except ValueError:
        return RedirectResponse("/projects?error=Файл слишком большой (макс 10 МБ)", status_code=303)
    try:
        result = parse_excel_file(content, "", int(manager_id) if manager_id else None, db)
    except Exception:
        return RedirectResponse("/projects?error=invalid_excel", status_code=303)
    return RedirectResponse("/projects", status_code=303)


@router.post("/import-excel-section")
@limiter.limit("5/minute")
async def import_excel_section(request: Request, db: Session = Depends(get_db),
                                user: dict = Depends(require_login),
                                file: UploadFile = File(...),
                                project_type: str = Form(""),
                                manager_id: str = Form(""),
                                redirect_to: str = Form("/")):
    try:
        content = await read_limited(file, _MAX_EXCEL_BYTES)
    except ValueError:
        return RedirectResponse(f"{redirect_to}?error=Файл слишком большой (макс 10 МБ)", status_code=303)
    try:
        if project_type == "Реконструкция":
            result = import_reconstruct_excel(content, db)
        elif project_type == "Констракшн":
            result = import_construction_excel(content, db)
        else:
            result = parse_excel_file(content, project_type,
                                      int(manager_id) if manager_id else None, db)
        return RedirectResponse(
            f"{redirect_to}?msg=created:{result['created']},updated:{result['updated']}",
            status_code=303)
    except Exception as e:
        return RedirectResponse(f"{redirect_to}?error={str(e)[:80]}", status_code=303)


@router.get("/api/export/projects-excel")
async def export_excel(db: Session = Depends(get_db), type: str = None,
                       user: dict = Depends(require_login)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Проекты"
    hfill  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hfont  = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")
    headers = ["№", "ТК №", "Название проекта", "Город", "Тип", "Менеджер",
               "Статус", "Этап", "Дата начала", "Дата окончания", "Бюджет, руб."]
    ws.row_dimensions[1].height = 20
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont; c.alignment = center

    q = db.query(models.Project)
    if type:
        q = q.filter(models.Project.project_type == type)
    today = date.today()
    for i, p in enumerate(q.all(), 2):
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=p.tk_number)
        ws.cell(row=i, column=3, value=p.name)
        ws.cell(row=i, column=4, value=p.city)
        ws.cell(row=i, column=5, value=p.project_type)
        ws.cell(row=i, column=6, value=p.manager.name if p.manager else "")
        ws.cell(row=i, column=7, value=p.status)
        ws.cell(row=i, column=8, value=p.stage)
        ws.cell(row=i, column=9, value=p.start_date)
        ws.cell(row=i, column=10, value=p.end_date)
        ws.cell(row=i, column=11, value=p.budget)
        if p.end_date:
            days = (p.end_date - today).days
            color = "FFCCCC" if days < 0 else "FFEB9C" if days <= 7 else "D4EDDA"
            rf = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for col in range(1, 12):
                ws.cell(row=i, column=col).fill = rf

    for col, w in enumerate([4, 8, 40, 15, 18, 18, 14, 20, 14, 14, 15], 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=lenta_projects.xlsx"},
    )
