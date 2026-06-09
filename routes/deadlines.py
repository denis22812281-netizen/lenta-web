import io
from datetime import date

from fastapi import APIRouter, Request, Depends
from fastapi.responses import HTMLResponse, StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

import models
from database import get_db
from deps import templates, require_login

router = APIRouter()


@router.get("/deadlines", response_class=HTMLResponse)
async def deadlines(request: Request, db: Session = Depends(get_db),
                    manager_id: str = None, view: str = "all",
                    user: dict = Depends(require_login)):
    today = date.today()
    pq = db.query(models.Project).filter(
        models.Project.status != "Завершён",
        models.Project.end_date != None,
        models.Project.project_type == "Констракшн",
        (models.Project.opening_date == None) | (models.Project.opening_date > today)
    )
    if manager_id and str(manager_id).isdigit():
        pq = pq.filter(models.Project.manager_id == int(manager_id))
    projects = pq.order_by(models.Project.end_date).all()

    tq = db.query(models.Task).filter(
        models.Task.status != "Завершена", models.Task.deadline != None)
    if manager_id and str(manager_id).isdigit():
        tq = tq.filter(models.Task.assignee_id == int(manager_id))
    tasks = tq.order_by(models.Task.deadline).all()

    managers = db.query(models.Manager).filter(models.Manager.is_leader == False).all()
    return templates.TemplateResponse("deadlines.html", {
        "request": request, "user": user,
        "projects": projects, "tasks": tasks,
        "managers": managers, "today": today,
        "filter_manager_id": manager_id, "view": view,
    })


@router.get("/deadlines/export")
async def export_deadlines_excel(request: Request, db: Session = Depends(get_db),
                                  user: dict = Depends(require_login)):
    today = date.today()
    projects = db.query(models.Project).filter(
        models.Project.status != "Завершён",
        models.Project.end_date != None,
        models.Project.project_type == "Констракшн",
        (models.Project.opening_date == None) | (models.Project.opening_date > today),
    ).order_by(models.Project.end_date).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Дедлайны"

    hfill = PatternFill(start_color="0D2010", end_color="0D2010", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center")

    headers = ["ТК №", "Название", "Город", "Менеджер", "Статус", "Плановое открытие", "Осталось дней"]
    ws.row_dimensions[1].height = 18
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont; c.alignment = center

    crit  = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    warn  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    ok    = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    over  = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

    for i, p in enumerate(projects, 2):
        days = (p.end_date - today).days if p.end_date else None
        row = [
            p.tk_number, p.name, p.city or p.address,
            p.manager.name if p.manager else "",
            p.status,
            p.end_date.strftime("%d.%m.%Y") if p.end_date else "",
            days,
        ]
        for col, val in enumerate(row, 1):
            ws.cell(row=i, column=col, value=val)
        if days is not None:
            fill = over if days < 0 else crit if days <= 3 else warn if days <= 14 else ok
            for col in range(1, 8):
                ws.cell(row=i, column=col).fill = fill

    for col, w in enumerate([10, 35, 18, 20, 14, 18, 14], 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"deadlines_{today.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )
