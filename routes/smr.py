"""График СМР — создание, просмотр, управление задачами, email-подтверждения."""
import io, os, secrets
from datetime import date, timedelta, datetime
from datetime import timedelta as td

from fastapi import APIRouter, Request, Form, Depends, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

import models
from database import get_db
from deps import templates, get_current_user
from services.smr_template import get_template
from services.email_service import send_smr_confirmation, send_smr_task_done, send_smr_progress_report

router = APIRouter()

APP_URL = os.getenv("APP_URL", "https://lenta-web-production.up.railway.app").rstrip("/")


# ── Список всех графиков ─────────────────────────────────────────────────────

@router.get("/smr", response_class=HTMLResponse)
async def smr_list(request: Request, db: Session = Depends(get_db),
                   search: str = "", manager_id: str = "", proj_type: str = ""):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)

    q = db.query(models.Project).filter(
        models.Project.project_type.in_(["Констракшн", "Реконструкция"])
    )
    if proj_type in ("Констракшн", "Реконструкция"):
        q = q.filter(models.Project.project_type == proj_type)
    if search:
        q = q.filter(models.Project.tk_number.contains(search))
    if manager_id and manager_id.isdigit():
        q = q.filter(models.Project.manager_id == int(manager_id))
    constr_projects = q.order_by(
        models.Project.project_type, models.Project.end_date.nullslast()
    ).all()

    schedules = {s.project_id: s for s in db.query(models.SmrSchedule).all()}
    managers  = db.query(models.Manager).order_by(models.Manager.name).all()

    projects = []
    for proj in constr_projects:
        sch = schedules.get(proj.id)
        done  = sum(1 for t in sch.tasks if t.status == "Выполнено") if sch else 0
        total = len(sch.tasks) if sch else 0
        projects.append({"proj": proj, "schedule": sch, "done": done, "total": total})

    return templates.TemplateResponse("smr_list.html", {
        "request": request, "user": user,
        "projects": projects, "today": date.today(),
        "managers": managers,
        "search": search, "filter_manager_id": manager_id,
    })


# ── Экспорт в Excel ──────────────────────────────────────────────────────────

@router.get("/smr/export")
async def smr_export(request: Request, db: Session = Depends(get_db),
                     search: str = "", manager_id: str = ""):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)

    q = db.query(models.Project).filter(
        models.Project.project_type.in_(["Констракшн", "Реконструкция"])
    )
    if search:
        q = q.filter(models.Project.tk_number.contains(search))
    if manager_id and manager_id.isdigit():
        q = q.filter(models.Project.manager_id == int(manager_id))
    projects = q.order_by(models.Project.project_type, models.Project.end_date.nullslast()).all()
    schedules = {s.project_id: s for s in db.query(models.SmrSchedule).all()}

    wb = Workbook()

    # ── Сводный лист ──
    ws_sum = wb.active
    ws_sum.title = "Сводка"

    # Стили — светлый фон для читаемости в Excel
    hfill     = PatternFill(start_color="1A5C22", end_color="1A5C22", fill_type="solid")
    hfont     = Font(color="FFFFFF", bold=True, size=11)
    msfill    = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    msfont    = Font(color="7B4400", bold=True, size=10)
    done_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    work_fill = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
    over_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    norm_font = Font(color="212529", size=10)
    thin      = Side(style="thin", color="BBBBBB")
    brd       = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap      = Alignment(vertical="center", wrap_text=True)

    headers = ["ТК №", "Город", "Менеджер", "Всего этапов", "Выполнено",
               "В работе", "Просрочено", "% выполнения",
               "ВПК 1", "ВПК 2", "Открытие"]
    col_w   = [10, 18, 22, 14, 12, 12, 12, 14, 14, 14, 14]

    for ci, (h, w) in enumerate(zip(headers, col_w), 1):
        c = ws_sum.cell(1, ci, h)
        c.fill, c.font, c.alignment, c.border = hfill, hfont, center, brd
        ws_sum.column_dimensions[get_column_letter(ci)].width = w
    ws_sum.row_dimensions[1].height = 26
    ws_sum.freeze_panes = "A2"

    for ri, proj in enumerate(projects, 2):
        sch   = schedules.get(proj.id)
        total = len(sch.tasks) if sch else 0
        done  = sum(1 for t in sch.tasks if t.status == "Выполнено") if sch else 0
        work  = sum(1 for t in sch.tasks if t.status == "В работе")  if sch else 0
        over  = sum(1 for t in sch.tasks if t.status == "Просрочено") if sch else 0
        pct   = f"{int(done/total*100)}%" if total else "—"

        # Даты ключевых вех
        vpk1 = vpk2 = opening = "—"
        if sch:
            for t in sch.tasks:
                if t.is_milestone:
                    if "ВПК 1" in t.name and t.end_plan:
                        vpk1 = t.end_plan.strftime("%d.%m.%Y")
                    elif "ВПК 2" in t.name and t.end_plan:
                        vpk2 = t.end_plan.strftime("%d.%m.%Y")
                    elif "Открытие" in t.name and t.end_plan:
                        opening = t.end_plan.strftime("%d.%m.%Y")

        row_fill = done_fill if (total and done == total) else PatternFill(fill_type=None)

        vals = [proj.tk_number, proj.city or "—",
                proj.manager.name if proj.manager else "—",
                total, done, work, over, pct, vpk1, vpk2, opening]
        for ci, v in enumerate(vals, 1):
            c = ws_sum.cell(ri, ci, v)
            c.fill, c.font, c.border = row_fill, norm_font, brd
            c.alignment = center if ci > 3 else wrap

    # ── Листы по объектам ──
    used_titles = set()
    for proj in projects:
        sch = schedules.get(proj.id)
        if not sch:
            continue
        # К = Констракшн, Р = Реконструкция — уникальный префикс
        prefix = "K" if proj.project_type == "Констракшн" else "R"
        base   = f"{prefix} TK{proj.tk_number}"
        title  = base[:31]
        # Защита от дублей
        if title in used_titles:
            title = f"{base[:28]}_{len(used_titles)}"[:31]
        used_titles.add(title)
        ws = wb.create_sheet(title=title)

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 48
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 18

        # Заголовок листа
        ws.merge_cells("A1:E1")
        header_text = f"ТК {proj.tk_number}  {proj.city or ''}  {proj.manager.name if proj.manager else ''}"
        c = ws.cell(1, 1, header_text)
        c.fill = hfill
        c.font = Font(color="FFFFFF", bold=True, size=12)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

        # Заголовки колонок
        for ci, h in enumerate(["#", "Этап работ", "Начало (план)", "Окончание (план)", "Статус"], 1):
            c = ws.cell(2, ci, h)
            c.fill, c.font, c.alignment, c.border = hfill, hfont, center, brd
        ws.row_dimensions[2].height = 22
        ws.freeze_panes = "A3"

        for ri, task in enumerate(sch.tasks, 3):
            is_ms    = task.is_milestone
            rfill    = msfill if is_ms else PatternFill(fill_type=None)
            rfont    = msfont if is_ms else norm_font
            num_font = Font(color="888888", size=9) if not is_ms else msfont

            n = ws.cell(ri, 1, ri - 2)
            n.fill, n.font, n.alignment, n.border = rfill, num_font, center, brd

            c = ws.cell(ri, 2, ("◆ " if is_ms else "  ") + task.name)
            c.fill, c.font, c.border, c.alignment = rfill, rfont, brd, wrap

            for ci, val in enumerate([
                task.start_plan.strftime("%d.%m.%Y") if task.start_plan else "—",
                task.end_plan.strftime("%d.%m.%Y")   if task.end_plan   else "—",
                task.status
            ], 3):
                c = ws.cell(ri, ci, val)
                c.fill, c.font, c.border, c.alignment = rfill, rfont, brd, center

            # Статус — цвет ячейки
            status_fills = {
                "Выполнено":   done_fill,
                "В работе":    work_fill,
                "Просрочено":  over_fill,
            }
            if task.status in status_fills:
                ws.cell(ri, 5).fill = status_fills[task.status]
                ws.cell(ri, 5).font = norm_font

            ws.row_dimensions[ri].height = 16

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    # Имя файла только ASCII — кириллица в Content-Disposition ненадёжна
    fname = f"SMR_schedules_{date.today().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


# ── Создать график по шаблону ─────────────────────────────────────────────────

@router.post("/smr/create/{project_id}")
async def smr_create(project_id: int, request: Request,
                     start_date: str = Form(...),
                     db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)

    proj = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not proj:
        raise HTTPException(status_code=404)

    # Если уже есть — удалить старый
    existing = db.query(models.SmrSchedule).filter(
        models.SmrSchedule.project_id == project_id).first()
    if existing:
        db.delete(existing)
        db.flush()

    try:
        base = date.fromisoformat(start_date)
    except ValueError:
        return RedirectResponse(f"/projects/{project_id}", status_code=303)

    schedule = models.SmrSchedule(project_id=project_id)
    db.add(schedule)
    db.flush()

    template = get_template(proj.project_type)
    for i, (name, s_day, e_day, is_ms) in enumerate(template):
        db.add(models.SmrTask(
            schedule_id=schedule.id,
            name=name,
            order=i,
            start_plan=base + timedelta(days=s_day),
            end_plan=base + timedelta(days=e_day),
            is_milestone=is_ms,
            status="Запланировано",
        ))
    db.commit()
    return RedirectResponse(f"/smr/{project_id}", status_code=303)


# ── Просмотр графика ─────────────────────────────────────────────────────────

@router.get("/smr/{project_id}", response_class=HTMLResponse)
async def smr_view(project_id: int, request: Request,
                   db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)

    proj = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not proj:
        raise HTTPException(status_code=404)

    schedule = db.query(models.SmrSchedule).filter(
        models.SmrSchedule.project_id == project_id).first()

    today = date.today()
    gantt_start = gantt_end = None
    if schedule and schedule.tasks:
        dates = [t.start_plan for t in schedule.tasks if t.start_plan] + \
                [t.end_plan   for t in schedule.tasks if t.end_plan]
        if dates:
            gantt_start = min(dates)
            gantt_end   = max(dates)

    return templates.TemplateResponse("smr_schedule.html", {
        "request": request, "user": user,
        "proj": proj, "schedule": schedule,
        "today": today,
        "gantt_start": gantt_start,
        "gantt_end":   gantt_end,
        "timedelta": timedelta,
    })


# ── Обновить статус задачи (AJAX) ────────────────────────────────────────────

@router.post("/api/smr/task/{task_id}/status")
async def smr_task_status(task_id: int, request: Request,
                          db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"error": "Не авторизован"}, status_code=401)
    data = await request.json()
    task = db.query(models.SmrTask).filter(models.SmrTask.id == task_id).first()
    if not task:
        return JSONResponse({"error": "Не найдено"}, status_code=404)

    old_status = task.status
    new_status  = data.get("status", task.status)
    task.status = new_status
    db.commit()

    # При выполнении — отправляем отчёт получателям задачи
    if new_status == "Выполнено" and old_status != "Выполнено":
        proj = task.schedule.project
        plan_date = task.end_plan.strftime("%d.%m.%Y") if task.end_plan else "—"
        completed_by = user.get("display_name", "")
        for email in [task.notify_email1, task.notify_email2]:
            if email:
                send_smr_task_done(
                    to_email=email,
                    task_name=task.name,
                    project_name=proj.name,
                    tk_number=proj.tk_number,
                    plan_date=plan_date,
                    completed_by=completed_by,
                )

    return {"ok": True, "status": task.status}


# ── Обновить email ответственных (AJAX) ──────────────────────────────────────

@router.post("/api/smr/task/{task_id}/emails")
async def smr_task_emails(task_id: int, request: Request,
                          db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"error": "Не авторизован"}, status_code=401)
    data = await request.json()
    task = db.query(models.SmrTask).filter(models.SmrTask.id == task_id).first()
    if not task:
        return JSONResponse({"error": "Не найдено"}, status_code=404)
    task.notify_email1 = data.get("email1", "").strip().lower()
    task.notify_email2 = data.get("email2", "").strip().lower()
    db.commit()
    return {"ok": True}


# ── Отправить запрос на подтверждение ────────────────────────────────────────

@router.post("/api/smr/task/{task_id}/send-confirm")
async def smr_send_confirm(task_id: int, request: Request,
                           db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return JSONResponse({"error": "Не авторизован"}, status_code=401)

    task = db.query(models.SmrTask).filter(models.SmrTask.id == task_id).first()
    if not task:
        return JSONResponse({"error": "Не найдено"}, status_code=404)

    proj = task.schedule.project
    sent = []

    for email in [task.notify_email1, task.notify_email2]:
        if not email:
            continue
        token = secrets.token_hex(32)
        db.add(models.SmrConfirmation(task_id=task.id, token=token, email=email))
        confirm_url = f"{APP_URL}/smr/confirm/{token}"
        reject_url  = f"{APP_URL}/smr/confirm/{token}?action=reject"
        try:
            send_smr_confirmation(
                to_email=email,
                task_name=task.name,
                project_name=proj.name,
                tk_number=proj.tk_number,
                plan_date=task.end_plan.strftime("%d.%m.%Y") if task.end_plan else "—",
                confirm_url=confirm_url,
                reject_url=reject_url,
            )
            sent.append(email)
        except Exception as e:
            pass

    db.commit()
    return {"ok": True, "sent": sent}


# ── Публичная страница подтверждения (без авторизации) ───────────────────────

@router.get("/smr/confirm/{token}", response_class=HTMLResponse)
async def smr_confirm_page(token: str, request: Request,
                           action: str = "confirm",
                           db: Session = Depends(get_db)):
    conf = db.query(models.SmrConfirmation).filter(
        models.SmrConfirmation.token == token).first()

    if not conf:
        return templates.TemplateResponse("smr_confirm.html", {
            "request": request, "error": "Ссылка недействительна или устарела."
        })

    already_done = bool(conf.action)
    task = conf.task
    proj = task.schedule.project if task and task.schedule else None

    # Если уже ответил — показываем результат
    if already_done:
        return templates.TemplateResponse("smr_confirm.html", {
            "request": request, "conf": conf,
            "task": task, "proj": proj, "already_done": True,
        })

    # Нажал "Выполнено" — сразу фиксируем
    if action != "reject":
        conf.action       = "confirmed"
        conf.responded_at = datetime.utcnow()
        if task:
            task.status = "Выполнено"
        db.commit()
        return templates.TemplateResponse("smr_confirm.html", {
            "request": request, "conf": conf,
            "task": task, "proj": proj, "already_done": False,
        })

    # Нажал "Не выполнено" — показываем форму комментария
    return templates.TemplateResponse("smr_confirm.html", {
        "request": request, "conf": conf,
        "task": task, "proj": proj,
        "already_done": False,
        "show_comment_form": True,   # ← ключевой флаг
        "token": token,
    })


@router.post("/smr/confirm/{token}", response_class=HTMLResponse)
async def smr_confirm_submit(token: str, request: Request,
                              db: Session = Depends(get_db),
                              comment: str = Form("")):
    """POST — сохраняет комментарий при отклонении."""
    conf = db.query(models.SmrConfirmation).filter(
        models.SmrConfirmation.token == token).first()

    if not conf:
        return templates.TemplateResponse("smr_confirm.html", {
            "request": request, "error": "Ссылка недействительна."
        })

    task = conf.task
    proj = task.schedule.project if task and task.schedule else None

    if not conf.action:
        conf.action       = "rejected"
        conf.responded_at = datetime.utcnow()
        if task:
            task.status         = "Просрочено"
            task.reject_comment = comment.strip()
        db.commit()

    return templates.TemplateResponse("smr_confirm.html", {
        "request": request, "conf": conf,
        "task": task, "proj": proj,
        "already_done": False,
        "comment_saved": True,
    })


# ── Отчёт о ходе выполнения графика ─────────────────────────────────────────

@router.post("/api/smr/{project_id}/send-report")
async def smr_send_report(project_id: int, request: Request,
                          db: Session = Depends(get_db)):
    """Отправляет отчёт по графику СМР на указанный email."""
    user = get_current_user(request)
    if not user:
        return JSONResponse({"error": "Не авторизован"}, status_code=401)

    data   = await request.json()
    to_email = data.get("email", "").strip().lower()
    if not to_email:
        return JSONResponse({"error": "Email не указан"}, status_code=400)

    proj = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not proj:
        return JSONResponse({"error": "Проект не найден"}, status_code=404)

    sch = db.query(models.SmrSchedule).filter(
        models.SmrSchedule.project_id == project_id).first()
    if not sch:
        return JSONResponse({"error": "График не создан"}, status_code=404)

    tasks = [
        {
            "name":        t.name,
            "status":      t.status,
            "end_plan":    t.end_plan.strftime("%d.%m.%Y") if t.end_plan else "",
            "is_milestone": t.is_milestone,
        }
        for t in sch.tasks
    ]

    ok = send_smr_progress_report(
        to_email=to_email,
        tk_number=proj.tk_number,
        project_name=proj.name,
        city=proj.city or "",
        manager_name=proj.manager.name if proj.manager else "—",
        report_date=date.today().strftime("%d.%m.%Y"),
        tasks=tasks,
        sent_by=user.get("display_name", ""),
    )
    return {"ok": ok, "sent_to": to_email}


# ── База контактов ───────────────────────────────────────────────────────────

@router.get("/smr/contacts", response_class=HTMLResponse)
async def smr_contacts(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    contacts = db.query(models.SmrContact).order_by(models.SmrContact.name).all()
    return templates.TemplateResponse("smr_contacts.html", {
        "request": request, "user": user, "contacts": contacts,
    })


@router.post("/smr/contacts/add")
async def smr_contact_add(request: Request, db: Session = Depends(get_db),
                           name: str = Form(...), email: str = Form(...),
                           position: str = Form("")):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    db.add(models.SmrContact(
        name=name.strip(),
        email=email.strip().lower(),
        position=position.strip(),
    ))
    db.commit()
    return RedirectResponse("/smr/contacts", status_code=303)


@router.post("/smr/contacts/{contact_id}/delete")
async def smr_contact_delete(contact_id: int, request: Request,
                              db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    c = db.query(models.SmrContact).filter(models.SmrContact.id == contact_id).first()
    if c:
        db.delete(c)
        db.commit()
    return RedirectResponse("/smr/contacts", status_code=303)


@router.get("/api/smr/contacts")
async def smr_contacts_search(request: Request, db: Session = Depends(get_db),
                               q: str = ""):
    """Поиск контактов по имени — для автодополнения."""
    user = get_current_user(request)
    if not user:
        return JSONResponse({"contacts": []})
    query = db.query(models.SmrContact)
    if q and len(q) >= 1:
        query = query.filter(models.SmrContact.name.ilike(f"%{q}%"))
    contacts = query.order_by(models.SmrContact.name).limit(10).all()
    return JSONResponse({"contacts": [
        {"id": c.id, "name": c.name, "email": c.email, "position": c.position}
        for c in contacts
    ]})


# ── Удалить график ────────────────────────────────────────────────────────────

@router.post("/smr/delete/{project_id}")
async def smr_delete(project_id: int, request: Request,
                     db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    sch = db.query(models.SmrSchedule).filter(
        models.SmrSchedule.project_id == project_id).first()
    if sch:
        db.delete(sch)
        db.commit()
    return RedirectResponse(f"/projects/{project_id}", status_code=303)
