import io
from datetime import datetime
from pathlib import Path

import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session

import models
from database import get_db
from deps import get_current_user, require_admin, require_login, templates
from utils.excel import match_manager
from utils.files import read_limited

router = APIRouter()


@router.get("/kso", response_class=HTMLResponse)
async def kso_view(request: Request, db: Session = Depends(get_db),
                   user: dict = Depends(require_login)):
    manager_id = request.query_params.get("manager_id")
    search = request.query_params.get("search")
    tab = request.query_params.get("tab", "objects")
    q = db.query(models.KsoObject)
    if manager_id and str(manager_id).isdigit():
        q = q.filter(models.KsoObject.manager_id == int(manager_id))
    if search:
        q = q.filter(models.KsoObject.tk_number.contains(search))
    objects = q.order_by(models.KsoObject.manager_id, models.KsoObject.tk_number).all()
    schedules = db.query(models.KsoSchedule).order_by(
        models.KsoSchedule.uploaded_at.desc()).all()
    managers = db.query(models.Manager).all()
    done_count = sum(1 for o in objects if o.done)
    return templates.TemplateResponse("kso.html", {
        "request": request, "user": user,
        "objects": objects, "schedules": schedules,
        "managers": managers, "tab": tab,
        "filter_manager_id": manager_id, "search": search or "",
        "done_count": done_count, "total": len(objects),
        "msg": request.query_params.get("msg"),
        "error": request.query_params.get("error"),
    })


@router.post("/kso/import")
async def kso_import(request: Request, db: Session = Depends(get_db),
                     file: UploadFile = File(...),
                     user: dict = Depends(require_login)):
    try:
        content = await read_limited(file, 10 * 1024 * 1024)
    except ValueError as e:
        return RedirectResponse(f"/kso?error={str(e)[:80]}&tab=objects", status_code=303)
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.worksheets[0]
        managers = db.query(models.Manager).all()
        created = 0
        col_tk = col_addr = col_mgr = None
        header_row = 1
        for r in range(1, 6):
            for c in range(1, 20):
                v = str(ws.cell(r, c).value or '').strip().lower()
                if 'тк' in v or 'номер' in v:
                    col_tk = c; header_row = r
                elif 'адрес' in v and not col_addr:
                    col_addr = c
                elif 'менеджер' in v and not col_mgr:
                    col_mgr = c
        col_tk = col_tk or 1; col_addr = col_addr or 2; col_mgr = col_mgr or 3
        for row_idx in range(header_row + 1, ws.max_row + 1):
            tk = str(ws.cell(row_idx, col_tk).value or '').strip()
            if not tk or tk in ('None', '—'):
                continue
            addr    = str(ws.cell(row_idx, col_addr).value or '').strip()
            mgr_val = str(ws.cell(row_idx, col_mgr).value or '').strip()
            mgr_id  = match_manager(mgr_val, managers)
            if not db.query(models.KsoObject).filter(
                    models.KsoObject.tk_number == tk).first():
                db.add(models.KsoObject(tk_number=tk, address=addr, manager_id=mgr_id))
                created += 1
        db.commit()
        return RedirectResponse(f"/kso?msg=Загружено: {created} объектов&tab=objects",
                                status_code=303)
    except Exception as e:
        return RedirectResponse(f"/kso?error={str(e)[:100]}&tab=objects", status_code=303)


@router.post("/kso/objects/{obj_id}/toggle")
async def kso_toggle(obj_id: int, request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return {"error": "Не авторизован"}
    obj = db.query(models.KsoObject).filter(models.KsoObject.id == obj_id).first()
    if not obj:
        raise HTTPException(status_code=404)
    obj.done = not obj.done
    db.commit()
    return {"done": obj.done}


@router.post("/api/kso/{obj_id}/comment")
async def kso_comment(obj_id: int, request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return {"error": "Не авторизован"}
    data = await request.json()
    obj = db.query(models.KsoObject).filter(models.KsoObject.id == obj_id).first()
    if not obj:
        raise HTTPException(status_code=404)
    obj.comment = data.get("comment", "")
    db.commit()
    return {"ok": True}


@router.post("/kso/objects/{obj_id}/delete")
async def kso_delete_object(obj_id: int, request: Request, db: Session = Depends(get_db),
                             user: dict = Depends(require_admin)):
    obj = db.query(models.KsoObject).filter(models.KsoObject.id == obj_id).first()
    if obj:
        db.delete(obj)
        db.commit()
    return RedirectResponse("/kso?tab=objects", status_code=303)


@router.post("/kso/schedules/upload")
async def kso_schedule_upload(request: Request, db: Session = Depends(get_db),
                               file: UploadFile = File(...), description: str = Form(""),
                               user: dict = Depends(require_login)):
    _ALLOWED_KSO = {".pdf", ".xlsx", ".xls", ".doc", ".docx", ".jpg", ".jpeg", ".png"}
    ext = Path(file.filename or "").suffix.lower()
    if ext not in _ALLOWED_KSO:
        return RedirectResponse("/kso?tab=schedules&msg=Недопустимый тип файла", status_code=303)
    try:
        content = await read_limited(file, 50 * 1024 * 1024)
    except ValueError:
        return RedirectResponse("/kso?tab=schedules&msg=Файл слишком большой (макс 50MB)", status_code=303)
    save_dir = Path("static/uploads/kso")
    save_dir.mkdir(parents=True, exist_ok=True)
    safe_name = f"{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
    safe_name = "".join(c if c.isalnum() or c in "._-" else "_" for c in safe_name)
    (save_dir / safe_name).write_bytes(content)
    db.add(models.KsoSchedule(
        original_name=file.filename, filename=safe_name,
        description=description, uploaded_by=user.get("display_name", ""),
    ))
    db.commit()
    return RedirectResponse("/kso?tab=schedules&msg=Файл загружен", status_code=303)


_KSO_UPLOAD_DIR = Path("static/uploads/kso")


def _safe_kso_path(filename: str) -> Path:
    upload_dir = _KSO_UPLOAD_DIR.resolve()
    path = (upload_dir / filename).resolve()
    if not path.is_relative_to(upload_dir):
        raise HTTPException(status_code=400, detail="Недопустимый путь файла")
    return path


@router.get("/kso/schedules/{sch_id}/download")
async def kso_schedule_download(sch_id: int, request: Request, db: Session = Depends(get_db),
                                 user: dict = Depends(require_login)):
    sch = db.query(models.KsoSchedule).filter(models.KsoSchedule.id == sch_id).first()
    if not sch:
        raise HTTPException(status_code=404)
    path = _safe_kso_path(sch.filename)
    if not path.exists():
        raise HTTPException(status_code=404, detail="Файл не найден")
    return StreamingResponse(open(path, "rb"),
        headers={"Content-Disposition": f'attachment; filename="{sch.original_name}"'})


@router.post("/kso/schedules/{sch_id}/delete")
async def kso_schedule_delete(sch_id: int, request: Request, db: Session = Depends(get_db),
                               user: dict = Depends(require_login)):
    sch = db.query(models.KsoSchedule).filter(models.KsoSchedule.id == sch_id).first()
    if sch:
        try:
            path = _safe_kso_path(sch.filename)
            if path.exists():
                path.unlink()
        except HTTPException:
            pass
        db.delete(sch)
        db.commit()
    return RedirectResponse("/kso?tab=schedules", status_code=303)
