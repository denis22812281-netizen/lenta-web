import io
from datetime import date, datetime

import openpyxl
from fastapi import APIRouter, Depends, File, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

import models
from database import get_db
from deps import get_current_user, require_login, templates
from utils.files import read_limited

router = APIRouter()


@router.get("/stats", response_class=HTMLResponse)
async def stats_page(request: Request, db: Session = Depends(get_db),
                     user: dict = Depends(require_login)):
    today    = date.today()
    managers = db.query(models.Manager).all()
    projects = db.query(models.Project).filter(
        models.Project.project_type == "Констракшн").all()

    def opening_color(p):
        if not p.opening_date or p.opening_date > today:
            return "active"
        if not p.end_date:
            return "on-time"
        if p.opening_date < p.end_date:
            return "early"
        if p.opening_date == p.end_date:
            return "on-time"
        return "late"

    manager_stats = []
    for m in managers:
        m_proj = [p for p in projects if p.manager_id == m.id]
        if not m_proj:
            continue
        manager_stats.append({
            "name": m.name,
            "early":   sum(1 for p in m_proj if opening_color(p) == "early"),
            "on_time": sum(1 for p in m_proj if opening_color(p) == "on-time"),
            "late":    sum(1 for p in m_proj if opening_color(p) == "late"),
            "active":  sum(1 for p in m_proj if not p.opening_date or p.opening_date > today),
            "total":   len(m_proj),
        })

    projects_with_stats = []
    for p in sorted(projects, key=lambda x: (x.manager_id or 0, x.end_date or date.max)):
        color = opening_color(p)
        delta = (p.opening_date - p.end_date).days if (p.opening_date and p.end_date) else None
        days_left = (p.end_date - today).days if (p.end_date and not p.opening_date) else None
        projects_with_stats.append({
            "id": p.id, "tk_number": p.tk_number,
            "address": p.address or p.city or "—",
            "format_type": p.format_type or "",
            "manager_name": p.manager.name if p.manager else "—",
            "end_date": p.end_date, "opening_date": p.opening_date,
            "color": color, "delta_days": delta, "days_left": days_left,
            "delay_reason": p.delay_reason or "",
        })

    return templates.TemplateResponse("stats.html", {
        "request": request, "user": user,
        "manager_stats": manager_stats,
        "projects_with_stats": projects_with_stats,
        "total_early":   sum(s["early"]   for s in manager_stats),
        "total_on_time": sum(s["on_time"] for s in manager_stats),
        "total_late":    sum(s["late"]    for s in manager_stats),
        "total_active":  sum(s["active"]  for s in manager_stats),
        "today": today,
    })


@router.post("/api/projects/{project_id}/delay-reason")
async def save_delay_reason(project_id: int, request: Request, db: Session = Depends(get_db)):
    from fastapi import HTTPException
    user = get_current_user(request)
    if not user:
        return {"error": "Не авторизован"}
    if not user.get("is_admin"):
        return {"error": "Нет доступа"}
    data = await request.json()
    p = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not p:
        raise HTTPException(status_code=404)
    p.delay_reason = data.get("reason", "")
    p.updated_at = datetime.utcnow()
    db.commit()
    return {"ok": True}


@router.post("/stats/upload")
async def stats_upload(request: Request, db: Session = Depends(get_db),
                       user: dict = Depends(require_login),
                       file: UploadFile = File(...)):
    try:
        content = await read_limited(file, 10 * 1024 * 1024)
    except ValueError:
        return RedirectResponse("/stats?error=Файл слишком большой (макс 10 МБ)", status_code=303)
    try:
        from utils.excel import safe_date
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        updated = 0
        today = date.today()
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                tk_num = open_date = None
                for val in row:
                    if isinstance(val, (int, float)) and 100 < val < 9999:
                        tk_num = str(int(val))
                    d = safe_date(val)
                    if d:
                        open_date = d
                if tk_num and open_date:
                    p = db.query(models.Project).filter(
                        models.Project.tk_number == tk_num).first()
                    if p:
                        p.opening_date = open_date
                        if open_date <= today:
                            p.status = "Завершён"
                        updated += 1
        db.commit()
    except Exception as e:
        return RedirectResponse(f"/stats?error={str(e)[:80]}", status_code=303)
    return RedirectResponse(f"/stats?updated={updated}", status_code=303)


@router.get("/stats/export")
async def stats_export(request: Request, db: Session = Depends(get_db),
                       user: dict = Depends(require_login),
                       date_from: str = "", date_to: str = ""):
    today_date = date.today()
    q = db.query(models.Project).filter(
        models.Project.project_type == "Констракшн",
        models.Project.opening_date != None,
        models.Project.opening_date <= today_date,
    )
    if date_from:
        try:
            q = q.filter(models.Project.opening_date >= datetime.strptime(date_from, "%Y-%m-%d").date())
        except ValueError:
            pass
    if date_to:
        try:
            q = q.filter(models.Project.opening_date <= datetime.strptime(date_to, "%Y-%m-%d").date())
        except ValueError:
            pass
    projects = q.order_by(models.Project.manager_id, models.Project.opening_date).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Статистика открытий"
    hfill     = PatternFill(start_color="1A5C22", end_color="1A5C22", fill_type="solid")
    hfont     = Font(color="FFFFFF", bold=True, size=11)
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fill_early   = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    fill_on_time = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_late    = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    headers    = ["№", "ТК №", "Адрес", "Формат", "Менеджер",
                  "План открытия", "Факт открытия", "Отклонение (дн)", "Результат", "Комментарий"]
    col_widths = [5, 12, 40, 10, 20, 16, 16, 16, 18, 35]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont; c.alignment = center
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[1].height = 30

    for i, p in enumerate(projects, 1):
        if p.end_date:
            if p.opening_date < p.end_date:
                result, rf, delta = "Раньше срока", fill_early, (p.opening_date - p.end_date).days
            elif p.opening_date == p.end_date:
                result, rf, delta = "Вовремя", fill_on_time, 0
            else:
                result, rf, delta = "С опозданием", fill_late, (p.opening_date - p.end_date).days
        else:
            result, rf, delta = "Вовремя", fill_on_time, None

        row_data = [
            i, p.tk_number or "", p.address or p.city or "",
            p.format_type or "", p.manager.name if p.manager else "—",
            p.end_date.strftime("%d.%m.%Y") if p.end_date else "—",
            p.opening_date.strftime("%d.%m.%Y"),
            delta if delta is not None else "—",
            result, p.delay_reason or "",
        ]
        row_num = i + 1
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row_num, column=col, value=val)
            c.fill = rf
            if col in (6, 7):
                c.alignment = Alignment(horizontal="center")
            if col == 8 and isinstance(val, int):
                c.alignment = Alignment(horizontal="center")
                if val < 0:
                    c.font = Font(color="16A34A", bold=True)
                elif val > 0:
                    c.font = Font(color="DC2626", bold=True)

    total_row = len(projects) + 2
    early_c  = sum(1 for p in projects if p.end_date and p.opening_date < p.end_date)
    ontime_c = sum(1 for p in projects if not p.end_date or p.opening_date == p.end_date)
    late_c   = sum(1 for p in projects if p.end_date and p.opening_date > p.end_date)
    ws.cell(row=total_row, column=1, value="ИТОГО").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=len(projects)).font = Font(bold=True)
    ws.cell(row=total_row, column=9,
            value=f"Раньше: {early_c} | Вовремя: {ontime_c} | Опозд.: {late_c}"
            ).font = Font(bold=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    period = f"{date_from or 'начало'}_{date_to or today_date.strftime('%Y-%m-%d')}"
    return StreamingResponse(output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=construction_stats_{period}.xlsx"})
