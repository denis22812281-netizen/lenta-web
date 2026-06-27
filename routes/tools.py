"""Data → Excel conversion tool (admin only): photo, text, multi-photo, compare, merge."""
import asyncio
import base64
import io
import json
import logging
import os
import re
import time as _time
import urllib.parse
import uuid
from datetime import date as dt
from datetime import timedelta
from typing import Any, List

import httpx
from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import AI_CLAUDE_MODEL, AI_GEMINI_MODEL, AI_MAX_OUTPUT_TOKENS
from deps import require_admin, templates

logger = logging.getLogger(__name__)
router = APIRouter()

_GEMINI_KEY    = os.getenv("GEMINI_API_KEY", "")
_ANTHROPIC_KEY = os.getenv("ANTHROPIC_API_KEY", "")
_GROQ_KEY      = os.getenv("GROQ_API_KEY", "")

_GEMINI_ENDPOINT = (
    "https://generativelanguage.googleapis.com/v1beta/models/"
    f"{AI_GEMINI_MODEL}:generateContent"
)

_COL_MIN_WIDTH = 12
_COL_MAX_WIDTH = 40
_COL_PADDING   = 4

# ── In-memory templates (per user, persistent until server restart) ───────────
# {user_id: [{id, name, output_type, mode, created_at}]}
_TEMPLATES: dict[int, list[dict]] = {}
_TEMPLATES_MAX = 20

def _template_save(user_id: int, name: str, output_type: str, mode: str) -> str:
    uid = str(uuid.uuid4())[:8]
    items = _TEMPLATES.get(user_id, [])
    # prevent duplicate names
    items = [i for i in items if i["name"] != name]
    items.append({"id": uid, "name": name, "output_type": output_type,
                  "mode": mode, "created_at": _time.strftime("%d.%m.%Y %H:%M")})
    _TEMPLATES[user_id] = items[-_TEMPLATES_MAX:]
    return uid

def _template_delete(user_id: int, template_id: str) -> bool:
    items = _TEMPLATES.get(user_id, [])
    before = len(items)
    _TEMPLATES[user_id] = [i for i in items if i["id"] != template_id]
    return len(_TEMPLATES[user_id]) < before

# ── In-memory history (per user, max 10 items, TTL 24h) ──────────────────────
_HISTORY: dict[int, list[dict]] = {}  # user_id → [{id, ts, filename, type, blob, size}]
_HISTORY_MAX = 10
_HISTORY_TTL = 86400  # 24 hours

def _history_save(user_id: int, filename: str, out_type: str, blob: bytes) -> str:
    uid = str(uuid.uuid4())[:8]
    now = _time.time()
    items = [i for i in _HISTORY.get(user_id, []) if now - i["ts"] < _HISTORY_TTL]
    items.append({"id": uid, "ts": now, "filename": filename, "type": out_type,
                  "blob": blob, "size": len(blob)})
    _HISTORY[user_id] = items[-_HISTORY_MAX:]
    return uid

def _history_get(user_id: int, item_id: str) -> dict | None:
    now = _time.time()
    for item in _HISTORY.get(user_id, []):
        if item["id"] == item_id and now - item["ts"] < _HISTORY_TTL:
            return item
    return None

_TABLE_PROMPT = (
    "Ты эксперт по извлечению данных. "
    "Извлеки ВСЕ данные таблицы. "
    "1) title — название таблицы. "
    "2) headers — только названия столбцов. "
    "3) rows — все строки данных, НЕ включая заголовок. "
    "Верни ТОЛЬКО валидный JSON:\n"
    '{"title":"Название","headers":["№","Столбец1","Столбец2"],'
    '"rows":[["1","значение1","значение2"]]}'
)

_PROMPTS: dict[str, str] = {
    "table": "Ты эксперт по извлечению данных из изображений. Посмотри на изображение и извлеки ВСЕ данные таблицы. "
             "1) title — реальное название. 2) headers — названия столбцов. 3) rows — строки данных без заголовка. "
             "Верни ТОЛЬКО валидный JSON:\n"
             '{"title":"Название","headers":["№","Столбец1","Столбец2"],"rows":[["1","a","b"]]}',
    "chart": "Ты эксперт по извлечению данных из изображений. Посмотри и извлеки данные для графика. "
             "Временной ряд → line, иначе → bar. Верни ТОЛЬКО валидный JSON:\n"
             '{"title":"Название","chart_type":"bar","categories":["янв","фев"],"series":[{"name":"Серия","values":[10,20]}]}',
    "diagram": "Ты эксперт по извлечению данных из изображений. Извлеки данные для круговой диаграммы. "
               "Верни ТОЛЬКО валидный JSON:\n"
               '{"title":"Диаграмма","labels":["А","Б","В"],"values":[30,50,20]}',
    "gantt": (
        "Ты эксперт по извлечению данных из изображений. "
        "Посмотри на изображение с графиком работ (диаграммой Ганта) и извлеки все строки. "
        "Для каждой строки: id (номер), name (название задачи/события), object (объект/код, если есть), "
        "start (дата начала в формате дд.мм), end (дата окончания в формате дд.мм). "
        "Если задача является вехой — в имени есть •, или start==end. "
        "Определи год из заголовка. "
        "Верни ТОЛЬКО валидный JSON:\n"
        '{"title":"График работ СМ 3185 | июль–сентябрь 2026","year":2026,'
        '"tasks":[{"id":1,"name":"СМР помещения БЭК","object":"СМ 3185","start":"06.07","end":"27.07"},'
        '{"id":2,"name":"• Закрытие СМ 3187","object":"СМ 3185","start":"26.07","end":"26.07"}]}'
    ),
    "auto": (
        "Ты эксперт по извлечению данных из изображений. "
        "Посмотри на изображение и определи тип данных: "
        "gantt — если видишь график работ с датами и задачами, "
        "table — если обычная таблица, "
        "chart — если временной ряд или категориальный бар-чарт, "
        "diagram — если круговая диаграмма с долями/процентами. "
        "Верни JSON с полем detected_type и всеми данными этого типа:\n"
        '{"detected_type":"table","title":"...","headers":[...],"rows":[[...]]} ИЛИ\n'
        '{"detected_type":"chart","title":"...","chart_type":"bar","categories":[...],"series":[...]} ИЛИ\n'
        '{"detected_type":"diagram","title":"...","labels":[...],"values":[...]} ИЛИ\n'
        '{"detected_type":"gantt","title":"...","year":2026,"tasks":[...]}'
    ),
}

# Text-only prompts (same structure, different instruction prefix)
_TEXT_PROMPTS: dict[str, str] = {
    k: v.replace("Посмотри на изображение и", "Ты получил текстовые данные. Проанализируй следующий текст и")
       .replace("из изображений", "из текста")
       .replace("Посмотри на изображение с", "Проанализируй текстовые данные с")
    for k, v in _PROMPTS.items()
}

_NUM_HEADERS = {"№", "n", "#", "no", "номер", "num", "п/п", "п.п."}


# ── AI callers ────────────────────────────────────────────────────────────────

def _parse_ai_json(raw: str) -> dict[str, Any]:
    """Strip markdown fences and parse JSON; repairs truncated responses."""
    raw = re.sub(r"^```(?:json)?\s*", "", raw.strip())
    raw = re.sub(r"\s*```$", "", raw)
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        fixed = raw.rstrip().rstrip(",")
        fixed += "]" * max(0, fixed.count("[") - fixed.count("]"))
        fixed += "}" * max(0, fixed.count("{") - fixed.count("}"))
        return json.loads(fixed)


async def _call_gemini(image_bytes: bytes, mime: str, output_type: str) -> dict[str, Any]:
    b64     = base64.standard_b64encode(image_bytes).decode("ascii")
    payload = {
        "contents": [{"parts": [
            {"inline_data": {"mime_type": mime, "data": b64}},
            {"text": _PROMPTS[output_type]},
        ]}],
        "generationConfig": {
            "maxOutputTokens": AI_MAX_OUTPUT_TOKENS,
            "temperature":     0.0,
        },
    }
    async with httpx.AsyncClient(timeout=60.0) as client:
        resp = await client.post(
            _GEMINI_ENDPOINT,
            params={"key": _GEMINI_KEY},
            json=payload,
        )
    if resp.status_code != 200:
        raise RuntimeError(f"Gemini {resp.status_code}: {resp.text[:400]}")

    logger.info("Gemini response: %d tokens used", resp.json().get("usageMetadata", {}).get("totalTokenCount", 0))
    raw = resp.json()["candidates"][0]["content"]["parts"][0]["text"]
    return _parse_ai_json(raw)


async def _call_claude(image_bytes: bytes, mime: str, output_type: str) -> dict[str, Any]:
    import anthropic
    client = anthropic.AsyncAnthropic(api_key=_ANTHROPIC_KEY)
    b64    = base64.standard_b64encode(image_bytes).decode("ascii")
    resp   = await client.messages.create(
        model=AI_CLAUDE_MODEL,
        max_tokens=AI_MAX_OUTPUT_TOKENS,
        messages=[{
            "role": "user",
            "content": [
                {"type": "image", "source": {"type": "base64", "media_type": mime, "data": b64}},
                {"type": "text",  "text": _PROMPTS[output_type]},
            ],
        }],
    )
    return _parse_ai_json(resp.content[0].text)


async def _call_ai(image_bytes: bytes, mime: str, output_type: str) -> dict[str, Any]:
    """Tries Gemini first (free tier), falls back to Claude. Retries once on failure."""
    if not _GEMINI_KEY and not _ANTHROPIC_KEY:
        raise RuntimeError("Не задан ни GEMINI_API_KEY, ни ANTHROPIC_API_KEY")

    caller   = _call_gemini if _GEMINI_KEY else _call_claude
    last_err: Exception | None = None

    for attempt in range(2):
        try:
            return await caller(image_bytes, mime, output_type)
        except Exception as exc:
            last_err = exc
            if attempt == 0:
                logger.warning("AI attempt 1 failed (%s), retrying in 2s…", exc)
                await asyncio.sleep(2)

    assert last_err is not None
    raise last_err


async def _call_text_ai(prompt: str) -> dict[str, Any]:
    """Text-only AI call: Groq first (fast/free), then Claude, then Gemini."""
    # Groq (fastest for text)
    if _GROQ_KEY:
        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                resp = await client.post(
                    "https://api.groq.com/openai/v1/chat/completions",
                    headers={"Authorization": f"Bearer {_GROQ_KEY}"},
                    json={
                        "model": "llama-3.3-70b-versatile",
                        "messages": [{"role": "user", "content": prompt}],
                        "max_tokens": AI_MAX_OUTPUT_TOKENS,
                        "temperature": 0.0,
                    },
                )
            if resp.status_code == 200:
                return _parse_ai_json(resp.json()["choices"][0]["message"]["content"])
        except Exception as exc:
            logger.warning("Groq text failed: %s", exc)

    # Claude fallback
    if _ANTHROPIC_KEY:
        import anthropic
        client = anthropic.AsyncAnthropic(api_key=_ANTHROPIC_KEY)
        resp   = await client.messages.create(
            model=AI_CLAUDE_MODEL,
            max_tokens=AI_MAX_OUTPUT_TOKENS,
            messages=[{"role": "user", "content": prompt}],
        )
        return _parse_ai_json(resp.content[0].text)

    # Gemini text fallback
    if _GEMINI_KEY:
        payload = {
            "contents": [{"parts": [{"text": prompt}]}],
            "generationConfig": {"maxOutputTokens": AI_MAX_OUTPUT_TOKENS, "temperature": 0.0},
        }
        async with httpx.AsyncClient(timeout=30.0) as client:
            resp = await client.post(_GEMINI_ENDPOINT, params={"key": _GEMINI_KEY}, json=payload)
        return _parse_ai_json(resp.json()["candidates"][0]["content"]["parts"][0]["text"])

    raise RuntimeError("Нет доступных AI-ключей")


# ── Excel builders ────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_BORDER_SIDE = Side(style="thin", color="CCCCCC")
_CELL_BORDER = Border(
    left=_BORDER_SIDE, right=_BORDER_SIDE,
    top=_BORDER_SIDE,  bottom=_BORDER_SIDE,
)


def _auto_col_width(ws: Any, ncols: int, nrows: int) -> None:
    for ci in range(1, ncols + 1):
        width = _COL_MIN_WIDTH
        for ri in range(1, nrows + 3):
            val = ws.cell(ri, ci).value
            if val:
                width = max(width, min(len(str(val)) + _COL_PADDING, _COL_MAX_WIDTH))
        ws.column_dimensions[get_column_letter(ci)].width = width


def _build_table(data: dict[str, Any]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Данные"

    title   = data.get("title", "Таблица")
    headers = data.get("headers", [])
    rows    = data.get("rows", [])
    ncols   = max(len(headers), max((len(r) for r in rows), default=0))

    # Auto-fix row numbering when the first column is an index column
    if headers and str(headers[0]).strip().lower() in _NUM_HEADERS:
        for i, row in enumerate(rows, 1):
            if row:
                row[0] = str(i)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(ncols, 1))
    title_cell            = ws.cell(1, 1, title)
    title_cell.font       = Font(bold=True, size=14, color="1E3A5F")
    title_cell.alignment  = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for ci, h in enumerate(headers, 1):
        cell           = ws.cell(2, ci, str(h))
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _CELL_BORDER
    ws.row_dimensions[2].height = 22

    for ri, row in enumerate(rows, 3):
        for ci, val in enumerate(row, 1):
            cell           = ws.cell(ri, ci, val)
            cell.border    = _CELL_BORDER
            cell.alignment = Alignment(wrap_text=True)
            if ri % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F0F4F8")

    _auto_col_width(ws, ncols, len(rows))
    return wb


def _build_chart(data: dict[str, Any]) -> Workbook:
    wb      = Workbook()
    ws_data = wb.active
    ws_data.title = "Данные"

    title       = data.get("title", "График")
    chart_type  = data.get("chart_type", "bar").lower()
    categories  = data.get("categories", [])
    series_list = data.get("series", [])

    ws_data.cell(1, 1, "Категория").font = _HEADER_FONT
    ws_data.cell(1, 1).fill = _HEADER_FILL
    for si, s in enumerate(series_list, 2):
        cell      = ws_data.cell(1, si, s.get("name", f"Серия {si - 1}"))
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    for ri, cat in enumerate(categories, 2):
        ws_data.cell(ri, 1, cat)
        for si, s in enumerate(series_list, 2):
            vals = s.get("values", [])
            ws_data.cell(ri, si, vals[ri - 2] if (ri - 2) < len(vals) else None)

    chart = LineChart() if chart_type == "line" else BarChart()
    chart.style          = 10
    chart.title          = title
    chart.y_axis.title   = "Значение"
    if isinstance(chart, BarChart):
        chart.type = "col"

    nrows    = len(categories) + 1
    cats_ref = Reference(ws_data, min_col=1, min_row=2, max_row=nrows)
    for si in range(len(series_list)):
        data_ref = Reference(ws_data, min_col=si + 2, min_row=1, max_row=nrows)
        chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width  = 20
    chart.height = 14

    ws_chart = wb.create_sheet("График")
    ws_chart.add_chart(chart, "B2")
    return wb


def _build_diagram(data: dict[str, Any]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Данные"

    title  = data.get("title", "Диаграмма")
    labels = data.get("labels", [])
    values = data.get("values", [])

    ws.cell(1, 1, "Категория").font = _HEADER_FONT
    ws.cell(1, 1).fill = _HEADER_FILL
    ws.cell(1, 2, "Значение").font = _HEADER_FONT
    ws.cell(1, 2).fill = _HEADER_FILL

    for ri, (label, val) in enumerate(zip(labels, values), 2):
        ws.cell(ri, 1, label)
        ws.cell(ri, 2, val)

    pie       = PieChart()
    pie.title = title
    pie.style = 10

    nrows      = len(labels) + 1
    labels_ref = Reference(ws, min_col=1, min_row=2, max_row=nrows)
    data_ref   = Reference(ws, min_col=2, min_row=1, max_row=nrows)
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(labels_ref)
    pie.width  = 18
    pie.height = 14

    ws_chart = wb.create_sheet("Диаграмма")
    ws_chart.add_chart(pie, "B2")
    return wb


def _build_gantt(data: dict[str, Any]) -> Workbook:
    """Cell-coloring Gantt — same visual style as native Excel Gantt charts."""
    title = data.get("title", "График работ")
    tasks = data.get("tasks", [])
    year  = int(data.get("year", dt.today().year))

    def _parse(s: str) -> dt | None:
        try:
            d, m = s.strip().split(".")
            return dt(year, int(m), int(d))
        except Exception:
            return None

    all_dates = [_parse(t.get(k, "")) for t in tasks for k in ("start", "end")]
    all_dates = [d for d in all_dates if d]
    if not all_dates:
        return _build_table(data)

    min_d, max_d = min(all_dates), max(all_dates)
    dates = []
    cur = min_d
    while cur <= max_d:
        dates.append(cur)
        cur += timedelta(days=1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ганта"
    FIXED = 5  # №, Задача, Объект, Начало, Конец

    # ── Row 1: title ──────────────────────────────────────────────────────────
    total = FIXED + len(dates)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total)
    tc = ws.cell(1, 1, title)
    tc.font      = Font(bold=True, size=13, color="FFFFFF")
    tc.fill      = PatternFill("solid", fgColor="1F3864")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # ── Row 2: month headers (merged per month) ───────────────────────────────
    mfill = PatternFill("solid", fgColor="2E4057")
    mfont = Font(bold=True, color="FFFFFF", size=9)
    month_start: dict = {}
    for i, d in enumerate(dates, FIXED + 1):
        key = (d.year, d.month)
        if key not in month_start:
            month_start[key] = i
    month_items = sorted(month_start.items())
    _RU_MONTHS = {1:"ЯНВАРЬ",2:"ФЕВРАЛЬ",3:"МАРТ",4:"АПРЕЛЬ",5:"МАЙ",6:"ИЮНЬ",
                  7:"ИЮЛЬ",8:"АВГУСТ",9:"СЕНТЯБРЬ",10:"ОКТЯБРЬ",11:"НОЯБРЬ",12:"ДЕКАБРЬ"}
    for idx, ((yr, mo), col_start) in enumerate(month_items):
        col_end = month_items[idx + 1][1] - 1 if idx + 1 < len(month_items) else FIXED + len(dates)
        if col_start < col_end:
            ws.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_end)
        c = ws.cell(2, col_start, f"{_RU_MONTHS[mo]} {yr}")
        c.fill      = mfill
        c.font      = mfont
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # ── Row 3: fixed-col headers + day numbers ────────────────────────────────
    hfill = PatternFill("solid", fgColor="2E4057")
    hfont = Font(bold=True, color="FFFFFF", size=9)
    wknd_hfill = PatternFill("solid", fgColor="C9572A")
    for ci, h in enumerate(["№", "Задача / Событие", "Объект", "Начало", "Конец"], 1):
        c = ws.cell(3, ci, h)
        c.fill      = hfill
        c.font      = hfont
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, d in enumerate(dates, FIXED + 1):
        c = ws.cell(3, i, d.day)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(size=7, bold=d.weekday() >= 5, color="FFFFFF" if d.weekday() >= 5 else "2E4057")
        c.fill = wknd_hfill if d.weekday() >= 5 else PatternFill("solid", fgColor="E8ECF0")
    ws.row_dimensions[3].height = 14

    # ── Task rows ─────────────────────────────────────────────────────────────
    BAR_PALETTE = ["2E75B6", "70AD47", "4472C4", "5B9BD5", "A9D18E", "9DC3E6"]
    MILESTONE_C = "ED7D31"
    WKND_ROW    = PatternFill("solid", fgColor="FFF0E8")
    THIN        = Side(style="thin", color="D0D0D0")
    BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    for ri, task in enumerate(tasks, 4):
        is_ms = task.get("name", "").startswith("•") or task.get("is_milestone", False)
        color = MILESTONE_C if is_ms else BAR_PALETTE[(ri - 4) % len(BAR_PALETTE)]
        bar_fill = PatternFill("solid", fgColor=color)
        row_bg   = PatternFill("solid", fgColor="F7F9FC") if ri % 2 == 0 else None

        ws.cell(ri, 1, task.get("id", ri - 3)).alignment = Alignment(horizontal="center", vertical="center")
        nc = ws.cell(ri, 2, task.get("name", ""))
        nc.font      = Font(bold=is_ms, size=9)
        nc.alignment = Alignment(vertical="center")
        ws.cell(ri, 3, task.get("object", "")).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(ri, 4, task.get("start", "")).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(ri, 5, task.get("end",   "")).alignment = Alignment(horizontal="center", vertical="center")
        for ci in range(1, 6):
            ws.cell(ri, ci).border = BORDER
            if row_bg and not ws.cell(ri, ci).fill.fgColor.rgb != "00000000":
                ws.cell(ri, ci).fill = row_bg

        s, e = _parse(task.get("start", "")), _parse(task.get("end", ""))
        for i, d in enumerate(dates, FIXED + 1):
            cell = ws.cell(ri, i)
            cell.border = BORDER
            if s and e and s <= d <= e:
                cell.fill = bar_fill
            elif d.weekday() >= 5:
                cell.fill = WKND_ROW
        ws.row_dimensions[ri].height = 15

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 9
    ws.column_dimensions["D"].width = 7
    ws.column_dimensions["E"].width = 7
    for i in range(FIXED + 1, FIXED + len(dates) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 2.2

    ws.freeze_panes = "F4"
    return wb


def _build_comparison(data1: dict, data2: dict) -> Workbook:
    """Two tables side-by-side + Diff sheet with highlighted changes."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Источник 1"
    ws2 = wb.create_sheet("Источник 2")
    wsd = wb.create_sheet("Разница")

    def _fill_sheet(ws, data):
        headers = data.get("headers", [])
        rows    = data.get("rows", [])
        title   = data.get("title", "")
        if title:
            ws.merge_cells(start_row=1, start_column=1, end_row=1,
                           end_column=max(len(headers), 1))
            tc = ws.cell(1, 1, title)
            tc.font = Font(bold=True, size=12)
            tc.alignment = Alignment(horizontal="center")
        for ci, h in enumerate(headers, 1):
            c = ws.cell(2, ci, h)
            c.fill = _HEADER_FILL
            c.font = _HEADER_FONT
            c.alignment = Alignment(horizontal="center")
        for ri, row in enumerate(rows, 3):
            for ci, val in enumerate(row, 1):
                ws.cell(ri, ci, val)
        _auto_col_width(ws, max(len(headers), 1), len(rows))

    _fill_sheet(ws1, data1)
    _fill_sheet(ws2, data2)

    # Diff: compare row-by-row by first column value
    rows1 = {str(r[0]): r for r in data1.get("rows", []) if r}
    rows2 = {str(r[0]): r for r in data2.get("rows", []) if r}
    all_keys = list(dict.fromkeys(list(rows1) + list(rows2)))
    headers = data1.get("headers", data2.get("headers", []))

    RED_FILL   = PatternFill("solid", fgColor="FFCCCC")
    GREEN_FILL = PatternFill("solid", fgColor="CCFFCC")
    YELLOW_FILL = PatternFill("solid", fgColor="FFFACC")

    for ci, h in enumerate(headers, 1):
        c = wsd.cell(1, ci, h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center")

    for ri, key in enumerate(all_keys, 2):
        r1 = rows1.get(key)
        r2 = rows2.get(key)
        row = r2 or r1
        for ci, val in enumerate(row or [], 1):
            c = wsd.cell(ri, ci, val)
            if r1 is None:
                c.fill = GREEN_FILL   # added in source 2
            elif r2 is None:
                c.fill = RED_FILL     # removed in source 2
            elif r1 != r2 and ci < len(r1) and ci < len(r2) and r1[ci-1] != r2[ci-1]:
                c.fill = YELLOW_FILL  # changed value

    _auto_col_width(wsd, len(headers), len(all_keys))
    return wb


_BUILDERS: dict[str, Any] = {
    "table":   _build_table,
    "chart":   _build_chart,
    "diagram": _build_diagram,
    "gantt":   _build_gantt,
}

_TYPE_RU: dict[str, str] = {
    "table":   "таблица",
    "chart":   "график",
    "diagram": "диаграмма",
    "gantt":   "ганта",
    "auto":    "авто",
}


# ── Routes ────────────────────────────────────────────────────────────────────

@router.get("/tools/photo-to-excel", response_class=HTMLResponse)
async def photo_to_excel_page(request: Request, user: dict = Depends(require_admin)):
    return templates.TemplateResponse("photo_to_excel.html", {
        "request": request, "user": user,
    })


@router.post("/api/tools/photo-to-excel")
async def photo_to_excel_api(
    request:     Request,
    user:        dict       = Depends(require_admin),
    photo:       UploadFile = File(...),
    output_type: str        = Form("table"),
):
    try:
        if output_type not in _BUILDERS:
            output_type = "table"

        image_bytes  = await photo.read()
        content_type = photo.content_type or "image/jpeg"
        if content_type not in ("image/jpeg", "image/png", "image/gif", "image/webp"):
            content_type = "image/jpeg"

        data = await _call_ai(image_bytes, content_type, output_type)
        wb   = _BUILDERS[output_type](data)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        raw_title   = data.get("title", "data")
        ascii_title = re.sub(r"[^\w\- ]", "", raw_title, flags=re.ASCII)[:40].strip() or "data"
        filename    = f"{ascii_title}_{output_type}.xlsx"
        encoded     = urllib.parse.quote(f"{raw_title[:40]}_{_TYPE_RU[output_type]}.xlsx")
        disposition = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{encoded}'

        blob = buf.getvalue()
        _history_save(user["id"], filename, output_type, blob)

        return StreamingResponse(
            io.BytesIO(blob),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": disposition},
        )
    except Exception as exc:
        logger.error("photo-to-excel error: %s", exc, exc_info=True)
        return JSONResponse({"error": str(exc)}, status_code=500)


@router.post("/api/tools/photo-to-excel/preview")
async def photo_to_excel_preview(
    request:     Request,
    user:        dict       = Depends(require_admin),
    photo:       UploadFile = File(...),
    output_type: str        = Form("auto"),
):
    """Return extracted JSON for frontend preview before downloading Excel."""
    try:
        image_bytes  = await photo.read()
        content_type = photo.content_type or "image/jpeg"
        if content_type not in ("image/jpeg", "image/png", "image/gif", "image/webp"):
            content_type = "image/jpeg"

        otype = output_type if output_type in _PROMPTS else "auto"
        data  = await _call_ai(image_bytes, content_type, otype)

        # Auto-detection: resolve detected_type
        if otype == "auto":
            detected = data.get("detected_type", "table")
            data["_resolved_type"] = detected
        else:
            data["_resolved_type"] = otype

        return JSONResponse({"ok": True, "data": data})
    except Exception as exc:
        logger.error("preview error: %s", exc, exc_info=True)
        return JSONResponse({"error": str(exc)}, status_code=500)


@router.post("/api/tools/text-to-excel")
async def text_to_excel_api(
    request:     Request,
    user:        dict = Depends(require_admin),
    text:        str  = Form(...),
    output_type: str  = Form("table"),
):
    """Convert pasted/typed text → Excel using text-only AI."""
    try:
        if not text.strip():
            return JSONResponse({"error": "Текст не может быть пустым"}, status_code=400)

        otype = output_type if output_type in _PROMPTS else "table"
        prompt = _TEXT_PROMPTS.get(otype, _TEXT_PROMPTS["table"])
        full_prompt = f"{prompt}\n\nТекст для анализа:\n{text.strip()[:8000]}"

        data = await _call_text_ai(full_prompt)

        if otype == "auto":
            otype = data.get("detected_type", "table")
            if otype not in _BUILDERS:
                otype = "table"

        wb  = _BUILDERS.get(otype, _build_table)(data)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        raw_title   = data.get("title", "data")
        ascii_title = re.sub(r"[^\w\- ]", "", raw_title, flags=re.ASCII)[:40].strip() or "data"
        filename    = f"{ascii_title}_{otype}.xlsx"
        encoded     = urllib.parse.quote(f"{raw_title[:40]}_{_TYPE_RU.get(otype, otype)}.xlsx")
        disposition = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{encoded}'

        blob = buf.getvalue()
        _history_save(user["id"], filename, otype, blob)

        return StreamingResponse(
            io.BytesIO(blob),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": disposition},
        )
    except Exception as exc:
        logger.error("text-to-excel error: %s", exc, exc_info=True)
        return JSONResponse({"error": str(exc)}, status_code=500)


@router.post("/api/tools/multi-photo-excel")
async def multi_photo_excel_api(
    request:     Request,
    user:        dict             = Depends(require_admin),
    photos:      List[UploadFile] = File(...),
    output_type: str              = Form("table"),
):
    """Multiple photos → one Excel file, each photo becomes a separate sheet."""
    try:
        if not photos:
            return JSONResponse({"error": "Нет файлов"}, status_code=400)

        otype = output_type if output_type in _PROMPTS else "table"
        wb_final = Workbook()
        wb_final.remove(wb_final.active)

        tasks_list = []
        for photo in photos[:10]:  # max 10
            image_bytes  = await photo.read()
            content_type = photo.content_type or "image/jpeg"
            if content_type not in ("image/jpeg", "image/png", "image/gif", "image/webp"):
                content_type = "image/jpeg"
            tasks_list.append(_call_ai(image_bytes, content_type, otype))

        results = await asyncio.gather(*tasks_list, return_exceptions=True)

        for idx, (photo, result) in enumerate(zip(photos, results), 1):
            if isinstance(result, Exception):
                logger.warning("multi-photo sheet %d error: %s", idx, result)
                ws = wb_final.create_sheet(f"Лист {idx} (ошибка)")
                ws.cell(1, 1, f"Ошибка обработки: {result}")
                continue
            wb_sheet = _BUILDERS.get(otype, _build_table)(result)
            src_ws   = wb_sheet.active
            sheet_name = re.sub(r"[^\w\s]", "", photo.filename or f"Лист {idx}")[:30] or f"Лист {idx}"
            dst_ws = wb_final.create_sheet(sheet_name)
            for row in src_ws.iter_rows():
                for cell in row:
                    nc = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        nc.font   = cell.font.copy()
                        nc.fill   = cell.fill.copy()
                        nc.border = cell.border.copy()
                        nc.alignment = cell.alignment.copy()

        if not wb_final.sheetnames:
            return JSONResponse({"error": "Не удалось обработать ни один файл"}, status_code=500)

        buf = io.BytesIO()
        wb_final.save(buf)
        buf.seek(0)

        blob = buf.getvalue()
        fname = f"multi_{len(photos)}_sheets_{otype}.xlsx"
        _history_save(user["id"], fname, otype, blob)

        return StreamingResponse(
            io.BytesIO(blob),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )
    except Exception as exc:
        logger.error("multi-photo error: %s", exc, exc_info=True)
        return JSONResponse({"error": str(exc)}, status_code=500)


@router.post("/api/tools/compare-excel")
async def compare_excel_api(
    request: Request,
    user:    dict       = Depends(require_admin),
    photo1:  UploadFile = File(...),
    photo2:  UploadFile = File(...),
):
    """Two photos → Excel with 3 sheets: Source1, Source2, Diff (color-coded)."""
    try:
        b1, b2 = await photo1.read(), await photo2.read()
        m1 = photo1.content_type or "image/jpeg"
        m2 = photo2.content_type or "image/jpeg"
        for m in [m1, m2]:
            if m not in ("image/jpeg", "image/png", "image/gif", "image/webp"):
                m = "image/jpeg"

        data1, data2 = await asyncio.gather(
            _call_ai(b1, m1, "table"),
            _call_ai(b2, m2, "table"),
        )
        wb  = _build_comparison(data1, data2)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        blob  = buf.getvalue()
        fname = "сравнение.xlsx"
        _history_save(user["id"], fname, "compare", blob)

        return StreamingResponse(
            io.BytesIO(blob),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename*=UTF-8\'\'{urllib.parse.quote(fname)}'},
        )
    except Exception as exc:
        logger.error("compare error: %s", exc, exc_info=True)
        return JSONResponse({"error": str(exc)}, status_code=500)


@router.post("/api/tools/merge-excel")
async def merge_excel_api(
    request:     Request,
    user:        dict             = Depends(require_admin),
    photos:      List[UploadFile] = File(...),
):
    """Multiple photos → one merged table (AI deduplicates and unifies structure)."""
    try:
        if not photos:
            return JSONResponse({"error": "Нет файлов"}, status_code=400)

        tasks = [_call_ai(await p.read(), p.content_type or "image/jpeg", "table")
                 for p in photos[:5]]
        results = await asyncio.gather(*tasks, return_exceptions=True)
        tables  = [r for r in results if not isinstance(r, Exception)]

        if not tables:
            return JSONResponse({"error": "Не удалось обработать файлы"}, status_code=500)

        if len(tables) == 1:
            merged = tables[0]
        else:
            # Send all tables to AI to merge
            merge_prompt = (
                "У тебя есть несколько таблиц. Объедини их в одну, убери дублирующиеся строки, "
                "унифицируй заголовки. Верни ТОЛЬКО валидный JSON:\n"
                '{"title":"Объединённая таблица","headers":[...],"rows":[[...]]}\n\n'
                "Таблицы:\n" +
                "\n".join(f"Таблица {i+1}:\n{json.dumps(t, ensure_ascii=False)}" for i, t in enumerate(tables))
            )
            merged = await _call_text_ai(merge_prompt)

        wb  = _build_table(merged)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        blob  = buf.getvalue()
        fname = "объединение.xlsx"
        _history_save(user["id"], fname, "merge", blob)

        return StreamingResponse(
            io.BytesIO(blob),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename*=UTF-8\'\'{urllib.parse.quote(fname)}'},
        )
    except Exception as exc:
        logger.error("merge error: %s", exc, exc_info=True)
        return JSONResponse({"error": str(exc)}, status_code=500)


@router.get("/api/tools/history")
async def conversion_history(request: Request, user: dict = Depends(require_admin)):
    """Return last N conversion history items for current user."""
    now = _time.time()
    items = [
        {"id": i["id"], "filename": i["filename"], "type": i["type"],
         "size": i["size"], "age_min": int((now - i["ts"]) / 60)}
        for i in _HISTORY.get(user["id"], [])
        if now - i["ts"] < _HISTORY_TTL
    ]
    return JSONResponse({"items": list(reversed(items))})


@router.get("/api/tools/history/{item_id}")
async def download_history_item(item_id: str, request: Request,
                                user: dict = Depends(require_admin)):
    """Re-download a previously converted file."""
    item = _history_get(user["id"], item_id)
    if not item:
        return JSONResponse({"error": "Файл не найден или истёк срок хранения"}, status_code=404)
    encoded = urllib.parse.quote(item["filename"])
    return StreamingResponse(
        io.BytesIO(item["blob"]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{item["filename"]}"; filename*=UTF-8\'\'{encoded}'},
    )


# ── Templates ─────────────────────────────────────────────────────────────────

@router.get("/api/tools/templates")
async def list_templates(request: Request, user: dict = Depends(require_admin)):
    """List saved conversion templates for current user."""
    return JSONResponse({"templates": _TEMPLATES.get(user["id"], [])})


@router.post("/api/tools/templates")
async def save_template(request: Request, user: dict = Depends(require_admin)):
    """Save a named template (output_type + mode)."""
    data = await request.json()
    name = (data.get("name") or "").strip()[:60]
    if not name:
        return JSONResponse({"error": "Название обязательно"}, status_code=400)
    output_type = data.get("output_type", "auto")
    mode        = data.get("mode", "photo")
    if output_type not in _PROMPTS:
        output_type = "auto"
    uid = _template_save(user["id"], name, output_type, mode)
    return JSONResponse({"ok": True, "id": uid})


@router.delete("/api/tools/templates/{template_id}")
async def delete_template(template_id: str, request: Request,
                          user: dict = Depends(require_admin)):
    """Delete a saved template by ID."""
    ok = _template_delete(user["id"], template_id)
    if not ok:
        return JSONResponse({"error": "Шаблон не найден"}, status_code=404)
    return JSONResponse({"ok": True})
