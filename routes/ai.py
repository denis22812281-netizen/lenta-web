import io
import os
from datetime import date

import openpyxl
from fastapi import APIRouter, Depends, File, Request, UploadFile
from fastapi.responses import HTMLResponse
from sqlalchemy.orm import Session

import models
from database import get_db
from deps import get_current_user, require_api_user, require_login, templates
from utils.files import read_limited

_MAX_AI_EXCEL_BYTES = 10 * 1024 * 1024  # 10 MB

router = APIRouter()


@router.get("/ai", response_class=HTMLResponse)
async def ai_page(request: Request, user: dict = Depends(require_login)):
    return templates.TemplateResponse("ai_chat.html", {"request": request, "user": user})


@router.post("/api/ai/chat")
async def ai_chat(request: Request, db: Session = Depends(get_db),
                  user: dict = Depends(require_api_user)):
    body = await request.json()
    user_message = body.get("message", "").strip()
    if not user_message:
        return {"error": "Пустое сообщение"}
    provider = body.get("provider", "groq")
    history  = body.get("history", [])

    today    = date.today()
    all_projects = db.query(models.Project).all()
    active   = [p for p in all_projects if p.status == "Активный"]
    done     = [p for p in all_projects if p.status == "Завершён"]
    urgent   = [p for p in active if p.end_date and 0 <= (p.end_date - today).days <= 7
                and not (p.opening_date and p.opening_date <= today)]
    overdue_p = [p for p in active if p.end_date and p.end_date < today
                 and not (p.opening_date and p.opening_date <= today)]
    overdue_tasks = db.query(models.Task).filter(
        models.Task.status != "Завершена",
        models.Task.deadline != None,
        models.Task.deadline < today).limit(15).all()
    open_tasks = db.query(models.Task).filter(models.Task.status != "Завершена").count()
    managers   = db.query(models.Manager).all()

    ctx = f"""Сегодня: {today.strftime('%d.%m.%Y')}
ПРОЕКТЫ: Всего {len(all_projects)} | Активных {len(active)} | Завершённых {len(done)}
Реконструкции активных: {sum(1 for p in active if p.project_type=='Реконструкция')}
Констракшн активных: {sum(1 for p in active if p.project_type=='Констракшн')}
СРОЧНЫЕ (≤7 дней): {len(urgent)}
{chr(10).join(f'  • {p.name} | {p.manager.name if p.manager else "—"} | {p.end_date.strftime("%d.%m.%Y")}' for p in urgent[:8])}
ПРОСРОЧЕННЫЕ проекты: {len(overdue_p)}
{chr(10).join(f'  • {p.name} | -{(today-p.end_date).days} дн' for p in overdue_p[:5])}
ЗАДАЧИ: открытых {open_tasks}, просроченных {len(overdue_tasks)}
{chr(10).join(f'  • {t.title} | {t.assignee.name if t.assignee else "—"} | -{(today-t.deadline).days} дн' for t in overdue_tasks[:5])}
МЕНЕДЖЕРЫ ({len(managers)} чел):
{chr(10).join(f'  • {m.name}{"  [руководитель]" if m.is_leader else ""}: активных {sum(1 for p in m.projects if p.status=="Активный")}' for m in managers)}"""

    system_prompt = (
        "Ты — ИИ-ассистент системы управления проектами компании ЛЕНТА (сеть гипермаркетов России). "
        "Помогаешь команде из 9 человек: руководитель Гаврин Игорь + 8 менеджеров. "
        "Анализируешь статусы, дедлайны, риски. Отвечай на русском языке, используй markdown. "
        f"\n\n=== ТЕКУЩИЕ ДАННЫЕ ===\n{ctx}"
    )

    msg_history = []
    for h in history[:-1]:
        if h.get("role") in ("user", "assistant"):
            msg_history.append({"role": h["role"], "content": h.get("content", "")})
    msg_history.append({"role": "user", "content": user_message})

    reply = ""
    try:
        if provider == "claude":
            import anthropic as ant
            api_key = os.getenv("ANTHROPIC_API_KEY", "")
            if not api_key:
                return {"reply": "⚠️ ANTHROPIC_API_KEY не задан."}
            resp = ant.Anthropic(api_key=api_key).messages.create(
                model="claude-sonnet-4-6", max_tokens=2048,
                system=system_prompt, messages=msg_history)
            reply = resp.content[0].text
        elif provider == "groq":
            from openai import OpenAI
            api_key = os.getenv("GROQ_API_KEY", "")
            if not api_key:
                return {"reply": "⚠️ GROQ_API_KEY не задан."}
            resp = OpenAI(api_key=api_key, base_url="https://api.groq.com/openai/v1"
                          ).chat.completions.create(
                model="llama-3.3-70b-versatile", max_tokens=2048,
                messages=[{"role": "system", "content": system_prompt}] + msg_history)
            reply = resp.choices[0].message.content
        elif provider == "deepseek":
            from openai import OpenAI
            api_key = os.getenv("DEEPSEEK_API_KEY", "")
            if not api_key:
                return {"reply": "⚠️ DEEPSEEK_API_KEY не задан."}
            resp = OpenAI(api_key=api_key, base_url="https://api.deepseek.com"
                          ).chat.completions.create(
                model="deepseek-chat", max_tokens=2048,
                messages=[{"role": "system", "content": system_prompt}] + msg_history)
            reply = resp.choices[0].message.content
    except Exception as e:
        reply = f"⚠️ Ошибка ({provider}): {str(e)[:300]}"

    if reply and not reply.startswith("⚠️"):
        user_name = user.get("display_name", "")
        db.add(models.AiChatMessage(user_name=user_name, role="user",
                                    text=user_message, provider=provider))
        db.add(models.AiChatMessage(user_name=user_name, role="assistant",
                                    text=reply, provider=provider))
        db.commit()
    return {"reply": reply}


@router.get("/api/ai/history")
async def ai_history(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return {"messages": []}
    user_name = user.get("display_name", "")
    msgs = db.query(models.AiChatMessage).filter(
        models.AiChatMessage.user_name == user_name
    ).order_by(models.AiChatMessage.id.desc()).limit(100).all()
    msgs.reverse()
    return {"messages": [
        {"role": m.role, "text": m.text, "time": m.created_at.strftime("%H:%M")}
        for m in msgs
    ]}


@router.post("/api/ai/clear-history")
async def ai_clear_history(request: Request, db: Session = Depends(get_db),
                            user: dict = Depends(require_api_user)):
    db.query(models.AiChatMessage).filter(
        models.AiChatMessage.user_name == user.get("display_name", "")).delete()
    db.commit()
    return {"ok": True}


@router.post("/api/ai/check-excel")
async def ai_check_excel(request: Request, db: Session = Depends(get_db),
                          user: dict = Depends(require_api_user),
                          file: UploadFile = File(...)):
    api_key = os.getenv("ANTHROPIC_API_KEY", "")
    if not api_key:
        return {"report": "ANTHROPIC_API_KEY не задан."}
    try:
        content = await read_limited(file, _MAX_AI_EXCEL_BYTES)
    except ValueError:
        return {"report": "Файл слишком большой (макс 10 МБ)"}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
        issues = []
        for sheet_name in wb.sheetnames:
            for row in wb[sheet_name].iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        issues.append(f"Лист '{sheet_name}', {cell.coordinate}: {cell.value[:60]}")
        formula_text = "\n".join(issues[:50]) if issues else "Формул не обнаружено."
    except Exception as e:
        return {"report": f"Ошибка чтения файла: {e}"}
    try:
        import anthropic
        response = anthropic.Anthropic(api_key=api_key).messages.create(
            model="claude-sonnet-4-6", max_tokens=2048,
            system="Ты — эксперт по Excel. Проверяй формулы, отвечай на русском.",
            messages=[{"role": "user", "content":
                f"Проверь формулы:\n{formula_text}\nСоставь краткий отчёт о проблемах."}])
        report = response.content[0].text
    except Exception as e:
        report = f"Ошибка ИИ: {str(e)[:200]}"
    return {"report": report, "formulas_found": len(issues)}
