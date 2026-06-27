import asyncio
import io
import json
import logging as _logging
import os
from datetime import date, datetime, timedelta
from pathlib import Path
from urllib.parse import quote

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

import models
from database import get_db
from deps import get_current_user, require_admin, require_api_user, require_login, templates
from services.cache import cache_delete, cache_get, cache_set
from services.cloud_storage import upload_photo
from services.email_service import notify_opening_photos, notify_precheck_report, notify_vpk_report
from utils.files import check_magic_bytes

_vpk_logger = _logging.getLogger(__name__)

# Список email для VPK-уведомлений. Формат: email:Имя,email2:Имя2
_VPK_NOTIFY_EMAILS = []
for _entry in os.getenv("NOTIFY_VPK_EMAILS", "").split(","):
    _entry = _entry.strip()
    if not _entry:
        continue
    if ":" in _entry:
        _e, _n = _entry.split(":", 1)
        _VPK_NOTIFY_EMAILS.append((_e.strip(), _n.strip()))
    else:
        _VPK_NOTIFY_EMAILS.append((_entry, _entry.split("@")[0]))
_vpk_logger.warning("VPK: NOTIFY_VPK_EMAILS = %s", _VPK_NOTIFY_EMAILS)

# Email для предосмотра и фото открытия — только один получатель
_PRECHECK_EMAIL = os.getenv("NOTIFY_PRECHECK_EMAIL", "").strip()
_OPENING_EMAIL  = _PRECHECK_EMAIL  # тот же env var

router = APIRouter()


# ─── ВПК ─────────────────────────────────────────────────────────────────────

@router.get("/vpk", response_class=HTMLResponse)
async def vpk_page(request: Request, db: Session = Depends(get_db),
                   user: dict = Depends(require_login), tab: str = "vpk1"):
    from types import SimpleNamespace

    # VPK criteria change rarely — cache 10 minutes
    c1_cached = await cache_get("vpk:criteria:1")
    c2_cached = await cache_get("vpk:criteria:2")
    if c1_cached is None:
        criteria1 = db.query(models.VpkCriterion).filter(
            models.VpkCriterion.vpk_type == 1).order_by(models.VpkCriterion.order).all()
        await cache_set("vpk:criteria:1",
                        [{"id": c.id, "name": c.name, "order": c.order} for c in criteria1], ttl=600)
    else:
        criteria1 = [SimpleNamespace(**d) for d in c1_cached]
    if c2_cached is None:
        criteria2 = db.query(models.VpkCriterion).filter(
            models.VpkCriterion.vpk_type == 2).order_by(models.VpkCriterion.order).all()
        await cache_set("vpk:criteria:2",
                        [{"id": c.id, "name": c.name, "order": c.order} for c in criteria2], ttl=600)
    else:
        criteria2 = [SimpleNamespace(**d) for d in c2_cached]

    projects  = db.query(models.Project).filter(
        models.Project.project_type == "Констракшн").order_by(models.Project.tk_number).all()
    reports   = db.query(models.VpkReport).order_by(
        models.VpkReport.submitted_at.desc()).limit(50).all()
    pre_reports = db.query(models.PreVpkReport).order_by(
        models.PreVpkReport.submitted_at.desc()).limit(20).all()
    name = user.get("display_name", "")
    total_reports = db.query(models.VpkReport).count()
    read_by_me = db.query(models.VpkReportRead).filter(
        models.VpkReportRead.reader_name == name).count()
    unread = max(0, total_reports - read_by_me)
    vpk_t = int(request.query_params.get("vpk_t", 1))
    precheck_criteria = criteria1 if vpk_t == 1 else criteria2
    return templates.TemplateResponse("vpk.html", {
        "request": request, "user": user,
        "tab": tab, "projects": projects,
        "criteria1": criteria1, "criteria2": criteria2,
        "reports": reports, "unread": unread,
        "pre_reports": pre_reports,
        "vpk_t": vpk_t,
        "precheck_criteria": precheck_criteria,
        "msg": request.query_params.get("msg"),
    })


@router.post("/vpk/submit")
async def vpk_submit(request: Request, background_tasks: BackgroundTasks,
                     db: Session = Depends(get_db), user: dict = Depends(require_login)):
    form      = await request.form()
    project_id = form.get("project_id")
    vpk_type   = int(form.get("vpk_type", 1))
    report = models.VpkReport(
        project_id=int(project_id) if project_id else None,
        vpk_type=vpk_type,
        submitted_by=user.get("display_name", ""),
        read_gavrin=False, read_mesmer=False,
    )
    db.add(report)
    db.flush()
    criteria = db.query(models.VpkCriterion).filter(
        models.VpkCriterion.vpk_type == vpk_type
    ).order_by(models.VpkCriterion.order).all()
    for c in criteria:
        is_done = form.get(f"criterion_{c.id}") == "on"
        comment = str(form.get(f"comment_{c.id}", "") or "").strip()
        photo_path = ""
        photo_file = form.get(f"photo_{c.id}")
        if photo_file and hasattr(photo_file, "filename") and photo_file.filename:
            ext = Path(photo_file.filename).suffix.lower() or ".jpg"
            content = await photo_file.read()
            try:
                check_magic_bytes(content, photo_file.filename)
            except ValueError:
                continue  # skip invalid file, don't block the whole report
            fname = f"{report.id}_{c.id}{ext}"
            photo_path = upload_photo(content, "vpk", fname)
        db.add(models.VpkReportItem(
            report_id=report.id, criterion_id=c.id,
            criterion_name=c.name, done=is_done,
            comment=comment, photo_path=photo_path,
        ))
    db.commit()
    proj = db.query(models.Project).filter(models.Project.id == project_id).first()
    tk   = proj.tk_number if proj else "—"
    proj_name = proj.name if proj else ""

    # Email-уведомления — всем кто имеет email: лидеры + менеджер проекта
    items        = db.query(models.VpkReportItem).filter(
        models.VpkReportItem.report_id == report.id).all()
    done         = sum(1 for i in items if i.done)
    total        = len(items)
    submitted_at = report.submitted_at.strftime("%d.%m.%Y %H:%M") if report.submitted_at else ""
    submitter    = user.get("display_name", "")
    failed_items = [
        {"name": i.criterion_name, "comment": i.comment or "", "photo_path": i.photo_path or ""}
        for i in items if not i.done
    ]

    recipients = {}
    # Получатели: отправитель + менеджер проекта (лидеры временно отключены)
    for mgr in db.query(models.Manager).filter(
            models.Manager.email != "", models.Manager.email.isnot(None)).all():
        if mgr.name == submitter or (proj and proj.manager_id == mgr.id):
            recipients[mgr.email] = mgr.name

    _vpk_logger.info("VPK submit: отправка email получателям %s", list(recipients.keys()))

    def _send_emails():
        for email, name in recipients.items():
            notify_vpk_report(
                to_email=email, recipient_name=name,
                vpk_type=vpk_type, tk_number=tk, project_name=proj_name,
                submitted_by=submitter, done=done, total=total,
                submitted_at=submitted_at, failed_items=failed_items,
            )

    background_tasks.add_task(_send_emails)

    return RedirectResponse(
        f"/vpk?tab=reports&msg={quote(f'Отчёт ВПК{vpk_type} по ТК {tk} отправлен')}",
        status_code=303)


@router.post("/vpk/precheck")
async def precheck_submit(request: Request, background_tasks: BackgroundTasks,
                          db: Session = Depends(get_db), user: dict = Depends(require_login)):
    form      = await request.form()
    project_id = form.get("project_id")
    vpk_type   = int(form.get("vpk_type", 1))

    report = models.PreVpkReport(
        project_id=int(project_id) if project_id else None,
        vpk_type=vpk_type,
        submitted_by=user.get("display_name", ""),
    )
    db.add(report)
    db.flush()

    criteria = db.query(models.VpkCriterion).filter(
        models.VpkCriterion.vpk_type == vpk_type
    ).order_by(models.VpkCriterion.order).all()

    _all_keys = [k for k, v in form.multi_items() if isinstance(v, str)]
    _vpk_logger.warning("PRECHECK all_keys(%d): %s", len(_all_keys), _all_keys[:30])
    _vpk_logger.warning("PRECHECK form: project_id=%r tk_text=%r",
                        project_id, form.get("tk_text"))
    precheck_json = str(form.get("precheck_json", "{}") or "{}")
    try:
        precheck_states = json.loads(precheck_json)
    except Exception:
        precheck_states = {}
    _vpk_logger.warning("PRECHECK JSON received (%d keys): %s",
                        len(precheck_states), precheck_states)

    for c in criteria:
        state_val = precheck_states.get(str(c.id), "not_checked")
        status = state_val if state_val in ("done", "not_done", "not_checked") else "not_checked"
        comment = str(form.get(f"comment_{c.id}", "") or "").strip()
        photo_path = ""
        if status == "not_done":
            photo_file = form.get(f"photo_{c.id}")
            if photo_file and hasattr(photo_file, "filename") and photo_file.filename:
                ext = Path(photo_file.filename).suffix.lower() or ".jpg"
                fname = f"pre_{report.id}_{c.id}{ext}"
                photo_path = upload_photo(await photo_file.read(), "precheck", fname)
        db.add(models.PreVpkReportItem(
            report_id=report.id, criterion_id=c.id,
            criterion_name=c.name, status=status,
            comment=comment, photo_path=photo_path,
        ))

    db.commit()

    if project_id:
        proj = db.query(models.Project).filter(models.Project.id == int(project_id)).first()
    else:
        tk_text = str(form.get("tk_text", "") or "").strip()
        proj = db.query(models.Project).filter(models.Project.tk_number == tk_text).first() if tk_text else None
    tk   = proj.tk_number if proj else "—"
    proj_name = proj.name if proj else ""
    vpk_date_str = proj.vpk_date.strftime("%d.%m.%Y") if (proj and proj.vpk_date) else ""

    items       = db.query(models.PreVpkReportItem).filter(
        models.PreVpkReportItem.report_id == report.id).all()
    failed_items = [
        {"name": i.criterion_name, "comment": i.comment or "", "photo_path": i.photo_path or ""}
        for i in items if i.status == "not_done"
    ]
    ok_count   = sum(1 for i in items if i.status == "done")
    skip_count = sum(1 for i in items if i.status == "not_checked")

    # Получатели: фиксированный email (Denis) + email отправителя из таблицы менеджеров
    submitter = user.get("display_name", "")
    to_emails: set[str] = set()
    if _PRECHECK_EMAIL:
        to_emails.add(_PRECHECK_EMAIL)
    sub_mgr = db.query(models.Manager).filter(
        models.Manager.name == submitter,
        models.Manager.email.isnot(None),
        models.Manager.email != "",
    ).first()
    if sub_mgr:
        to_emails.add(sub_mgr.email)

    if to_emails:
        for em in to_emails:
            background_tasks.add_task(
                notify_precheck_report, em, vpk_type, tk, proj_name,
                submitter, vpk_date_str, failed_items, ok_count, skip_count,
            )
    else:
        _vpk_logger.warning("Precheck: email не отправлен — NOTIFY_PRECHECK_EMAIL не задан и у менеджера нет email")

    return RedirectResponse(
        f"/vpk?tab=precheck&msg={quote(f'Предосмотр ВПК{vpk_type} по ТК {tk} отправлен')}",
        status_code=303,
    )


@router.post("/vpk/reports/{report_id}/read")
async def vpk_mark_read(report_id: int, request: Request, db: Session = Depends(get_db),
                        user: dict = Depends(require_api_user)):
    report = db.query(models.VpkReport).filter(models.VpkReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404)
    name = user.get("display_name", "")
    existing = db.query(models.VpkReportRead).filter(
        models.VpkReportRead.report_id == report_id,
        models.VpkReportRead.reader_name == name,
    ).first()
    if not existing:
        db.add(models.VpkReportRead(report_id=report_id, reader_name=name))
        db.commit()
    return {"ok": True}


@router.get("/api/vpk/unread")
async def vpk_unread(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request)
    if not user:
        return {"reports": []}
    name = user.get("display_name", "")
    if not name:
        return {"reports": []}
    read_ids = {r.report_id for r in db.query(models.VpkReportRead).filter(
        models.VpkReportRead.reader_name == name).all()}
    reports = db.query(models.VpkReport).filter(
        models.VpkReport.id.notin_(read_ids)).all()
    result = []
    for r in reports:
        tk   = r.project.tk_number if r.project else "—"
        done = sum(1 for i in r.items if i.done)
        result.append({
            "id": r.id,
            "title": f"Новый отчёт ВПК{r.vpk_type} по ТК {tk}",
            "body": f"Подал: {r.submitted_by} | Выполнено: {done}/{len(r.items)} критериев",
        })
    return {"reports": result}


# ─── Отчёты ВПК ──────────────────────────────────────────────────────────────

@router.get("/reports", response_class=HTMLResponse)
async def reports_page(request: Request, db: Session = Depends(get_db),
                       user: dict = Depends(require_login),
                       vpk_type: str = "", manager_name: str = ""):
    q = db.query(models.VpkReport).order_by(models.VpkReport.submitted_at.desc())
    if vpk_type in ("1", "2"):
        q = q.filter(models.VpkReport.vpk_type == int(vpk_type))
    if manager_name:
        q = q.filter(models.VpkReport.submitted_by == manager_name)
    reports = q.limit(100).all()
    authors = sorted({a[0] for a in db.query(models.VpkReport.submitted_by).distinct().all() if a[0]})
    return templates.TemplateResponse("reports.html", {
        "request": request, "user": user,
        "reports": reports, "authors": authors,
        "filter_vpk_type": vpk_type, "filter_manager": manager_name,
    })


@router.get("/reports/export")
async def reports_export(request: Request, db: Session = Depends(get_db),
                         user: dict = Depends(require_login),
                         date_from: str = "", date_to: str = "", vpk_type: str = ""):
    q = db.query(models.VpkReport).order_by(models.VpkReport.submitted_at.desc())
    if vpk_type in ("1", "2"):
        q = q.filter(models.VpkReport.vpk_type == int(vpk_type))
    if date_from:
        try:
            q = q.filter(models.VpkReport.submitted_at >=
                         datetime.strptime(date_from, "%Y-%m-%d"))
        except ValueError:
            pass
    if date_to:
        try:
            q = q.filter(models.VpkReport.submitted_at <
                         datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1))
        except ValueError:
            pass
    reports = q.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёты ВПК"
    hfill  = PatternFill(start_color="1A5C22", end_color="1A5C22", fill_type="solid")
    hfont  = Font(color="FFFFFF", bold=True, size=12)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap   = Alignment(vertical="center", wrap_text=True)
    headers    = ["Менеджер", "Номер ТК", "Критерий ВПК", "Комментарий"]
    col_widths = [25, 12, 60, 50]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont; c.alignment = center
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"
    row_num = 2
    for r in reports:
        tk  = r.project.tk_number if r.project else "—"
        mgr = r.project.manager.name if (r.project and r.project.manager) else "—"
        for item in r.items:
            ws.cell(row=row_num, column=1, value=mgr).alignment = wrap
            ws.cell(row=row_num, column=2, value=tk).alignment  = center
            ws.cell(row=row_num, column=3, value=item.criterion_name).alignment = wrap
            ws.cell(row=row_num, column=4, value=item.comment or "").alignment  = wrap
            ws.row_dimensions[row_num].height = 18
            row_num += 1
        row_num += 1
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    period = f"{date_from or 'all'}_{date_to or 'now'}"
    return StreamingResponse(output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=vpk_reports_{period}.xlsx"})


@router.post("/reports/clear-all")
async def clear_all_reports(request: Request, db: Session = Depends(get_db),
                             user: dict = Depends(require_admin)):
    db.query(models.VpkReportItem).delete()
    db.query(models.VpkReport).delete()
    db.commit()
    return RedirectResponse("/reports", status_code=303)


# ─── Фото открытия ───────────────────────────────────────────────────────────

@router.get("/opening", response_class=HTMLResponse)
async def opening_page(request: Request, db: Session = Depends(get_db),
                       user: dict = Depends(require_login)):
    """Страница управления фото открытий."""
    projects = db.query(models.Project).filter(
        models.Project.project_type == "Констракшн"
    ).order_by(models.Project.tk_number).all()
    selected_id = request.query_params.get("project_id")
    selected_proj = None
    photos = []
    if selected_id:
        selected_proj = db.query(models.Project).filter(
            models.Project.id == int(selected_id)).first()
        if selected_proj:
            photos = db.query(models.OpeningPhoto).filter(
                models.OpeningPhoto.project_id == selected_proj.id,
            ).order_by(models.OpeningPhoto.uploaded_at.desc()).all()
    return templates.TemplateResponse("opening.html", {
        "request": request, "user": user,
        "projects": projects,
        "selected_proj": selected_proj,
        "photos": photos,
        "msg": request.query_params.get("msg"),
    })


@router.post("/api/vpk/opening/upload-one")
async def opening_upload_one(request: Request, db: Session = Depends(get_db),
                              user: dict = Depends(require_api_user)):
    """Загрузка одного фото открытия (AJAX, один за раз). /api/ — CSRF exempt."""
    form = await request.form()
    project_id = form.get("project_id")
    if not project_id:
        return {"error": "project_id required"}

    proj = db.query(models.Project).filter(models.Project.id == int(project_id)).first()
    if not proj:
        return {"error": "ТК не найден"}

    photo_file = form.get("photo")
    if not photo_file or not hasattr(photo_file, "filename") or not photo_file.filename:
        return {"error": "Файл не получен"}

    raw = await photo_file.read()
    if not raw:
        return {"error": "Пустой файл"}

    try:
        check_magic_bytes(raw, photo_file.filename)
    except ValueError as e:
        return {"error": str(e)}

    ext = Path(photo_file.filename).suffix.lower() or ".jpg"
    ts  = int(datetime.utcnow().timestamp() * 1000)
    fname = f"open_{proj.id}_{ts}{ext}"
    photo_path = await asyncio.to_thread(upload_photo, raw, "opening", fname)

    photo = models.OpeningPhoto(
        project_id=proj.id,
        photo_path=photo_path,
        uploaded_by=user.get("display_name", ""),
        is_featured=False,
    )
    db.add(photo)
    db.commit()
    db.refresh(photo)

    from services.cloud_storage import media_url as _mu
    _vpk_logger.info("Opening: загружено фото %s для ТК %s", photo.id, proj.tk_number)
    return {"ok": True, "photo_id": photo.id, "url": _mu(photo_path)}


@router.post("/api/vpk/opening/{photo_id}/feature")
async def opening_toggle_feature(photo_id: int, request: Request,
                                  db: Session = Depends(get_db),
                                  user: dict = Depends(require_api_user)):
    """Toggle is_featured для фото (AJAX)."""
    photo = db.query(models.OpeningPhoto).filter(models.OpeningPhoto.id == photo_id).first()
    if not photo:
        raise HTTPException(status_code=404)
    photo.is_featured = not photo.is_featured
    db.commit()
    return {"ok": True, "is_featured": photo.is_featured}


@router.post("/api/vpk/opening/{photo_id}/delete")
async def opening_delete_photo(photo_id: int, request: Request,
                                db: Session = Depends(get_db),
                                user: dict = Depends(require_api_user)):
    """Удалить фото."""
    photo = db.query(models.OpeningPhoto).filter(models.OpeningPhoto.id == photo_id).first()
    if not photo:
        raise HTTPException(status_code=404)
    db.delete(photo)
    db.commit()
    return {"ok": True}


@router.post("/opening/send-report")
async def opening_send_report(request: Request, background_tasks: BackgroundTasks,
                               db: Session = Depends(get_db),
                               user: dict = Depends(require_login)):
    """Отправить email-отчёт: ВСЕ фото проекта (starred отмечены как лучшие в письме)."""
    form = await request.form()
    project_id = form.get("project_id")
    if not project_id:
        return RedirectResponse(f"/opening?msg={quote('Выберите ТК')}", status_code=303)

    proj = db.query(models.Project).filter(models.Project.id == int(project_id)).first()
    if not proj:
        return RedirectResponse(f"/opening?msg={quote('ТК не найден')}", status_code=303)

    # Сначала starred, потом остальные — чтобы лучшие шли первыми в письме
    all_photos = db.query(models.OpeningPhoto).filter(
        models.OpeningPhoto.project_id == proj.id,
    ).order_by(models.OpeningPhoto.is_featured.desc(), models.OpeningPhoto.uploaded_at).all()

    if not all_photos:
        return RedirectResponse(
            f"/opening?project_id={project_id}&msg={quote('Сначала загрузите фото')}",
            status_code=303,
        )

    photo_urls = [p.photo_path for p in all_photos]
    city = proj.city or ""
    submitter = user.get("display_name", "")

    open_emails: set[str] = set()
    if _OPENING_EMAIL:
        open_emails.add(_OPENING_EMAIL)
    sub_mgr_op = db.query(models.Manager).filter(
        models.Manager.name == submitter,
        models.Manager.email.isnot(None),
        models.Manager.email != "",
    ).first()
    if sub_mgr_op:
        open_emails.add(sub_mgr_op.email)

    if open_emails:
        for em in open_emails:
            background_tasks.add_task(
                notify_opening_photos,
                em, proj.tk_number, city, submitter, photo_urls, proj.id,
            )
        _vpk_logger.info("Opening report: отправка %d фото → %s", len(photo_urls), open_emails)
        msg = f"Отчёт отправлен — {len(photo_urls)} фото"
    else:
        msg = "NOTIFY_PRECHECK_EMAIL не задан"

    return RedirectResponse(
        f"/opening?project_id={project_id}&msg={quote(msg)}",
        status_code=303,
    )


@router.get("/api/vpk/opening/photos")
async def opening_photos_api(request: Request, project_id: int = 0,
                              db: Session = Depends(get_db)):
    """JSON список фото для проекта (для обновления галереи без перезагрузки)."""
    user = get_current_user(request)
    if not user:
        return {"photos": []}
    photos = db.query(models.OpeningPhoto).filter(
        models.OpeningPhoto.project_id == project_id,
    ).order_by(models.OpeningPhoto.uploaded_at.desc()).all()
    return {"photos": [
        {"id": p.id, "url": p.photo_path, "featured": p.is_featured,
         "by": p.uploaded_by, "at": p.uploaded_at.strftime("%d.%m %H:%M") if p.uploaded_at else ""}
        for p in photos
    ]}


@router.get("/opening/{project_id}", response_class=HTMLResponse)
async def opening_gallery_page(project_id: int, request: Request,
                                db: Session = Depends(get_db)):
    """Публичная галерея всех фото открытия (доступна по ссылке без входа)."""
    proj = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not proj:
        raise HTTPException(status_code=404)
    all_photos = db.query(models.OpeningPhoto).filter(
        models.OpeningPhoto.project_id == project_id,
    ).order_by(models.OpeningPhoto.is_featured.desc(), models.OpeningPhoto.uploaded_at).all()
    featured = [p for p in all_photos if p.is_featured]
    rest = [p for p in all_photos if not p.is_featured]
    user = get_current_user(request)
    return templates.TemplateResponse("opening_gallery.html", {
        "request": request,
        "user": user,
        "project": proj,
        "featured": featured,
        "rest": rest,
        "all_photos": all_photos,
    })


@router.get("/opening/{project_id}/download-zip")
async def opening_download_zip(project_id: int, db: Session = Depends(get_db)):
    """Скачать все фото открытия одним ZIP-архивом. Публичный доступ (как галерея)."""
    import io
    import zipfile
    from concurrent.futures import ThreadPoolExecutor
    from urllib.parse import quote

    import httpx as _httpx

    proj = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not proj:
        raise HTTPException(status_code=404)

    photos = db.query(models.OpeningPhoto).filter(
        models.OpeningPhoto.project_id == project_id,
    ).order_by(models.OpeningPhoto.is_featured.desc(), models.OpeningPhoto.uploaded_at).all()

    if not photos:
        raise HTTPException(status_code=404, detail="Нет фото")

    def _fetch_one(args):
        idx, photo = args
        try:
            path = photo.photo_path or ""
            if path.startswith("http"):
                r = _httpx.get(path, timeout=20, follow_redirects=True)
                data = r.content if r.status_code == 200 else None
                ext = ".jpg"
            else:
                full = Path("static") / path
                data = full.read_bytes() if full.exists() else None
                ext = full.suffix or ".jpg"
            if not data:
                return None
            prefix = "лучшее_" if photo.is_featured else "фото_"
            return f"{prefix}{idx:03d}{ext}", data
        except Exception as exc:
            _vpk_logger.warning("ZIP: ошибка загрузки фото %s: %s", photo.id, exc)
            return None

    def _build_zip() -> bytes:
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_STORED) as zf:
            with ThreadPoolExecutor(max_workers=10) as pool:
                results = pool.map(_fetch_one, enumerate(photos, 1))
            for item in results:
                if item:
                    name, data = item
                    zf.writestr(name, data)
        buf.seek(0)
        return buf.read()

    zip_data = await asyncio.to_thread(_build_zip)
    city = f"_{proj.city}" if proj.city else ""
    fname = f"Открытие_ТК{proj.tk_number}{city}.zip"

    from fastapi.responses import Response as _Resp
    return _Resp(
        content=zip_data,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"},
    )
