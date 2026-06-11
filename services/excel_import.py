"""Excel import functions for Реконструкции and Констракшн."""
import io
from datetime import datetime, date

import openpyxl
from sqlalchemy.orm import Session

import models
from config import STAGE_NAMES
from utils.excel import safe_date, row_to_dict, match_manager


def _detect_cols(ws, hdr_rows, defaults):
    """
    Определяет номера колонок по заголовкам листа.
    Возвращает defaults если колонка не найдена.
    hdr_rows — строки с заголовками (1-based), например [2, 3].
    """
    def find_exact(keywords, min_c=1):
        kws = [k.lower() for k in keywords]
        for ri in hdr_rows:
            for ci in range(min_c, ws.max_column + 1):
                v = str(ws.cell(ri, ci).value or "").strip().lower()
                if v in kws:
                    return ci
        return None

    def find_contains(keywords, min_c=1):
        kws = [k.lower() for k in keywords]
        for ri in hdr_rows:
            for ci in range(min_c, ws.max_column + 1):
                v = str(ws.cell(ri, ci).value or "").strip().lower()
                if any(kw in v for kw in kws):
                    return ci
        return None

    def start_end_after(from_c):
        s = e = None
        for ri in hdr_rows:
            for ci in range(from_c, min(from_c + 15, ws.max_column + 1)):
                v = str(ws.cell(ri, ci).value or "").strip().lower()
                if v == "старт" and s is None:
                    s = ci
                elif v == "окончание" and s is not None and e is None:
                    e = ci
                    return s, e
        return s, e

    r = dict(defaults)

    g = find_contains(["сбор исходных данных"])
    if g:
        s, e = start_end_after(g)
        if s: r["sid_s"] = s
        if e: r["sid_e"] = e

    g = find_contains(["зонирование"])
    if g:
        s, e = start_end_after(g)
        if s: r["zon_s"] = s
        if e: r["zon_e"] = e

    c = find_contains(["старт закрытия", "старт, закрытие"], min_c=30)
    if c: r["clos"] = c

    c = find_exact(["впк 1"], min_c=25) or find_contains(["впк 1"], min_c=25)
    if c: r["vpk"] = c

    c = find_exact(["открытие"], min_c=30)
    if c: r["opening"] = c

    c = find_exact(["мп"], min_c=35)
    if c: r["mgr"] = c

    c = find_exact(["мпс"], min_c=35)
    if c: r["mgr2"] = c

    return r


# Маппинг stage_key → ключ колонки конца в SHEETS (для чтения цвета ячейки)
_STAGE_COL_KEY = [
    ("sid",     "sid_e"),
    ("zoning",  "zon_e"),
    ("mp",      "mp_e"),
    ("tp",      "tp_e"),
    ("viz",     "viz_e"),
    ("audit",   "audit_e"),
    ("pjf",     "pjf_e"),
    ("ds",      "ds"),
    ("tz",      "tz_e"),
    ("closure", "clos"),
    ("vpk",     "vpk"),
    ("opening", "opening"),
]

# Зелёные RGB-цвета Excel → этап выполнен
_GREEN_RGB = {
    "92D050", "00B050", "70AD47", "339966", "00CC00",
    "C6EFCE", "4CAF50", "2ECC71", "27AE60", "1ABC9C",
    "00FF00", "228B22", "008000", "006400", "33CC33",
    "CCFFCC", "99FF99", "66CC66", "038303", "00B300",
}


def _cell_excel_color(ws, row: int, col: int) -> str:
    """Возвращает 'done' если ячейка залита зелёным, иначе ''."""
    if not col:
        return ""
    try:
        cell = ws.cell(row=row, column=col)
        fill = cell.fill
        if not fill or fill.fill_type != "solid":
            return ""
        fg = fill.fgColor
        if fg.type == "rgb":
            rgb = fg.rgb[-6:].upper()
            if rgb in _GREEN_RGB:
                return "done"
    except Exception:
        pass
    return ""


def import_reconstruct_excel(content: bytes, db: Session) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    managers = db.query(models.Manager).all()
    today = date.today()
    created = updated = 0
    # Собираем (project_obj, stage_key, is_done) для синхронизации после flush
    _stage_sync: list = []

    # Статичные параметры листов + дефолтные позиции колонок (авто-определяются из заголовков)
    SHEETS = [
        {"idx": 9,  "label": "Реконструкции 2026",
         "hdr": [2, 3], "start": 4, "tk": 1, "city": 4, "addr": 5, "area": 12,
         "sid_s": 14, "sid_e": 15, "zon_s": 16, "zon_e": 17,
         "mp_s": 18,  "mp_e": 19,  "tp_s": 20,  "tp_e": 21,
         "viz_s": 23, "viz_e": 24, "audit_s": 25, "audit_e": 26,
         "pjf_s": 27, "pjf_e": 28, "ds": 29,
         "tz_s": 30,  "tz_e": 31,
         "clos": 36, "vpk": 40, "opening": 41, "pjf_code": 45,
         "mgr": 52, "mgr2": 51, "status_col": 53},
        {"idx": 10, "label": "Лайт реконструкции 2026",
         "hdr": [2, 3], "start": 4, "tk": 1, "city": 4, "addr": 5, "area": 12,
         "sid_s": 13, "sid_e": 14, "zon_s": 15, "zon_e": 16,
         "mp_s": 17,  "mp_e": 18,  "tp_s": 19,  "tp_e": 20,
         "viz_s": 22, "viz_e": 23, "audit_s": 24, "audit_e": 25,
         "pjf_s": 26, "pjf_e": 27, "ds": 28,
         "tz_s": 29,  "tz_e": 30,
         "clos": 35, "vpk": 36, "opening": 37, "pjf_code": 38,
         "mgr": 42, "mgr2": 41, "status_col": 43},
        {"idx": 11, "label": "Рисковые объекты 2026 АЛ",
         "hdr": [1, 2], "start": 3, "tk": 1, "city": 2, "addr": 3, "area": 7,
         "sid_s": 9,  "sid_e": 10, "zon_s": 11, "zon_e": 12,
         "mp_s": 13,  "mp_e": 14,  "tp_s": 15,  "tp_e": 16,
         "viz_s": 18, "viz_e": 19, "audit_s": 20, "audit_e": 21,
         "pjf_s": 22, "pjf_e": 23, "ds": 24,
         "tz_s": 25,  "tz_e": 26,
         "clos": 32, "vpk": 33, "opening": 34, "pjf_code": 35,
         "mgr": 41, "mgr2": 40, "status_col": 42},
        {"idx": 12, "label": "Рисковые объекты Малая площадь",
         "hdr": [1, 2], "start": 3, "tk": 1, "city": 2, "addr": 3, "area": 7,
         "sid_s": 9,  "sid_e": 10, "zon_s": 11, "zon_e": 12,
         "mp_s": 13,  "mp_e": 14,  "tp_s": 15,  "tp_e": 16,
         "viz_s": 18, "viz_e": 19, "audit_s": 20, "audit_e": 21,
         "pjf_s": 22, "pjf_e": 23, "ds": 24,
         "tz_s": 25,  "tz_e": 26,
         "clos": 32, "vpk": 33, "opening": 34, "pjf_code": 0,
         "mgr": 37, "mgr2": 36, "status_col": 38},
    ]

    for sheet_cfg in SHEETS:
        idx = sheet_cfg["idx"]
        if idx >= len(wb.worksheets):
            continue
        ws = wb.worksheets[idx]

        # Авто-определение колонок по заголовкам
        cols = _detect_cols(ws, sheet_cfg["hdr"], sheet_cfg)

        for row_idx in range(sheet_cfg["start"], ws.max_row + 1):
            tk_val = ws.cell(row_idx, sheet_cfg["tk"]).value
            if not tk_val or not isinstance(tk_val, (int, float)):
                continue
            tk_num = str(int(tk_val))

            city_val = ws.cell(row_idx, sheet_cfg["city"]).value
            city = str(city_val).strip() if city_val and str(city_val).strip() not in ("", "#REF!", "#N/A") else ""

            addr_val = ws.cell(row_idx, sheet_cfg["addr"]).value
            address = str(addr_val).strip() if addr_val else ""

            area_val = ws.cell(row_idx, sheet_cfg["area"]).value
            area = float(area_val) if isinstance(area_val, (int, float)) else None

            sid_s   = safe_date(ws.cell(row_idx, cols["sid_s"]).value)
            sid_e   = safe_date(ws.cell(row_idx, cols["sid_e"]).value)
            zon_s   = safe_date(ws.cell(row_idx, cols["zon_s"]).value)
            zon_e   = safe_date(ws.cell(row_idx, cols["zon_e"]).value)
            mp_s    = safe_date(ws.cell(row_idx, cols["mp_s"]).value)
            mp_e    = safe_date(ws.cell(row_idx, cols["mp_e"]).value)
            tp_s    = safe_date(ws.cell(row_idx, cols["tp_s"]).value)
            tp_e    = safe_date(ws.cell(row_idx, cols["tp_e"]).value)
            viz_s   = safe_date(ws.cell(row_idx, cols["viz_s"]).value)
            viz_e   = safe_date(ws.cell(row_idx, cols["viz_e"]).value)
            aud_s   = safe_date(ws.cell(row_idx, cols["audit_s"]).value)
            aud_e   = safe_date(ws.cell(row_idx, cols["audit_e"]).value)
            pjf_s   = safe_date(ws.cell(row_idx, cols["pjf_s"]).value)
            pjf_e   = safe_date(ws.cell(row_idx, cols["pjf_e"]).value)
            ds      = safe_date(ws.cell(row_idx, cols["ds"]).value)
            tz_s    = safe_date(ws.cell(row_idx, cols["tz_s"]).value)
            tz_e    = safe_date(ws.cell(row_idx, cols["tz_e"]).value)
            clos    = safe_date(ws.cell(row_idx, cols["clos"]).value)
            vpk     = safe_date(ws.cell(row_idx, cols["vpk"]).value)
            opening = safe_date(ws.cell(row_idx, cols["opening"]).value)

            pjf_code_val = ""
            if cols.get("pjf_code"):
                raw = ws.cell(row_idx, cols["pjf_code"]).value
                pjf_code_val = str(raw).strip() if raw else ""

            # Последний комментарий — первая непустая текстовая ячейка в статус-колонках
            status_comment = ""
            sc = cols.get("status_col", 0)
            if sc:
                for c in range(sc, min(sc + 60, ws.max_column + 1)):
                    v = ws.cell(row_idx, c).value
                    if v and not hasattr(v, "year"):
                        txt = str(v).strip()
                        if txt:
                            status_comment = txt
                            break

            mgr_raw  = str(ws.cell(row_idx, cols["mgr"]).value  or "").strip()
            mgr_raw2 = str(ws.cell(row_idx, cols["mgr2"]).value or "").strip()
            manager_id = match_manager(mgr_raw, managers) or match_manager(mgr_raw2, managers)

            status = "Завершён" if (opening and opening <= today) else "Активный"
            proj_name = f"ТК {tk_num}" + (f" {city}" if city else "")
            sheet_label = sheet_cfg.get("label", "Реконструкция")

            existing = db.query(models.Project).filter(
                models.Project.tk_number == tk_num,
                models.Project.project_type == "Реконструкция",
            ).first()

            stage_fields = dict(
                sid_start=sid_s, sid_end=sid_e,
                zoning_start=zon_s, zoning_end=zon_e,
                mp_start=mp_s, mp_end=mp_e,
                tp_start=tp_s, tp_end=tp_e,
                visualization_start=viz_s, visualization_end=viz_e,
                audit_start=aud_s, audit_end=aud_e,
                pjf_approval_start=pjf_s, pjf_approval_end=pjf_e,
                ds_signing_date=ds,
                tz_start=tz_s, tz_end=tz_e,
                closure_date=clos, vpk_date=vpk, opening_date=opening,
            )

            if existing:
                if city:    existing.city    = city
                if address: existing.address = address
                if area:    existing.area    = area
                existing.project_type  = "Реконструкция"
                existing.format_type   = sheet_label
                existing.start_date    = sid_s
                existing.end_date      = opening
                existing.status_comment = status_comment
                if pjf_code_val:
                    existing.pjf_code = pjf_code_val
                for field, val in stage_fields.items():
                    setattr(existing, field, val)
                if existing.status != "Приостановлен":
                    existing.status = status
                if manager_id:
                    existing.manager_id = manager_id
                proj_obj = existing
                updated += 1
            else:
                proj_obj = models.Project(
                    name=proj_name, tk_number=tk_num,
                    city=city, address=address,
                    project_type="Реконструкция",
                    format_type=sheet_label,
                    manager_id=manager_id, status=status,
                    start_date=sid_s, end_date=opening, area=area,
                    pjf_code=pjf_code_val,
                    status_comment=status_comment,
                    **stage_fields,
                )
                db.add(proj_obj)
                created += 1

            # Собираем цвета ячеек для синхронизации статусов этапов
            for stage_key, col_key in _STAGE_COL_KEY:
                col = cols.get(col_key, 0)
                color = _cell_excel_color(ws, row_idx, col) if col else ""
                _stage_sync.append((proj_obj, stage_key, color == "done"))

        db.flush()

    # После flush все project_id назначены — синхронизируем статусы из Excel
    now = datetime.utcnow()
    for proj_obj, stage_key, is_done in _stage_sync:
        if not proj_obj.id:
            continue
        rec = db.query(models.ReconStageStatus).filter_by(
            project_id=proj_obj.id,
            stage_key=stage_key,
        ).first()
        if rec:
            rec.is_done = is_done
            if is_done:
                rec.done_by = rec.done_by or "Excel"
                rec.done_at = rec.done_at or now
            else:
                rec.done_by = ""
                rec.done_at = None
        elif is_done:
            db.add(models.ReconStageStatus(
                project_id=proj_obj.id,
                stage_key=stage_key,
                is_done=True,
                done_by="Excel",
                done_at=now,
            ))

    try:
        db.commit()
    except Exception:
        db.rollback()
        raise
    return {"created": created, "updated": updated}


def import_construction_excel(content: bytes, db: Session) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    managers = db.query(models.Manager).all()
    today = date.today()
    created = updated = 0

    target_sheets = [ws for ws in wb.worksheets
                     if ws.title.strip() in ('2026', '2027')
                     or '2026' in ws.title or '2027' in ws.title]
    seen = set()
    target_sheets = [ws for ws in target_sheets
                     if ws.title not in seen and not seen.add(ws.title)]
    if not target_sheets:
        target_sheets = [wb.worksheets[-1]]

    skipped_fmt = skipped_mgr = rows_with_tk = 0
    sample_formats: set = set()
    sample_managers: set = set()

    for ws in target_sheets:
        header_row = 4
        found_hdr = False
        for r in range(1, 11):
            if found_hdr:
                break
            for c in range(1, 15):
                v = str(ws.cell(r, c).value or '')
                if 'Номер' in v or 'Адрес' in v:
                    header_row = r
                    found_hdr = True
                    break

        col: dict = {}
        for r in range(max(1, header_row - 1), header_row + 2):
            for c in range(1, 60):
                v = str(ws.cell(r, c).value or '').strip()
                vl = v.lower()
                if not v:
                    continue
                if ('номер тк' in vl or vl == 'тк') and 'tk' not in col:
                    col['tk'] = c
                elif 'адрес' in vl and 'addr' not in col and len(v) < 30:
                    col['addr'] = c
                elif 'тип формата' in vl and 'fmt' not in col:
                    col['fmt'] = c
                elif ('формат' in vl and len(v) < 20) and 'fmt' not in col:
                    col['fmt'] = c
                elif ('регион' in vl or 'город' in vl) and 'city' not in col:
                    col['city'] = c
                elif 'площадь' in vl and 'тз' in vl and 'area' not in col:
                    col['area'] = c
                elif ('приёмка' in vl or 'приемка' in vl) and 'reception' not in col:
                    col['reception'] = c
                elif 'выход на смр' in vl and 'cmp' not in col:
                    col['cmp'] = c
                elif vl.startswith('впк') and 'vpk' not in col:
                    col['vpk'] = c
                elif 'первоначальная' in vl and 'open_plan' not in col:
                    col['open_plan'] = c
                elif 'фактическая' in vl and 'open_fact' not in col:
                    col['open_fact'] = c
                elif 'статус' in vl and ('откр' in vl or 'open' in vl) and 'open_status' not in col:
                    col['open_status'] = c
                elif 'менеджер' in vl and 'ос' in vl and 'mgr_os' not in col:
                    col['mgr_os'] = c

        fmt_from_header = 'fmt' in col
        col.setdefault('tk', 2)
        col.setdefault('addr', 3)
        col.setdefault('reception', 14)
        col.setdefault('cmp', 16)
        col.setdefault('vpk', 19)
        col.setdefault('open_plan', 20)
        col.setdefault('open_fact', 21)
        col.setdefault('open_status', 22)
        col.setdefault('mgr_os', 38)

        for row_idx in range(header_row + 1, ws.max_row + 1):
            tk_val = ws.cell(row_idx, col['tk']).value
            if not tk_val:
                continue
            tk_num = str(tk_val).strip()
            if not tk_num or tk_num in ('None', 'Номер ТК', 'ТК') or len(tk_num) > 20:
                continue

            rows_with_tk += 1
            address  = str(ws.cell(row_idx, col['addr']).value or '').strip()
            fmt_type = str(ws.cell(row_idx, col['fmt']).value or '').strip() if col.get('fmt') else ''
            if fmt_type and len(sample_formats) < 8:
                sample_formats.add(fmt_type)
            mgr_raw = str(ws.cell(row_idx, col['mgr_os']).value or '').strip()
            if mgr_raw and len(sample_managers) < 8:
                sample_managers.add(mgr_raw)
            city = str(ws.cell(row_idx, col['city']).value or '').strip() if col.get('city') else ''
            area_val = ws.cell(row_idx, col['area']).value if col.get('area') else None
            area = float(area_val) if isinstance(area_val, (int, float)) else None

            reception = safe_date(ws.cell(row_idx, col['reception']).value)
            cmp_date  = safe_date(ws.cell(row_idx, col['cmp']).value)
            vpk_date  = safe_date(ws.cell(row_idx, col['vpk']).value)
            open_plan = safe_date(ws.cell(row_idx, col['open_plan']).value)
            open_fact = safe_date(ws.cell(row_idx, col['open_fact']).value)
            open_st   = str(ws.cell(row_idx, col['open_status']).value or '').strip()

            manager_id = match_manager(mgr_raw, managers)

            fmt_upper = fmt_type.upper()
            if fmt_from_header and fmt_type and fmt_upper not in ('SM', 'UTKONOS'):
                skipped_fmt += 1
                continue
            if not manager_id:
                skipped_mgr += 1
                continue

            opening = open_fact or open_plan
            status  = "Завершён" if (opening and opening <= today) else "Активный"
            proj_name = f"ТК {tk_num}" + (f" {city}" if city else "")

            existing = db.query(models.Project).filter(
                models.Project.tk_number == tk_num,
                models.Project.project_type == "Констракшн"
            ).first()
            if existing:
                if address:  existing.address     = address
                if city:     existing.city        = city
                if area:     existing.area        = area
                if fmt_type: existing.format_type = fmt_type
                if open_st:  existing.open_status = open_st
                existing.start_date   = reception or cmp_date
                existing.closure_date = cmp_date
                existing.vpk_date     = vpk_date
                existing.end_date     = open_plan or open_fact
                existing.opening_date = open_fact
                if existing.status != "Приостановлен":
                    existing.status = status
                existing.manager_id = manager_id
                updated += 1
            else:
                db.add(models.Project(
                    name=proj_name, tk_number=tk_num,
                    city=city, address=address,
                    project_type="Констракшн",
                    manager_id=manager_id, status=status,
                    format_type=fmt_type, open_status=open_st,
                    start_date=reception or cmp_date, closure_date=cmp_date,
                    vpk_date=vpk_date, end_date=open_plan or open_fact,
                    opening_date=open_fact, area=area,
                ))
                created += 1

    try:
        db.commit()
    except Exception:
        db.rollback()
        raise
    return {
        "created": created, "updated": updated,
        "sheets": ", ".join(ws.title for ws in target_sheets),
        "rows_with_tk": rows_with_tk,
        "skipped_fmt": skipped_fmt, "skipped_mgr": skipped_mgr,
        "sample_formats": sorted(sample_formats),
        "sample_managers": sorted(sample_managers),
    }


def parse_excel_file(content: bytes, project_type: str,
                     manager_id_default: int | None, db: Session) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    created = updated = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        project_name = sheet_name
        tk_num = city_name = ""

        for row in ws.iter_rows(max_row=10, values_only=True):
            for cell in row:
                if cell and isinstance(cell, str):
                    text = cell.strip()
                    if text.upper().startswith(("ТК", "TK")):
                        project_name = text
                        parts = text.split()
                        if len(parts) >= 2:
                            tk_num = parts[1]
                        if len(parts) >= 3:
                            city_name = " ".join(parts[2:])

        all_dates = []
        for row in ws.iter_rows(max_row=100, values_only=True):
            for cell in row:
                d = safe_date(cell)
                if d:
                    all_dates.append(d)

        proj_start = min(all_dates) if all_dates else None
        proj_end   = max(all_dates) if all_dates else None

        existing = None
        if tk_num:
            existing = db.query(models.Project).filter(
                models.Project.tk_number == tk_num).first()
        if not existing and project_name:
            existing = db.query(models.Project).filter(
                models.Project.name == project_name).first()

        if existing:
            if proj_start:
                existing.start_date = proj_start
            if proj_end:
                existing.end_date = proj_end
            if city_name and not existing.city:
                existing.city = city_name
            if project_type and not existing.project_type:
                existing.project_type = project_type
            db.flush()
            _sync_stages(ws, existing.id, db)
            updated += 1
        else:
            p = models.Project(
                name=project_name, tk_number=tk_num, city=city_name,
                project_type=project_type,
                manager_id=manager_id_default,
                status="Активный", start_date=proj_start, end_date=proj_end,
            )
            db.add(p)
            db.flush()
            _sync_stages(ws, p.id, db)
            created += 1

    try:
        db.commit()
    except Exception:
        db.rollback()
        raise
    return {"created": created, "updated": updated}


def _sync_stages(ws, project_id: int, db: Session):
    existing_names = {s.name for s in db.query(models.ProjectStage).filter(
        models.ProjectStage.project_id == project_id).all()}
    order = db.query(models.ProjectStage).filter(
        models.ProjectStage.project_id == project_id).count()

    for row in ws.iter_rows(values_only=True):
        if not row[0] or not isinstance(row[0], str):
            continue
        cell_text = row[0].strip()
        for sn in STAGE_NAMES:
            if sn.split()[0].lower() in cell_text.lower():
                start = end = None
                for v in row[1:]:
                    d = safe_date(v)
                    if d and not start:
                        start = d
                    elif d and not end:
                        end = d
                        break
                if cell_text not in existing_names:
                    db.add(models.ProjectStage(
                        project_id=project_id, name=cell_text,
                        start_date=start, end_date=end, order=order))
                    order += 1
                else:
                    s = db.query(models.ProjectStage).filter(
                        models.ProjectStage.project_id == project_id,
                        models.ProjectStage.name == cell_text).first()
                    if s:
                        if start:
                            s.start_date = start
                        if end:
                            s.end_date = end
                break
