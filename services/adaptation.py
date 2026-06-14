"""Excel generation for Adaptation Card using openpyxl."""
import io
from pathlib import Path

import openpyxl

TEMPLATE_PATH = Path(__file__).parent.parent / "static" / "adaptation_template.xlsx"
SHEET_NAME = "Чек-лист"

# ── Dropdown options ──────────────────────────────────────────────────────────
OPTS_YES_NO_OTHER   = ["да", "нет", "иное"]
OPTS_YES_NO         = ["да", "нет"]
OPTS_RECON_TYPE     = ["Реконструкция", "Уплотнение", "Изменение МП", "Сокращение площадей", "Реконструкция с сокращением площадей"]
OPTS_OWNERSHIP      = ["аренда", "собственность"]
OPTS_HO_TYPE        = ["Встройка", "Вынос", "Встройка+Вынос"]
OPTS_BUILDING_LOC   = ["отдельностоящее", "внутри ТЦ", "пристроенное к жилому дому"]
OPTS_CEILING        = ["Подвесной армстронг", "Грильято", "Подвесной потолок отсутствует", "Частично подвесной, частично отсутствует", "Иное"]
OPTS_BUILDING_TYPE  = ["МКД", "ТЦ", "жилой дом", "пристроенное к жилому дому"]
OPTS_REPLACE        = ["да", "нет", "иное", "полная замена", "частичная замена"]
OPTS_REPLACE_REPAIR = ["да", "нет", "иное", "полная замена", "частичная замена", "полный ремонт"]
OPTS_REPAIR_FULL    = ["да", "нет", "иное", "полная замена", "частичная замена", "полный ремонт", "локальный ремонт"]
OPTS_WASH_ROBOT     = ["Робот", "Поломоечная машина", "нет"]

# ── Form field definitions ────────────────────────────────────────────────────
# key → (cell, label, options or None for free text)
FIELDS: list[tuple[str, str, str, list | None]] = [
    # Section 1: Общая информация
    ("tk_number",        "B5",  "Номер ТК",                                    None),
    ("recon_type",       "B6",  "Суть проекта",                                OPTS_RECON_TYPE),
    ("current_rto",      "B7",  "Текущий РТО (дневной)",                       None),
    ("projected_rto",    "B8",  "Прогнозируемый РТО (дневной)",                None),
    ("zoning_dev",       "B9",  "Разработка зонирования",                      OPTS_YES_NO_OTHER),
    ("zoning_wishes",    "B10", "Пожелания по зонированию",                    OPTS_YES_NO_OTHER),
    ("mp_revision",      "B11", "Пересмотр МП",                                OPTS_YES_NO_OTHER),
    ("real_estate",      "B12", "Недвижимость",                                OPTS_OWNERSHIP),
    ("open_facade",      "B13", "Открытый фасад",                              OPTS_YES_NO_OTHER),
    ("second_entrance",  "B14", "Второй вход",                                 OPTS_YES_NO_OTHER),
    ("ho_type",          "B15", "Тип ХО",                                      OPTS_HO_TYPE),
    ("sign_change",      "B16", "Замена вывески",                              OPTS_YES_NO_OTHER),
    ("led_facade",       "B17", "Светодиодная лента на фасаде (неон)",         OPTS_YES_NO_OTHER),
    ("light_structures", "B18", "Световые конструкции",                        OPTS_REPLACE_REPAIR),
    ("stela",            "B19", "Стелла",                                      OPTS_REPLACE_REPAIR),
    ("building_location","B20", "Строение",                                    OPTS_BUILDING_LOC),
    ("ceiling_tz",       "B21", "Тип потолков ТЗ",                             OPTS_CEILING),
    ("ceiling_bek",      "B22", "Тип потолков БЭК",                            OPTS_CEILING),
    ("ceiling_height",   "B23", "Высота потолков",                             None),
    ("building_type",    "B24", "Тип здания",                                  OPTS_BUILDING_TYPE),
    ("ownership_type",   "B25", "Тип собственности",                           OPTS_OWNERSHIP),
    ("online_zone",      "B26", "Зона онлайн",                                 OPTS_YES_NO),
    ("wall_demolition",  "B27", "Демонтаж старого оформления стен/колонн",     OPTS_REPLACE_REPAIR),
    ("vitrage_prep",     "B28", "Подготовка витражей к новой оклейке",         OPTS_REPLACE_REPAIR),
    ("service_options",  "B29", "Сервисные опции в ТЗ",                        None),

    # Section 2: Производство / Кухня (column B = ГО, C = размер)
    ("bakery",           "B31", "Пекарня СП",                                  OPTS_YES_NO_OTHER),
    ("bakery_size",      "C31", "Пекарня — размер",                            None),
    ("tandyr",           "B32", "Тандыр",                                      OPTS_YES_NO_OTHER),
    ("tandyr_size",      "C32", "Тандыр — размер",                             None),
    ("grill",            "B33", "Гриль",                                       OPTS_YES_NO_OTHER),
    ("grill_size",       "C33", "Гриль — размер",                              None),
    ("fish_fresh",       "B34", "Рыба на льду",                                OPTS_YES_NO_OTHER),
    ("fish_fresh_size",  "C34", "Рыба на льду — размер",                       None),
    ("fish_smoked",      "B35", "Рыба копчёная",                               OPTS_YES_NO_OTHER),
    ("fish_smoked_size", "C35", "Рыба копчёная — размер",                      None),
    ("meat_chilled",     "B36", "Мясо охлаждённое",                            OPTS_YES_NO_OTHER),
    ("meat_chilled_size","C36", "Мясо охлаждённое — размер",                   None),
    ("meat_shop",        "B37", "Мясная лавка",                                OPTS_YES_NO_OTHER),
    ("meat_shop_size",   "C37", "Мясная лавка — размер",                       None),
    ("cheese",           "B38", "Сыр",                                         OPTS_YES_NO_OTHER),
    ("cheese_size",      "C38", "Сыр — размер",                                None),
    ("sausage",          "B39", "Колбаса",                                     OPTS_YES_NO_OTHER),
    ("sausage_size",     "C39", "Колбаса — размер",                            None),
    ("salads",           "B40", "Салаты / ГБ",                                 OPTS_YES_NO_OTHER),
    ("salads_size",      "C40", "Салаты — размер",                             None),
    ("pizza",            "B41", "Пицца",                                       OPTS_YES_NO_OTHER),
    ("pizza_size",       "C41", "Пицца — размер",                              None),
    ("shawarma",         "B42", "Шаверма",                                     OPTS_YES_NO_OTHER),
    ("shawarma_size",    "C42", "Шаверма — размер",                            None),
    ("pancakes",         "B43", "Блины",                                       OPTS_YES_NO_OTHER),
    ("pancakes_size",    "C43", "Блины — размер",                              None),
    ("coffee",           "B44", "Кофе-машина",                                 OPTS_YES_NO_OTHER),
    ("coffee_size",      "C44", "Кофе — размер",                               None),
    ("cafe_zone",        "B45", "Зона кафе",                                   OPTS_YES_NO_OTHER),
    ("cafe_zone_size",   "C45", "Зона кафе — размер",                          None),

    # Section 3: Оборудование ТЗ
    ("cash_tables",      "B52", "Кассовые столы",                              OPTS_REPLACE),
    ("cash_size",        "C52", "Кассовые столы — кол-во",                     None),
    ("kso",              "B53", "КСО",                                         OPTS_REPLACE),
    ("kso_size",         "C53", "КСО — кол-во",                                None),
    ("marmite",          "B54", "Мармиты",                                     OPTS_REPLACE),
    ("grill_eq",         "B55", "Гриль (оборудование)",                        OPTS_REPLACE),
    ("oven",             "B56", "Печь",                                        OPTS_REPLACE),
    ("convection_oven",  "B57", "Пароконвектомат",                             OPTS_REPLACE),
    ("cool_shelves",     "B58", "Холодильные горки",                           OPTS_REPLACE),
    ("cool_shelves_1500","B59", "Холодильные горки H 1500",                    OPTS_REPLACE),
    ("cool_doors",       "B60", "Наличие дверей в холодильных горках",         OPTS_REPLACE),
    ("ho_ikra",          "B61", "ХО Икра",                                     OPTS_REPLACE),
    ("lari",             "B62", "Лари",                                        OPTS_REPLACE),
    ("vizory",           "B63", "Визоры",                                      OPTS_REPLACE),
    ("showcases",        "B64", "Витрины",                                     OPTS_REPLACE),
    ("bonety",           "B65", "Бонеты",                                      OPTS_REPLACE),
    ("freezer_cabinets", "B66", "Морозильные шкафы",                           OPTS_REPLACE),
    ("eq_bumpers",       "B67", "Отбойники оборудования",                      OPTS_REPLACE),
    ("ldsp",             "B68", "ЛДСП",                                        OPTS_REPLACE),
    ("shelving",         "B69", "Стеллажное оборудование",                     OPTS_REPLACE),
    ("shelving_sof",     "B70", "Стеллажное оборудование СОФ",                 OPTS_REPLACE),
    ("shelving_bakery",  "B71", "Стеллажное оборудование Пекарня",             OPTS_REPLACE),
    ("shopping_carts",   "B72", "Тележки покупательские",                      OPTS_REPLACE),
    ("price_holders",    "B73", "Замена ценникодержателей",                    ["да"]),
    ("shelving_recount", "B74", "Пересчёт стеллажного оборудования (БУ)",      OPTS_YES_NO),

    # Section 4: Оборудование ИТ
    ("ad_monitors",      "B77", "Рекламные мониторы над кассами",              OPTS_REPLACE),
    ("weight_equip",     "B78", "Весовое оборудование",                        OPTS_REPLACE),
    ("price_checkers",   "B79", "Прайс-чекеры",                                OPTS_REPLACE),
    ("sks",              "B80", "СКС",                                         OPTS_REPLACE),
    ("announcement",     "B81", "Система коммерческого оповещения",            OPTS_REPLACE),
    ("server_room",      "B82", "Серверная",                                   OPTS_REPLACE),
    ("kkt",              "B83", "Замена ККТ",                                  OPTS_REPLACE),
    ("sticker_printer",  "B84", "Принтер печати стикеров",                     OPTS_REPLACE),
    ("cellular",         "B85", "Усиление сотовой связи",                      OPTS_REPLACE),
    ("wifi",             "B86", "Вай-фай",                                     OPTS_REPLACE),
    ("computers",        "B87", "Компьютеры",                                  OPTS_REPLACE),

    # Section 5: Оборудование Бэкофиса
    ("office_furniture", "B90", "Мебель офиса",                                OPTS_REPLACE),
    ("metal_furniture",  "B91", "Мебель металлическая",                        OPTS_REPLACE),
    ("tech_equipment",   "B92", "Технологическое оборудование",                OPTS_REPLACE),
    ("bek_shelving",     "B93", "Стеллажное оборудование БЭК",                 OPTS_REPLACE),
    ("cargo_shelving",   "B94", "Грузовые стеллажи в БЭК",                    OPTS_REPLACE),
    ("defect_fridge",    "B95", "Холодильник для брака",                       OPTS_REPLACE),
    ("cryo_coolers",     "B96", "Холодильные камеры (воздухоохладители)",      OPTS_REPLACE),
    ("cryo_panels",      "B97", "Холодильные камеры (панели)",                 OPTS_REPLACE),
    ("cryo_doors",       "B98", "Двери и фурнитура холодильных камер",         OPTS_REPLACE),
    ("cryo_shelving",    "B99", "Стеллажи в холодильных камерах",              OPTS_REPLACE),
    ("chm_replace",      "B100","ЦХМ замена",                                  OPTS_REPLACE),
    ("chm_relocate",     "B101","ЦХМ изменение месторасположения",             OPTS_REPLACE),
    ("condenser",        "B102","Конденсаторные блоки",                        OPTS_REPLACE),
    ("cargo_scales_15",  "B103","Грузовые весы 1,5 т.",                        OPTS_REPLACE),
    ("cargo_scales_30",  "B104","Грузовые весы до 30 кг",                      OPTS_REPLACE),
    ("press",            "B105","Пресс",                                       OPTS_REPLACE),
    ("hydraulic_carts",  "B106","Гидравлические тележки",                      OPTS_REPLACE),

    # Section 6: Помещения Бэкофиса
    ("room_receiving",   "B110","Приёмка",                                     OPTS_REPAIR_FULL),
    ("room_dining",      "B111","Комната приёма пищи",                         OPTS_REPAIR_FULL),
    ("room_office",      "B112","Офис",                                        OPTS_REPAIR_FULL),
    ("room_changing",    "B113","Раздевалки",                                  OPTS_REPAIR_FULL),
    ("room_wc",          "B114","Сан. Узлы",                                   OPTS_REPAIR_FULL),
    ("room_floor",       "B115","Напольное покрытие (бэк)",                    OPTS_REPAIR_FULL),
    ("room_cashroom",    "B116","Главная касса",                               OPTS_REPAIR_FULL),
    ("room_bank_equip",  "B117","Замена банковского оборудования",             OPTS_REPAIR_FULL),

    # Section 7: Помещение ТЗ
    ("tz_floor",         "B120","Напольное покрытие (ТЗ)",                     OPTS_REPAIR_FULL),
    ("tz_lighting",      "B121","Освещение",                                   OPTS_REPAIR_FULL),
    ("tz_accent_light",  "B122","Акцентное освещение",                         OPTS_REPAIR_FULL),
    ("tz_walls",         "B123","Look & Feel (покраска стен)",                 OPTS_REPAIR_FULL),
    ("tz_entrance_grp",  "B124","Входная группа",                              OPTS_REPAIR_FULL),
    ("tz_auto_doors",    "B125","Входные автоматические двери",                OPTS_REPAIR_FULL),
    ("tz_mech_doors",    "B126","Входные механические двери",                  OPTS_REPAIR_FULL),
    ("tz_porch",         "B127","Крыльцо",                                     OPTS_REPAIR_FULL),
    ("tz_parking",       "B128","Парковка",                                    OPTS_REPAIR_FULL),
    ("tz_vitrages",      "B129","Витражи (открытый фасад)",                    OPTS_REPAIR_FULL),
    ("tz_imposty",       "B130","Импосты открытого фасада",                    OPTS_REPAIR_FULL),
    ("tz_ceiling",       "B131","Потолок",                                     OPTS_REPAIR_FULL),
    ("tz_announcement",  "B132","Система оповещения",                          OPTS_REPAIR_FULL),
    ("tz_fire_alarm",    "B133","Пожарная сигнализация",                       OPTS_REPAIR_FULL),
    ("tz_smoke",         "B134","Система дымоудаления",                        OPTS_REPAIR_FULL),
    ("tz_aupt",          "B135","АУПТ",                                        OPTS_REPAIR_FULL),
    ("tz_radiators",     "B136","Радиаторное отопление",                       OPTS_REPAIR_FULL),
    ("tz_heating",       "B137","Система отопления",                           OPTS_REPAIR_FULL),
    ("tz_conditioners",  "B138","Кондиционеры",                                OPTS_REPAIR_FULL),
    ("tz_partitions",    "B139","Новые перегородки",                           OPTS_REPAIR_FULL),
    ("tz_air_curtains",  "B140","Тепловые завесы",                             OPTS_REPAIR_FULL),
    ("tz_lift_tables",   "B141","Подъёмные столы",                             OPTS_REPAIR_FULL),
    ("tz_lifts",         "B142","Лифты/подъёмники",                            OPTS_REPAIR_FULL),
    ("tz_sectional_gates","B143","Секционные ворота",                          OPTS_REPAIR_FULL),
    ("tz_roller_gates",  "B144","Рулонные ворота",                             OPTS_REPAIR_FULL),
    ("tz_windows",       "B145","Стеклопакеты",                                OPTS_REPAIR_FULL),
    ("tz_mgn",           "B146","Конструкции для МГН",                         OPTS_REPAIR_FULL),
    ("tz_fire_doors",    "B147","Противопожарные двери",                       OPTS_REPAIR_FULL),
    ("tz_grshe",         "B148","ГРЩ",                                         OPTS_REPAIR_FULL),
    ("tz_wash_robot",    "B149","Поломоечная машина/робот",                    OPTS_WASH_ROBOT),

    # Section 8: Маркетинг
    ("mkt_facade_photos","B151","Получены фото фасада и входной группы",       OPTS_YES_NO),
    ("mkt_trc_photos",   "B152","Фото входной зоны ТРЦ (если в ТРЦ)",         OPTS_YES_NO),
    ("mkt_trc_navigation","B153","Нужна доп. навигация внутри ТРЦ",           OPTS_YES_NO),
    ("mkt_pattern",      "B154","Паттерн в ТЗ, нужно обновление?",            OPTS_YES_NO),
    ("mkt_ashel_mpc",    "B155","Лёгкая ашель и МНЦ над промо бинами",        OPTS_YES_NO),
    ("mkt_ashel_sof",    "B156","Сохранить ашель над СОФ",                    OPTS_YES_NO),
    ("mkt_posm",         "B157","Заказ POSM материалов для оформления",        OPTS_YES_NO),

    # Section 9: Иная важная информация (строки 159-177) — фиксированные поля
    ("who_demolition",   "A159","Кто проводит демонтаж, особенности по стройке",  None),
    ("work_restrictions","B160","Ограничения по работам (жилой дом и т.д.)",       None),
    ("lenta_online",     "B161","Лента Онлайн",                                    OPTS_YES_NO_OTHER),
    ("cctv",             "B162","Система видеонаблюдения",                         OPTS_YES_NO_OTHER),
    ("security_alarm",   "B163","Охранная сигнализация",                           OPTS_YES_NO_OTHER),
    ("antitheft_frames", "B164","Антикражные рамки",                               OPTS_YES_NO_OTHER),
    ("post1",            "B165","Пост №1",                                         OPTS_YES_NO_OTHER),
    ("new_ho",           "B166","Новое ХО",                                        OPTS_YES_NO_OTHER),
    ("meat_workshop",    "B167","Мясной цех",                                      OPTS_YES_NO_OTHER),
    ("cryo_floors",      "B168","Полы в ХК",                                       OPTS_YES_NO_OTHER),
    ("mobile_signal",    "B169","Сотовая связь",                                   OPTS_YES_NO_OTHER),
    ("walkie_talkie",    "B170","Рации",                                            OPTS_YES_NO_OTHER),
    ("fire_safety_lenta","B171","Пожарная безопасность за ООО Лента",              OPTS_YES_NO_OTHER),
    ("fire_safety_ardd", "B172","Пожарная безопасность за АРДД",                   OPTS_YES_NO_OTHER),
    ("receiving_room",   "B173","Комната приёмки товара",                          OPTS_YES_NO_OTHER),
    ("power_supply",     "B174","Электроснабжение",                                OPTS_YES_NO_OTHER),
    ("heat_supply",      "B175","Теплоснабжение",                                  OPTS_YES_NO_OTHER),
    ("water_supply",     "B176","Водоснабжение",                                   OPTS_YES_NO_OTHER),
    ("subrent_utilities","A177","Доп. вводы для субаренды (ХВС/ГВС/водоотведение/электроэнергия)", None),
]

# Free-text note sections with yellow headers in Excel
# key → (header_row, first_data_row, last_template_row, display_label)
FREE_TEXT_SECTIONS: dict[str, tuple[int, int, int, str]] = {
    "notes_main":    (178, 179, 183, "Иная важная информация"),
    "notes_fasad":   (184, 185, 191, "Фасад"),
    "notes_pb":      (192, 193, 195, "Пожарная безопасность"),
    "notes_elektro": (196, 197, 198, "Электроснабжение"),
    "notes_voda":    (199, 200, 203, "Водоснабжение"),
    "notes_teplo":   (204, 205, 210, "Теплоснабжение"),
}

# Quick lookup: key → (cell, label, options)
FIELD_MAP = {f[0]: (f[1], f[2], f[3]) for f in FIELDS}

# Dropdown options keyed by field name (for the HTML form)
DROPDOWN_OPTIONS: dict[str, list[str]] = {
    key: opts for key, cell, label, opts in FIELDS if opts
}


def get_template_path() -> Path:
    return TEMPLATE_PATH


def generate_excel(data: dict) -> bytes:
    """Open the adaptation template, write form values into 'Чек-лист', return bytes.

    All sheets, images and formatting from the original template are preserved.
    Fixed fields are written via FIELD_MAP.
    Free-text sections (multiline textarea) are split by newline and written into
    consecutive rows below each yellow header.
    """
    wb = openpyxl.load_workbook(str(TEMPLATE_PATH), keep_vba=False)

    ws = None
    for sheet in wb.worksheets:
        if sheet.title == SHEET_NAME:
            ws = sheet
            break
    if ws is None:
        ws = wb.worksheets[0]

    # Write fixed fields
    for key, value in data.items():
        entry = FIELD_MAP.get(key)
        if not entry or value is None or value == "":
            continue
        ws[entry[0]] = value

    # Write free-text sections (split by newline → individual rows in column B)
    for key, (header_row, start_row, end_row, _label) in FREE_TEXT_SECTIONS.items():
        text = data.get(key, "")
        if not text:
            continue
        lines = [ln for ln in text.replace("\r", "").split("\n") if ln.strip()]
        for i, line in enumerate(lines):
            ws.cell(row=start_row + i, column=1, value=line)  # A:F merged → write to A

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
